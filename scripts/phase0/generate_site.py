#!/usr/bin/env python3
"""Generate a static HTML site from the Phase-0 SQLite DB.

Output:
    site/
        index.html                       — leaderboard (sortable, men + ladies)
        styles.css                       — shared CSS
        players/<id>.html                — per-player page
        tournaments/<slug>.html          — roster ranking pages (one per entry
                                           in TOURNAMENT_ROSTERS)

Run from project root:
    python3 scripts/phase0/generate_site.py
"""

from __future__ import annotations

import difflib
import hashlib
import html
import json
import math
import os
import sqlite3
import sys
from pathlib import Path

# Anchor to project root regardless of cwd (script lives at scripts/phase0/).
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
DB_PATH = str(PROJECT_ROOT / "phase0.sqlite")
OUT_DIR = PROJECT_ROOT / "site"
MODEL = "openskill_pl"


def match_result(
    my_won: int, opp_won: int, my_games: int, opp_games: int
) -> tuple[str, str, str]:
    """Returns (cls, label, label_long) for one player's view of a match.

    Most matches resolve cleanly on sets: one side has won=1, the other won=0.
    But team-tournament rubbers split 1-1 on sets are stored with won=0 on
    BOTH sides; the rating engine breaks the tie via games-won (see
    rating.universal_score, PLAN.md §5.2). This helper mirrors that
    convention for the display layer so the UI stops calling the
    games-winner a "loss".

    Returns:
        cls: 'win' or 'loss' — preserves binary CSS classes (sort/streak
            logic stays untouched).
        label: 'W', 'L', 'W (g)', or 'L (g)' — the (g) suffix flags
            "won/lost on games tiebreak" so it's clear in the UI.
        label_long: 'Won', 'Lost', 'Won (games)', 'Lost (games)'.

    True 0/0 ties (equal sets AND equal games) are treated as 'L (g)' for
    both sides — extremely rare and the rating engine treats them as a
    draw anyway.
    """
    if my_won and not opp_won:
        return ("win", "W", "Won")
    if opp_won and not my_won:
        return ("loss", "L", "Lost")
    # Tied on sets — games-won tiebreak (matches the rating engine).
    my_g = my_games or 0
    opp_g = opp_games or 0
    if my_g > opp_g:
        return ("win", "W (g)", "Won (games)")
    return ("loss", "L (g)", "Lost (games)")

# Tournament roster pages. Each entry produces site/tournaments/<slug>.html
# from the named .xlsx (sheets `Men` and `Ladies`, two name-columns each).
# Add a new dict to publish another roster page.
TOURNAMENT_ROSTERS: list[dict] = [
    {
        "slug": "antes-tt-2026",
        "title": "VLTC next Tournament (Antes TT)",
        "menu_label": "Antes TT 2026 (roster)",
        "roster_xlsx": "_ANALYSIS_/NewTournamentRanking/Players List.xlsx",
        # Optional list of captain-supplied rankings. Each adds a sortable
        # column to the rated table. JSON is {"men": [...], "ladies": [...]}
        # where the name's index in the list = ordinal rank.
        "captain_rankings": [
            {
                "label": "Lonia",
                "json_path": "_ANALYSIS_/NewTournamentRanking/Captain-Lonia-Ranking/lonia_ranking.json",
            },
        ],
        "subtitle": (
            "Pre-tournament roster ranking. Players ordered by μ-3σ "
            "(conservative skill estimate). Class is the proposed slot for "
            "this tournament — 6 captains, so every 6 ranked players advance "
            "to the next slot (A1→A2→A3→A4→B1→…). Hover a class to see the "
            "player's previous-tournament class for reference. Click any "
            "column header to sort."
        ),
    },
]


def esc(s) -> str:
    return html.escape("" if s is None else str(s))


def write(path: Path, body: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(body, encoding="utf-8")


# --- CSS ---------------------------------------------------------------------

CSS = """
:root {
  --bg: #0f1115;
  --fg: #e6e6e6;
  --muted: #8b96a8;
  --accent: #4ea1ff;
  --win: #46c281;
  --loss: #e07a7a;
  --row-alt: #161a22;
  --card: #1a1f2a;
  --border: #2a3242;
}
* { box-sizing: border-box; }
html, body { -webkit-text-size-adjust: 100%; }
body {
  font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", system-ui, sans-serif;
  background: var(--bg); color: var(--fg);
  margin: 0; padding: 16px 24px; line-height: 1.45;
}
/* Layout used to be capped at 1200px which forced horizontal scroll on the
   wide match-log tables. Use the full viewport width (with body padding) so
   tables can lay out all columns naturally on desktop. Prose elements
   (h1/p) get their own readable max-width below. */
header { margin: 0 0 12px 0; }
header h1 { margin: 0 0 4px 0; font-size: 20px; max-width: 1100px; }
header p { margin: 0; color: var(--muted); font-size: 13px; max-width: 1100px; }
nav.topnav {
  margin: 0 0 12px 0;
  display: flex; gap: 6px; flex-wrap: wrap; align-items: center;
  padding: 8px 0; border-bottom: 1px solid var(--border);
}
nav.topnav a {
  color: var(--muted);
  padding: 6px 10px; border-radius: 6px;
  font-size: 13px; font-weight: 500;
}
nav.topnav a:hover { background: var(--card); color: var(--fg); text-decoration: none; }
nav.topnav a.active {
  background: var(--card); color: var(--fg);
  border: 1px solid var(--border);
}
nav.topnav .sep { color: var(--border); }
main { width: 100%; }
a { color: var(--accent); text-decoration: none; }
a:hover { text-decoration: underline; }
.controls { display: flex; gap: 8px; align-items: center; margin: 12px 0; flex-wrap: wrap; }
.controls input, .controls select {
  background: var(--card); color: var(--fg); border: 1px solid var(--border);
  padding: 8px 10px; border-radius: 6px; font-size: 14px;
  min-height: 38px;
}
.controls input { flex: 1 1 200px; min-width: 140px; }

/* Wrap every table in a div.table-wrap for horizontal scroll on mobile.
   Without this, wide tables blow out the viewport. */
.table-wrap {
  overflow-x: auto;
  -webkit-overflow-scrolling: touch;
  background: var(--card);
  border-radius: 8px;
  margin-bottom: 8px;
}
table {
  width: 100%;
  border-collapse: collapse;
  font-size: 13px;
  background: var(--card);
  /* No overflow:hidden here; the scroll container handles clipping.
     overflow:hidden also breaks position:sticky on <th>. */
}
th, td { padding: 6px 8px; text-align: left; border-bottom: 1px solid var(--border); white-space: nowrap; }
th { background: #20283a; color: var(--muted); font-weight: 600;
     position: sticky; top: 0; cursor: pointer; user-select: none; }
th:hover { color: var(--fg); }
tbody tr:nth-child(even) { background: var(--row-alt); }
tbody tr:hover { background: #21283a; }
td.num, th.num { text-align: right; font-variant-numeric: tabular-nums; }
.cls { font-weight: 600; color: var(--accent); }
.win { color: var(--win); }
.loss { color: var(--loss); }
.gender-M { color: #6cb3ff; }
.gender-F { color: #ff8db5; }
.player-link { color: var(--fg); }
.player-link:hover { color: var(--accent); }
.tag {
  display: inline-block; background: var(--bg); color: var(--muted);
  padding: 1px 6px; border-radius: 4px; font-size: 11px; margin-right: 4px;
}
.muted { color: var(--muted); }
footer { margin: 24px 0; color: var(--muted); font-size: 12px; max-width: 1100px; }

/* Player page */
.profile-head { display: flex; gap: 16px; align-items: flex-start; flex-wrap: wrap; margin-bottom: 12px; }
.profile-head .name { font-size: 26px; font-weight: 600; margin: 0 0 4px 0; line-height: 1.1; }
.profile-head .meta { color: var(--muted); font-size: 13px; }
.stat-grid {
  display: grid; grid-template-columns: repeat(auto-fit, minmax(100px, 1fr));
  gap: 6px; margin-bottom: 16px;
}
.stat {
  background: var(--card); border: 1px solid var(--border); padding: 10px 8px;
  border-radius: 6px; text-align: center;
}
.stat .label { color: var(--muted); font-size: 10px; text-transform: uppercase; letter-spacing: 0.05em; }
.stat .value { font-size: 18px; font-weight: 600; margin-top: 2px; font-variant-numeric: tabular-nums; }
.section-title { margin: 20px 0 8px 0; font-size: 14px; color: var(--muted);
                 text-transform: uppercase; letter-spacing: 0.05em; }
.chart {
  background: var(--card); border: 1px solid var(--border); border-radius: 8px;
  padding: 10px;
}
.chart svg { width: 100%; height: auto; display: block; }
.score { font-variant-numeric: tabular-nums; font-family: ui-monospace, "SF Mono", Menlo, monospace; }

/* Form badges (per-match W/L history) */
.form { display: inline-flex; gap: 3px; flex-wrap: wrap; }
.form .b {
  display: inline-block;
  width: 22px; height: 22px; line-height: 22px; text-align: center;
  border-radius: 4px; font-size: 11px; font-weight: 700;
  font-family: ui-monospace, monospace;
}
.form .b.w { background: rgba(70, 194, 129, 0.18); color: var(--win); }
.form .b.l { background: rgba(224, 122, 122, 0.18); color: var(--loss); }

/* Mobile tweaks */
@media (max-width: 700px) {
  body { padding: 10px; }
  header h1 { font-size: 18px; }
  .profile-head .name { font-size: 22px; }
  th, td { padding: 6px 8px; font-size: 12px; }
  .stat .value { font-size: 16px; }
  /* Hide low-value columns on the leaderboard for narrow screens */
  table.leaderboard th:nth-child(5),     /* μ */
  table.leaderboard td:nth-child(5),
  table.leaderboard th:nth-child(6),     /* σ */
  table.leaderboard td:nth-child(6),
  table.leaderboard th:nth-child(12),    /* games */
  table.leaderboard td:nth-child(12),
  table.leaderboard th:nth-child(13),    /* g% */
  table.leaderboard td:nth-child(13) { display: none; }
}

/* --- Changelog page --- */
.changelog { max-width: 760px; }
.changelog .filters {
  display: flex; flex-wrap: wrap; gap: 6px;
  margin: 14px 0 18px 0;
}
.changelog .filter {
  background: var(--card); color: var(--muted);
  border: 1px solid var(--border); border-radius: 999px;
  padding: 4px 12px; font-size: 12px; cursor: pointer;
  user-select: none;
}
.changelog .filter:hover { color: var(--fg); }
.changelog .filter.active {
  background: var(--accent); color: #0b1220;
  border-color: var(--accent);
}
.changelog .month-heading {
  margin: 22px 0 10px 0;
  font-size: 13px; color: var(--muted);
  text-transform: uppercase; letter-spacing: 0.06em;
  border-bottom: 1px solid var(--border);
  padding-bottom: 4px;
}
.changelog .entry {
  background: var(--card); border: 1px solid var(--border);
  border-radius: 8px; padding: 14px 16px; margin: 0 0 12px 0;
}
.changelog .entry.hidden { display: none; }
.changelog .entry-head {
  display: flex; align-items: center; gap: 10px; flex-wrap: wrap;
  margin-bottom: 4px;
}
.changelog .pill {
  display: inline-block; font-size: 11px; font-weight: 600;
  padding: 2px 9px; border-radius: 999px; letter-spacing: 0.04em;
  text-transform: uppercase;
}
.changelog .pill.kind-new      { background: #173626; color: #6fe09e; }
.changelog .pill.kind-improved { background: #1d2f4a; color: #6fb1ff; }
.changelog .pill.kind-fixed    { background: #3a2820; color: #ffae8a; }
.changelog .pill.aud-captains  { background: #2b2540; color: #c2b1ff; }
.changelog .pill.aud-admins    { background: #2a2a2a; color: #cccccc; }
.changelog .entry-date { color: var(--muted); font-size: 12px; }
.changelog .entry-title {
  font-size: 16px; font-weight: 600; margin: 0 0 4px 0;
}
.changelog .entry-summary { color: var(--fg); font-size: 14px; margin: 0; }
.changelog details {
  margin-top: 8px; color: var(--muted); font-size: 13px;
}
.changelog details summary {
  cursor: pointer; color: var(--accent); font-size: 12px;
  list-style: none;
}
.changelog details summary::-webkit-details-marker { display: none; }
.changelog details summary:hover { text-decoration: underline; }
.changelog details > p { margin: 8px 0 0 0; color: var(--fg); }
.changelog .entry-trace {
  margin-top: 10px; font-family: ui-monospace, SFMono-Regular, Menlo, monospace;
  font-size: 11px; color: var(--muted); word-break: break-all;
}
.changelog .entry-trace a { color: var(--muted); }
.changelog .entry-trace a:hover { color: var(--fg); }
.changelog .lead {
  color: var(--muted); font-size: 14px; max-width: 700px;
  margin: 0 0 6px 0;
}
.changelog .empty {
  color: var(--muted); padding: 24px; text-align: center;
  border: 1px dashed var(--border); border-radius: 8px;
}

/* --- How it works page --- */
.hwr { max-width: 760px; }
.hwr .lead { color: var(--muted); font-size: 14px; margin: 0 0 18px 0; }
.hwr section { margin: 26px 0; }
.hwr h2 {
  font-size: 17px; margin: 0 0 8px 0;
  border-bottom: 1px solid var(--border); padding-bottom: 4px;
}
.hwr p { font-size: 14px; line-height: 1.55; margin: 8px 0; }
.hwr ul { font-size: 14px; line-height: 1.55; margin: 8px 0; padding-left: 20px; }
.hwr li { margin: 4px 0; }
.hwr code, .hwr .code {
  font-family: ui-monospace, SFMono-Regular, Menlo, monospace;
  font-size: 13px; background: var(--card); padding: 1px 5px; border-radius: 3px;
  border: 1px solid var(--border);
}
.hwr .callout {
  background: var(--card); border: 1px solid var(--border);
  border-left: 3px solid var(--accent); border-radius: 6px;
  padding: 10px 14px; margin: 14px 0; font-size: 13.5px;
}
.hwr .callout strong { color: var(--accent); }
.hwr .diagram {
  background: var(--card); border: 1px solid var(--border);
  border-radius: 8px; padding: 16px; margin: 14px 0;
  text-align: center;
}
.hwr .diagram svg { max-width: 100%; height: auto; }
.hwr .diagram-caption {
  color: var(--muted); font-size: 12px; margin-top: 8px;
  font-style: italic;
}
/* --- Calculator --- */
.hwr .calc {
  background: var(--card); border: 1px solid var(--border);
  border-radius: 8px; padding: 18px; margin: 14px 0;
}
.hwr .calc-pair {
  display: grid; grid-template-columns: 90px 1fr 1fr; gap: 8px 12px;
  align-items: center; margin: 8px 0;
}
.hwr .calc-pair .pair-label {
  font-size: 12px; color: var(--muted); text-transform: uppercase;
  letter-spacing: 0.06em;
}
.hwr .calc input {
  background: var(--bg); color: var(--fg); border: 1px solid var(--border);
  border-radius: 4px; padding: 6px 10px; font-size: 13px; width: 100%;
}
.hwr .calc input:focus { outline: none; border-color: var(--accent); }
.hwr .calc input.invalid { border-color: var(--loss); }
.hwr .calc .calc-vs {
  text-align: center; color: var(--muted); font-size: 11px;
  text-transform: uppercase; letter-spacing: 0.08em; margin: 6px 0;
}
.hwr .calc-out {
  margin-top: 14px; padding: 14px;
  background: var(--bg); border: 1px solid var(--border); border-radius: 6px;
  font-size: 13.5px; line-height: 1.55;
}
.hwr .calc-out.placeholder { color: var(--muted); font-style: italic; }
.hwr .calc-bar {
  display: flex; height: 22px; border-radius: 4px; overflow: hidden;
  border: 1px solid var(--border); margin: 8px 0;
  font-size: 11px; font-weight: 600; color: #0b1220;
}
.hwr .calc-bar .a { background: var(--win); display: flex; align-items: center; justify-content: center; }
.hwr .calc-bar .b { background: var(--loss); display: flex; align-items: center; justify-content: center; }
.hwr .calc-table {
  display: grid; grid-template-columns: 1.4fr 0.6fr 0.6fr 0.7fr 0.9fr 0.9fr;
  gap: 4px 12px;
  font-family: ui-monospace, SFMono-Regular, Menlo, monospace;
  font-size: 12px; margin-top: 14px;
}
.hwr .calc-table .calc-th {
  color: var(--muted); text-transform: uppercase; font-size: 10.5px;
  letter-spacing: 0.06em; border-bottom: 1px solid var(--border);
  padding-bottom: 3px;
}
.hwr .calc-table .num { text-align: right; }
.hwr .calc-table .delta-pos { color: var(--win); }
.hwr .calc-table .delta-neg { color: var(--loss); }
.hwr .calc-table .delta-zero { color: var(--muted); }
.hwr .calc-section-head {
  margin-top: 14px; font-size: 12px; color: var(--muted);
  text-transform: uppercase; letter-spacing: 0.06em;
}
.hwr .calc-toggle {
  display: inline-flex; gap: 4px; margin: 10px 0 4px 0;
  padding: 3px; background: var(--bg); border: 1px solid var(--border);
  border-radius: 6px;
}
.hwr .calc-toggle button {
  background: transparent; color: var(--muted);
  border: 0; padding: 5px 12px; font-size: 12px; cursor: pointer;
  border-radius: 4px; font-weight: 600;
}
.hwr .calc-toggle button.active { background: var(--accent); color: #0b1220; }
@media (max-width: 600px) {
  .hwr .calc-pair { grid-template-columns: 1fr; }
  .hwr .calc-pair .pair-label { margin-top: 8px; }
  .hwr .calc-table {
    grid-template-columns: 1.6fr 0.7fr 0.9fr 0.9fr;
  }
  .hwr .calc-table .hide-mobile { display: none; }
}

/* --- Multiplier-stack diagram --- */
.hwr .stack-diagram {
  font-family: ui-monospace, SFMono-Regular, Menlo, monospace;
  font-size: 12px;
  background: var(--bg); border: 1px solid var(--border);
  border-radius: 6px; padding: 14px; margin: 8px 0;
}
.hwr .stack-row {
  display: grid; grid-template-columns: 150px 1fr 70px;
  align-items: center; gap: 12px; margin: 6px 0;
}
.hwr .stack-label {
  color: var(--muted); font-size: 11px; text-transform: uppercase;
  letter-spacing: 0.05em;
}
.hwr .stack-label .stack-mult {
  display: block; color: var(--fg); margin-top: 2px;
  font-size: 12px; text-transform: none; letter-spacing: 0;
  font-weight: 600;
}
.hwr .stack-bar-track {
  height: 16px; background: var(--card); border: 1px solid var(--border);
  border-radius: 3px; position: relative; overflow: hidden;
}
.hwr .stack-bar-fill {
  height: 100%; background: var(--accent); opacity: 0.7;
}
.hwr .stack-bar-fill.baseline { background: var(--muted); opacity: 0.5; }
.hwr .stack-bar-fill.up { background: var(--win); }
.hwr .stack-bar-fill.down { background: var(--loss); }
.hwr .stack-value {
  text-align: right; font-weight: 600;
  font-variant-numeric: tabular-nums;
}
.hwr .stack-final-row {
  border-top: 1px solid var(--border); margin-top: 10px; padding-top: 10px;
}
.hwr .stack-final-row .stack-label { color: var(--accent); font-weight: 600; }

/* --- Match-impact expansion (All-matches + Player pages) --- */
/* Tiny rank badge shown next to every player name in match listings,
   reflecting that player's rank in their gender bucket AT THE TIME of
   the match (not their current rank). */
.rank-tag {
  display: inline-block; background: rgba(78, 161, 255, 0.12);
  color: var(--accent); padding: 0 5px; border-radius: 3px;
  font-size: 10.5px; font-weight: 600; margin-left: 4px;
  font-variant-numeric: tabular-nums;
  vertical-align: 1px;
}
.rank-tag.muted { background: rgba(139, 150, 168, 0.15); color: var(--muted); }
/* Expander cell — clickable arrow that toggles the impact row. */
.expand-trigger {
  cursor: pointer; user-select: none;
  width: 22px; text-align: center;
  color: var(--muted); font-size: 11px;
  transition: transform 0.15s ease;
}
.expand-trigger:hover { color: var(--accent); }
tr.open .expand-trigger { transform: rotate(90deg); color: var(--accent); }
tr.impact-row { background: #11151d !important; }
tr.impact-row > td {
  padding: 0; border-bottom: 1px solid var(--border);
  white-space: normal;          /* override the global td nowrap */
}
tr.impact-row[hidden] { display: none; }
/* Match rows: allow long player names to wrap so the row grows vertically
   instead of overflowing horizontally. */
tr.match-row td { white-space: normal; }
tr.match-row td.score, tr.match-row td.num { white-space: nowrap; }

/* Impact box layout: 2-vs-2 visualisation.
   Desktop: [Side A pair] | VS | [Side B pair] in three columns.
   Mobile:  Side A pair stacked, then VS, then Side B pair stacked.
   Inside each side, the two partners stack vertically so the doubles pair
   reads as a unit. */
.impact-box {
  padding: 10px 14px 12px 32px;
  display: grid; gap: 8px;
  grid-template-columns: 1fr auto 1fr;
  grid-template-areas: "a vs b";
  align-items: stretch;
}
.impact-side {
  display: flex; flex-direction: column; gap: 6px;
}
.impact-side.side-A { grid-area: a; }
.impact-side.side-B { grid-area: b; }
.impact-vs {
  grid-area: vs;
  display: flex; align-items: center; justify-content: center;
  font-weight: 700; color: var(--muted);
  font-size: 13px; letter-spacing: 0.12em;
  padding: 0 6px;
  position: relative;
}
.impact-vs::before, .impact-vs::after {
  content: ""; position: absolute; left: 50%; width: 1px;
  background: var(--border); transform: translateX(-50%);
}
.impact-vs::before { top: 0; height: calc(50% - 14px); }
.impact-vs::after  { bottom: 0; height: calc(50% - 14px); }
.impact-player {
  background: var(--card); border: 1px solid var(--border);
  border-radius: 6px; padding: 8px 10px;
  font-size: 12px;
  display: flex; flex-direction: column; gap: 3px;
  word-break: break-word;       /* never overflow horizontally */
  min-width: 0;                  /* flex/grid child can shrink */
}
.impact-player .who {
  font-weight: 600; font-size: 13px;
  display: flex; justify-content: space-between; align-items: center; gap: 8px;
  flex-wrap: wrap;
}
.impact-player .side-tag {
  font-size: 9.5px; font-weight: 700;
  padding: 1px 5px; border-radius: 3px; letter-spacing: 0.04em;
  white-space: nowrap;
}
.impact-player .side-tag.A { background: rgba(78, 161, 255, 0.18); color: #6cb3ff; }
.impact-player .side-tag.B { background: rgba(170, 110, 220, 0.18); color: #c39cff; }
.impact-player .side-tag.win { background: rgba(70, 194, 129, 0.22); color: var(--win); }
.impact-player .side-tag.loss { background: rgba(224, 122, 122, 0.22); color: var(--loss); }
.impact-player .metric {
  display: flex; gap: 6px; font-variant-numeric: tabular-nums;
  font-family: ui-monospace, "SF Mono", Menlo, monospace;
  font-size: 11.5px;
  flex-wrap: wrap;
}
.impact-player .metric .k { color: var(--muted); width: 50px; flex-shrink: 0; }
.impact-player .metric .delta-up { color: var(--win); font-weight: 600; }
.impact-player .metric .delta-dn { color: var(--loss); font-weight: 600; }
.impact-player .metric .delta-z { color: var(--muted); }
.impact-player .commentary { color: var(--muted); font-size: 11px; line-height: 1.4; }
.impact-player .commentary .pass-up { color: var(--win); }
.impact-player .commentary .pass-dn { color: var(--loss); }
.impact-player .new-entry { color: var(--accent); font-style: italic; font-size: 11px; }
@media (max-width: 700px) {
  .impact-box {
    padding: 8px 10px 10px 18px;
    grid-template-columns: 1fr;
    grid-template-areas: "a" "vs" "b";
  }
  .impact-vs { padding: 6px 0; font-size: 12px; }
  .impact-vs::before, .impact-vs::after {
    width: calc(100% - 24px); height: 1px; left: 0; transform: none;
  }
  .impact-vs::before { top: 50%; }
  .impact-vs::after { bottom: 50%; left: auto; right: 0; }
  .rank-tag { font-size: 10px; padding: 0 4px; }
}
"""

# Cache-busting fingerprint for styles.css. GH Pages caches assets for 10 min;
# without a query-string version, browsers may serve the old CSS even after a
# fresh deploy. Recomputed automatically every time the CSS changes.
CSS_VERSION = hashlib.sha1(CSS.encode("utf-8")).hexdigest()[:10]


# --- Helpers -----------------------------------------------------------------


def render_nav(rel_root: str, active: str) -> str:
    """Top-of-page navigation bar.

    rel_root  -- prefix for hrefs ('' from index, '../' from sub-pages).
    active    -- which entry to highlight: 'index', 'matches', 'changelog',
                 'how-it-works', 'aliases', or 'tournament:<slug>'.
    """
    parts: list[str] = []
    cls_index = ' class="active"' if active == "index" else ""
    parts.append(f'<a href="{rel_root}index.html"{cls_index}>Leaderboard</a>')
    parts.append('<span class="sep">·</span>')
    cls_matches = ' class="active"' if active == "matches" else ""
    parts.append(f'<a href="{rel_root}matches.html"{cls_matches}>All matches</a>')
    parts.append('<span class="sep">·</span>')
    cls_dis = ' class="active"' if active == "disagreements" else ""
    parts.append(
        f'<a href="{rel_root}disagreements.html"{cls_dis} '
        f'title="Matches where vanilla PL and Decay-365 models predicted '
        f'different outcomes — where captain knowledge matters most.">'
        f'Model gaps</a>'
    )
    for t in TOURNAMENT_ROSTERS:
        slug = t["slug"]
        label = t["menu_label"]
        cls = ' class="active"' if active == f"tournament:{slug}" else ""
        parts.append('<span class="sep">·</span>')
        parts.append(f'<a href="{rel_root}tournaments/{esc(slug)}.html"{cls}>{esc(label)}</a>')
    parts.append('<span class="sep">·</span>')
    cls_al = ' class="active"' if active == "aliases" else ""
    parts.append(f'<a href="{rel_root}aliases.html"{cls_al}>Mapping</a>')
    parts.append('<span class="sep">·</span>')
    cls_cl = ' class="active"' if active == "changelog" else ""
    parts.append(f'<a href="{rel_root}changelog.html"{cls_cl}>What’s new</a>')
    parts.append('<span class="sep">·</span>')
    cls_hw = ' class="active"' if active == "how-it-works" else ""
    parts.append(f'<a href="{rel_root}how-it-works.html"{cls_hw}>How it works</a>')
    return f'<nav class="topnav">{"".join(parts)}</nav>'


def player_filename(pid: int) -> str:
    return f"players/{pid}.html"


def player_link(pid: int, name: str) -> str:
    return f'<a class="player-link" href="players/{pid}.html">{esc(name)}</a>'


def player_link_from_player_page(pid: int, name: str) -> str:
    return f'<a class="player-link" href="{pid}.html">{esc(name)}</a>'


def fetch_neighbour_index(
    conn: sqlite3.Connection,
) -> dict[str, list[dict]]:
    """Return ranked player lists per gender bucket.

    Buckets:
        'M' — men only (μ-3σ desc)
        'F' — ladies only
        'all' — fallback for players with NULL gender

    Each entry: {pid, name, gender, mu, sigma, mu3s, n, wins, captain_class}.
    Order is a stable rank-1 → rank-N within the bucket.
    """
    rows = conn.execute(
        """
        SELECT
            p.id, p.canonical_name, p.gender,
            r.mu, r.sigma, (r.mu - 3*r.sigma) AS mu3s,
            (SELECT COUNT(*) FROM match_sides ms
             JOIN matches m ON m.id = ms.match_id
             WHERE (ms.player1_id = p.id OR ms.player2_id = p.id)
               AND m.superseded_by_run_id IS NULL) AS n,
            (SELECT COUNT(*) FROM match_sides ms
             JOIN matches m ON m.id = ms.match_id
             WHERE (ms.player1_id = p.id OR ms.player2_id = p.id)
               AND m.superseded_by_run_id IS NULL AND ms.won = 1) AS wins,
            (SELECT pta.class_label
             FROM player_team_assignments pta
             JOIN tournaments t ON t.id = pta.tournament_id
             WHERE pta.player_id = p.id
             ORDER BY t.year DESC, t.id DESC LIMIT 1) AS captain_class
        FROM ratings r
        JOIN players p ON p.id = r.player_id
        WHERE r.model_name = ? AND p.merged_into_id IS NULL
        ORDER BY mu3s DESC
        """,
        (MODEL,),
    ).fetchall()

    buckets: dict[str, list[dict]] = {"M": [], "F": [], "all": []}
    for (pid, name, gender, mu, sigma, mu3s, n, wins, captain_class) in rows:
        if (n or 0) == 0:
            continue
        entry = {
            "pid": pid, "name": name, "gender": gender,
            "mu": mu, "sigma": sigma, "mu3s": mu3s,
            "n": n, "wins": wins, "captain_class": captain_class or "",
        }
        buckets["all"].append(entry)
        if gender in ("M", "F"):
            buckets[gender].append(entry)
    return buckets


def render_neighbours(neighbours: list[dict], me_pid: int) -> str:
    """Render a small "compare with peers" table: 3 above + me + 3 below."""
    pos = next(
        (i for i, e in enumerate(neighbours) if e["pid"] == me_pid), None
    )
    if pos is None:
        return ""
    lo = max(0, pos - 3)
    hi = min(len(neighbours), pos + 4)  # exclusive
    slice_ = neighbours[lo:hi]

    rows = []
    for i, e in enumerate(slice_, start=lo):
        rank = i + 1
        is_me = e["pid"] == me_pid
        n = e["n"]
        wins = e["wins"]
        losses = n - wins
        win_pct = f"{int(round(wins * 100 / n))}%" if n else "—"
        name_cell = (
            f'<strong>{esc(e["name"])}</strong>'
            if is_me
            else f'<a class="player-link" href="{e["pid"]}.html">{esc(e["name"])}</a>'
        )
        diff_cell = ""
        if not is_me:
            me_entry = neighbours[pos]
            d = e["mu3s"] - me_entry["mu3s"]
            cls = "win" if d < 0 else ("loss" if d > 0 else "muted")
            # d > 0 means peer is above me (better) — show as red gap from my POV
            sign = "+" if d >= 0 else ""
            diff_cell = f'<span class="{cls}">{sign}{d:.2f}</span>'
        else:
            diff_cell = '<span class="muted">—</span>'
        rows.append(
            f'<tr{" style=\"background:#22304a\"" if is_me else ""}>'
            f'<td class="num">{rank}</td>'
            f'<td class="cls">{esc(e["captain_class"])}</td>'
            f'<td>{name_cell}</td>'
            f'<td class="num">{e["mu"]:.2f}</td>'
            f'<td class="num">{e["sigma"]:.2f}</td>'
            f'<td class="num"><strong>{e["mu3s"]:.2f}</strong></td>'
            f'<td class="num">{diff_cell}</td>'
            f'<td class="num">{n}</td>'
            f'<td class="num"><span class="win">{wins}</span>-<span class="loss">{losses}</span></td>'
            f'<td class="num">{win_pct}</td>'
            f'</tr>'
        )
    return f"""
  <h2 class="section-title">Peers (3 above &middot; 3 below)</h2>
  <div class="table-wrap">
  <table>
    <thead><tr>
      <th class="num">#</th><th>Class</th><th>Player</th>
      <th class="num">μ</th><th class="num">σ</th><th class="num">μ-3σ</th>
      <th class="num">Δ vs me</th>
      <th class="num">n</th><th class="num">W-L</th><th class="num">win%</th>
    </tr></thead>
    <tbody>{''.join(rows)}</tbody>
  </table>
  </div>
"""


def fetch_player_lookup(conn: sqlite3.Connection) -> dict[int, tuple[int, str]]:
    """Map every player_id (including merged-out ones) to (canonical_id, name).

    Merged-out IDs resolve to their merge target's canonical name + ID, so a
    historical match referencing the old ID still shows the right name and
    links to the surviving page.
    """
    rows = conn.execute(
        "SELECT id, canonical_name, merged_into_id FROM players"
    ).fetchall()
    by_id = {r[0]: r for r in rows}
    out: dict[int, tuple[int, str]] = {}
    for pid, name, merged_into in rows:
        canon_id = pid
        # Walk merge chain (defensive against multi-hop merges).
        seen = {pid}
        while merged_into is not None and merged_into not in seen:
            seen.add(merged_into)
            target = by_id.get(merged_into)
            if target is None:
                break
            canon_id = target[0]
            merged_into = target[2]
        canon_name = by_id[canon_id][1]
        out[pid] = (canon_id, canon_name)
    return out


# --- Index page --------------------------------------------------------------


LEADERBOARD_SQL = """
SELECT
    p.id, p.canonical_name, p.gender,
    r.mu, r.sigma, (r.mu - 3*r.sigma) AS mu3s,
    (
        SELECT COUNT(*) FROM match_sides ms
        JOIN matches m ON m.id = ms.match_id
        WHERE (ms.player1_id = p.id OR ms.player2_id = p.id)
          AND m.superseded_by_run_id IS NULL
    ) AS n,
    (
        SELECT COUNT(*) FROM match_sides ms
        JOIN matches m ON m.id = ms.match_id
        WHERE (ms.player1_id = p.id OR ms.player2_id = p.id)
          AND m.superseded_by_run_id IS NULL AND ms.won = 1
    ) AS wins,
    (
        SELECT COALESCE(SUM(ms.games_won), 0) FROM match_sides ms
        JOIN matches m ON m.id = ms.match_id
        WHERE (ms.player1_id = p.id OR ms.player2_id = p.id)
          AND m.superseded_by_run_id IS NULL
    ) AS gw,
    (
        SELECT COALESCE(SUM(opp.games_won), 0) FROM match_sides ms
        JOIN matches m ON m.id = ms.match_id
        JOIN match_sides opp ON opp.match_id = ms.match_id AND opp.side <> ms.side
        WHERE (ms.player1_id = p.id OR ms.player2_id = p.id)
          AND m.superseded_by_run_id IS NULL
    ) AS gl,
    (
        SELECT MAX(m.played_on) FROM match_sides ms
        JOIN matches m ON m.id = ms.match_id
        WHERE (ms.player1_id = p.id OR ms.player2_id = p.id)
          AND m.superseded_by_run_id IS NULL
    ) AS last_played,
    (
        SELECT pta.class_label
        FROM player_team_assignments pta
        JOIN tournaments t ON t.id = pta.tournament_id
        WHERE pta.player_id = p.id
        ORDER BY t.year DESC, t.id DESC LIMIT 1
    ) AS captain_class,
    (
        SELECT GROUP_CONCAT(DISTINCT c.slug)
        FROM match_sides ms
        JOIN matches m ON m.id = ms.match_id
        JOIN tournaments t ON t.id = m.tournament_id
        JOIN clubs c ON c.id = t.club_id
        WHERE (ms.player1_id = p.id OR ms.player2_id = p.id)
          AND m.superseded_by_run_id IS NULL
    ) AS clubs
FROM ratings r
JOIN players p ON p.id = r.player_id
WHERE r.model_name = ?
  AND p.merged_into_id IS NULL
ORDER BY mu3s DESC
"""


def build_index(conn: sqlite3.Connection) -> str:
    rows = conn.execute(LEADERBOARD_SQL, (MODEL,)).fetchall()
    rows = [r for r in rows if r[6] > 0]  # n_matches > 0

    # Decay-365 challenger rank per gender. Sort each gender's pool by
    # decay μ-3σ DESC and map player_id → rank within that gender. Players
    # with NULL gender or missing decay row get None.
    decay_ratings = {
        row[0]: (row[1], row[2])
        for row in conn.execute(
            "SELECT player_id, mu, sigma FROM ratings "
            "WHERE model_name = 'openskill_pl_decay365'"
        ).fetchall()
    }
    decay_rank: dict[int, int] = {}
    for gender_filter in ("M", "F"):
        # Use the leaderboard output (already filtered to n>0) so the decay
        # rank reflects the same population as the displayed rows.
        candidates = [
            (pid, decay_ratings[pid])
            for (pid, _, gender, *_rest) in rows
            if gender == gender_filter and pid in decay_ratings
        ]
        candidates.sort(key=lambda x: -(x[1][0] - 3 * x[1][1]))
        for i, (pid, _) in enumerate(candidates, 1):
            decay_rank[pid] = i

    body_rows = []
    for rank, r in enumerate(rows, 1):
        (pid, name, gender, mu, sigma, mu3s, n, wins, gw, gl,
         last_played, captain_class, clubs) = r
        losses = n - wins
        win_pct = f"{int(round(wins * 100 / n))}%" if n else ""
        gw_pct = f"{int(round(gw * 100 / (gw + gl)))}%" if (gw + gl) else ""
        cls = captain_class or ""
        clubs_str = clubs or ""
        d_rank = decay_rank.get(pid)
        if d_rank is not None:
            decay_cell = (
                f'<td class="num" data-sort="{d_rank}">'
                f'{d_rank}</td>'
            )
        else:
            decay_cell = (
                '<td class="num muted" data-sort="999999">—</td>'
            )
        body_rows.append(
            f'<tr data-gender="{esc(gender or "")}" data-class="{esc(cls)}" '
            f'data-clubs="{esc(clubs_str)}">'
            f'<td class="num">{rank}</td>'
            f'<td class="cls">{esc(cls)}</td>'
            f'<td>{player_link(pid, name)}</td>'
            f'<td class="gender-{esc(gender or "")}">{esc(gender or "")}</td>'
            f'<td class="num">{mu:.2f}</td>'
            f'<td class="num">{sigma:.2f}</td>'
            f'<td class="num"><strong>{mu3s:.2f}</strong></td>'
            f'{decay_cell}'
            f'<td class="num">{n}</td>'
            f'<td class="num"><span class="win">{wins}</span>-<span class="loss">{losses}</span></td>'
            f'<td class="num">{win_pct}</td>'
            f'<td class="num">{gw}-{gl}</td>'
            f'<td class="num">{gw_pct}</td>'
            f'<td class="muted">{esc(clubs_str)}</td>'
            f'<td class="muted">{esc(last_played or "")}</td>'
            f'</tr>'
        )

    js = """
    <script>
    function applyFilters() {
      const g = document.getElementById('f-gender').value;
      const q = document.getElementById('f-search').value.toLowerCase();
      const cl = document.getElementById('f-club').value;
      const rows = document.querySelectorAll('tbody tr');
      let visible = 0;
      rows.forEach((row) => {
        const rg = row.dataset.gender;
        const rclubs = row.dataset.clubs;
        const txt = row.textContent.toLowerCase();
        const okG = !g || rg === g;
        const okC = !cl || (rclubs && rclubs.split(',').includes(cl));
        const okQ = !q || txt.includes(q);
        const show = okG && okC && okQ;
        row.style.display = show ? '' : 'none';
        if (show) visible++;
      });
      document.getElementById('count').textContent = visible + ' players';
    }
    document.addEventListener('DOMContentLoaded', () => {
      ['f-gender','f-search','f-club'].forEach((id) => {
        document.getElementById(id).addEventListener('input', applyFilters);
      });
      // Sort on header click
      document.querySelectorAll('th').forEach((th, idx) => {
        th.addEventListener('click', () => {
          const tbody = document.querySelector('tbody');
          const rows = Array.from(tbody.querySelectorAll('tr'));
          const asc = th.dataset.sort !== 'asc';
          th.dataset.sort = asc ? 'asc' : 'desc';
          const numeric = th.classList.contains('num');
          rows.sort((a,b) => {
            const av = a.cells[idx].textContent.trim();
            const bv = b.cells[idx].textContent.trim();
            if (numeric) {
              const an = parseFloat(av.replace(/[^0-9.\\-]/g,'')) || 0;
              const bn = parseFloat(bv.replace(/[^0-9.\\-]/g,'')) || 0;
              return asc ? an - bn : bn - an;
            }
            return asc ? av.localeCompare(bv) : bv.localeCompare(av);
          });
          rows.forEach(r => tbody.appendChild(r));
        });
      });
      applyFilters();
    });
    </script>
    """

    # Collect club slugs for the filter
    club_slugs = sorted({
        s for r in rows if r[12]
        for s in (r[12] or "").split(",") if s
    })
    club_options = "".join(
        f'<option value="{esc(c)}">{esc(c)}</option>' for c in club_slugs
    )

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1, viewport-fit=cover">
<meta name="theme-color" content="#0f1115">
<title>RallyRank — Leaderboard</title>
<link rel="stylesheet" href="styles.css?v={CSS_VERSION}">
</head>
<body>
<header>
  <h1>RallyRank — Doubles Leaderboard</h1>
  <p>OpenSkill Plackett-Luce ratings &middot; class column = captain-assigned slot from most recent team tournament</p>
</header>
{render_nav("", "index")}
<main>
  <div class="controls">
    <input id="f-search" type="search" placeholder="Search player...">
    <select id="f-gender">
      <option value="">All genders</option>
      <option value="M">Men</option>
      <option value="F">Ladies</option>
    </select>
    <select id="f-club">
      <option value="">All clubs</option>
      {club_options}
    </select>
    <span id="count" class="muted"></span>
  </div>
  <div class="table-wrap">
  <table class="leaderboard">
    <thead>
      <tr>
        <th class="num">#</th>
        <th>Class</th>
        <th>Player</th>
        <th>G</th>
        <th class="num">μ</th>
        <th class="num">σ</th>
        <th class="num">μ-3σ</th>
        <th class="num" title="Time-decay challenger rank (τ=365d, within same gender). Recency-weighted: old matches contribute exponentially less. See _ANALYSIS_/model_evaluation/SUMMARY.md for backtest results.">Decay #</th>
        <th class="num">n</th>
        <th class="num">W-L</th>
        <th class="num">win%</th>
        <th class="num">games</th>
        <th class="num">g%</th>
        <th>Clubs</th>
        <th>Last</th>
      </tr>
    </thead>
    <tbody>
      {''.join(body_rows)}
    </tbody>
  </table>
  </div>
</main>
<footer>
  Generated from {DB_PATH} &middot; click any player for trajectory and match history &middot;
  click any column header to sort.
</footer>
{js}
</body>
</html>
"""


# --- Player page -------------------------------------------------------------


PLAYER_INFO_SQL = """
SELECT id, canonical_name, gender FROM players WHERE id = ?
"""

PLAYER_RATING_SQL = """
SELECT mu, sigma FROM ratings WHERE player_id = ? AND model_name = ?
"""

PLAYER_MATCHES_SQL = """
SELECT
    m.id, m.played_on, m.division, m.round, m.walkover,
    t.id AS tour_id, t.name AS tour_name, t.year, c.name AS club_name, c.slug AS club_slug,
    ms.side, ms.player1_id AS my_p1, ms.player2_id AS my_p2,
    ms.games_won AS my_games, ms.sets_won AS my_sets, ms.won AS my_won,
    opp.player1_id AS opp_p1, opp.player2_id AS opp_p2,
    opp.games_won AS opp_games, opp.sets_won AS opp_sets,
    rh.mu_after, rh.sigma_after,
    opp.won AS opp_won
FROM match_sides ms
JOIN matches m ON m.id = ms.match_id
JOIN match_sides opp ON opp.match_id = m.id AND opp.side <> ms.side
JOIN tournaments t ON t.id = m.tournament_id
JOIN clubs c ON c.id = t.club_id
LEFT JOIN rating_history rh
    ON rh.match_id = m.id AND rh.player_id = ? AND rh.model_name = ?
WHERE (ms.player1_id = ? OR ms.player2_id = ?)
  AND m.superseded_by_run_id IS NULL
ORDER BY m.played_on, m.id
"""

PLAYER_CLASS_HISTORY_SQL = """
SELECT t.year, t.name, c.name AS club_name, c.slug AS club_slug,
       pta.team_letter, pta.captain_name, pta.class_label
FROM player_team_assignments pta
JOIN tournaments t ON t.id = pta.tournament_id
JOIN clubs c ON c.id = t.club_id
WHERE pta.player_id = ?
ORDER BY t.year, t.id
"""

SET_SCORES_SQL = """
SELECT set_number, side_a_games, side_b_games, was_tiebreak
FROM match_set_scores WHERE match_id = ? ORDER BY set_number
"""

# All raw-name aliases ever recorded for this player (post-merge: includes
# aliases that were transferred from absorbed losers).
PLAYER_ALIASES_SQL = """
SELECT pa.raw_name, pa.first_seen_at, sf.original_filename
FROM player_aliases pa
LEFT JOIN source_files sf ON sf.id = pa.source_file_id
WHERE pa.player_id = ?
ORDER BY pa.first_seen_at, pa.raw_name
"""

# Merge events where THIS player was the winner (others got absorbed).
# Reads the audit_log entries written by `merge_player_into`.
PLAYER_MERGES_IN_SQL = """
SELECT al.entity_id      AS loser_id,
       al.ts             AS merged_at,
       al.before_jsonb   AS before_json,
       al.after_jsonb    AS after_json
FROM audit_log al
WHERE al.action = 'player.merged'
  AND al.entity_type = 'players'
  AND json_extract(al.after_jsonb, '$.merged_into_id') = ?
ORDER BY al.ts
"""


def _resolve(name_lookup, pid):
    return name_lookup.get(pid, (pid, f"#{pid}"))


def render_partner(my_p1, my_p2, me_id, name_lookup) -> str:
    partner_id = my_p2 if my_p1 == me_id else my_p1
    if partner_id is None:
        return '<span class="muted">—</span>'
    cid, name = _resolve(name_lookup, partner_id)
    return player_link_from_player_page(cid, name)


def render_opponents(opp_p1, opp_p2, name_lookup) -> str:
    parts = []
    for pid in (opp_p1, opp_p2):
        if pid:
            cid, name = _resolve(name_lookup, pid)
            parts.append(player_link_from_player_page(cid, name))
    return " / ".join(parts) if parts else '<span class="muted">—</span>'


def render_score(side: str, set_scores: list) -> str:
    """Render set scores from THIS player's perspective (side A or B)."""
    parts = []
    for sn, a, b, tb in set_scores:
        my, opp = (a, b) if side == "A" else (b, a)
        marker = " <span class='muted'>(TB)</span>" if tb else ""
        parts.append(f"{my}-{opp}{marker}")
    return ", ".join(parts) if parts else '<span class="muted">—</span>'


def render_trajectory_svg(history: list[tuple[str, float, float]]) -> str:
    """history = [(date_iso, mu_after, sigma_after), ...] ordered chronologically."""
    if len(history) < 2:
        return '<p class="muted">Not enough rated matches to draw a trajectory.</p>'

    width, height = 920, 280
    pad_l, pad_r, pad_t, pad_b = 50, 16, 14, 30
    plot_w = width - pad_l - pad_r
    plot_h = height - pad_t - pad_b

    mus = [h[1] for h in history]
    sigs = [h[2] for h in history]
    mu_lo = min(m - 3*s for m, s in zip(mus, sigs)) - 1
    mu_hi = max(m for m in mus) + 2
    if mu_hi - mu_lo < 5:
        mu_hi = mu_lo + 5

    n = len(history)

    def x_at(i): return pad_l + (plot_w * i / (n - 1))
    def y_at(v): return pad_t + plot_h - (plot_h * (v - mu_lo) / (mu_hi - mu_lo))

    # μ line
    mu_pts = " ".join(f"{x_at(i):.1f},{y_at(mus[i]):.1f}" for i in range(n))
    # μ-3σ line (conservative ranking)
    cons_pts = " ".join(
        f"{x_at(i):.1f},{y_at(mus[i] - 3*sigs[i]):.1f}" for i in range(n)
    )

    # Y-axis ticks
    y_ticks = []
    step = 5
    t = int(mu_lo // step * step)
    while t <= mu_hi:
        if t >= mu_lo:
            y = y_at(t)
            y_ticks.append(
                f'<line x1="{pad_l}" y1="{y:.1f}" x2="{width-pad_r}" y2="{y:.1f}" '
                f'stroke="#2a3242" stroke-width="0.5"/>'
                f'<text x="{pad_l-6}" y="{y+3:.1f}" text-anchor="end" '
                f'fill="#8b96a8" font-size="10">{t}</text>'
            )
        t += step

    # X-axis: year markers (when the year in the date string changes)
    x_ticks = []
    last_year = None
    for i, (d, _, _) in enumerate(history):
        yr = d[:4] if d else None
        if yr and yr != last_year:
            x = x_at(i)
            x_ticks.append(
                f'<line x1="{x:.1f}" y1="{pad_t}" x2="{x:.1f}" y2="{pad_t+plot_h}" '
                f'stroke="#2a3242" stroke-width="0.5" stroke-dasharray="2,2"/>'
                f'<text x="{x:.1f}" y="{height-10}" text-anchor="middle" '
                f'fill="#8b96a8" font-size="10">{yr}</text>'
            )
            last_year = yr

    # σ trajectory (uncertainty over time) — uses a separate y-axis on the right
    sig_lo = max(0, min(sigs) - 0.5)
    sig_hi = max(sigs) + 0.5
    if sig_hi - sig_lo < 1.0:
        sig_hi = sig_lo + 1.0

    def y_sig(v):
        return pad_t + plot_h - (plot_h * (v - sig_lo) / (sig_hi - sig_lo))

    sig_pts = " ".join(f"{x_at(i):.1f},{y_sig(sigs[i]):.1f}" for i in range(n))

    # σ axis labels on right side
    sig_axis = []
    for v in (sig_lo, (sig_lo + sig_hi) / 2, sig_hi):
        y = y_sig(v)
        sig_axis.append(
            f'<text x="{width-pad_r+4}" y="{y+3:.1f}" text-anchor="start" '
            f'fill="#e07a7a" font-size="9">{v:.1f}</text>'
        )

    return f"""
<div class="chart">
<svg viewBox="0 0 {width} {height}" xmlns="http://www.w3.org/2000/svg" role="img" aria-label="rating trajectory">
  <rect x="0" y="0" width="{width}" height="{height}" fill="#1a1f2a"/>
  {''.join(y_ticks)}
  {''.join(x_ticks)}
  <polyline points="{cons_pts}" fill="none" stroke="#8b96a8" stroke-width="1" stroke-dasharray="3,3"/>
  <polyline points="{sig_pts}" fill="none" stroke="#e07a7a" stroke-width="1" stroke-dasharray="2,2" opacity="0.6"/>
  <polyline points="{mu_pts}" fill="none" stroke="#4ea1ff" stroke-width="2"/>
  {''.join(sig_axis)}
  <text x="{width-pad_r}" y="{pad_t+10}" text-anchor="end" fill="#4ea1ff" font-size="11">μ (skill, left axis)</text>
  <text x="{width-pad_r}" y="{pad_t+24}" text-anchor="end" fill="#8b96a8" font-size="11">μ-3σ (conservative)</text>
  <text x="{width-pad_r}" y="{pad_t+38}" text-anchor="end" fill="#e07a7a" font-size="11">σ (uncertainty, right axis)</text>
</svg>
</div>
"""


def _row_won(m) -> bool:
    """True iff this player effectively won — uses match_result so tied
    rubbers (sets 1-1) are decided by games-won tiebreak, matching the
    rating engine's convention."""
    my_won = m[15] or 0
    # opp_won was added at column 22 (see PLAYER_MATCHES_SQL). Older test
    # tuples may not include it; fall back to 0 (forces the helper to
    # consult games when my_won is also 0 — same as a true tie).
    opp_won = m[22] if len(m) > 22 and m[22] is not None else 0
    my_games = m[13] or 0
    opp_games = m[18] or 0
    cls, _label, _long = match_result(my_won, opp_won, my_games, opp_games)
    return cls == "win"


def compute_form(matches: list, last_n: int = 10) -> str:
    """Return HTML for the last N results as W/L badges (newest first).

    Tied rubbers render as "W (g)" / "L (g)" — the games-tiebreak winner
    gets the W badge to match how the rating engine scores the match.
    """
    recent = list(reversed(matches[-last_n:]))
    if not recent:
        return '<span class="muted">no matches</span>'
    parts = []
    for m in recent:
        my_won = m[15] or 0
        opp_won = m[22] if len(m) > 22 and m[22] is not None else 0
        my_games = m[13] or 0
        opp_games = m[18] or 0
        cls, label, _ = match_result(my_won, opp_won, my_games, opp_games)
        wl = "w" if cls == "win" else "l"
        # Form badges are tight; collapse "W (g)" → "W*" so the badge stays
        # one column wide. The full label still appears on the match row.
        ch = "W*" if label == "W (g)" else ("L*" if label == "L (g)" else label)
        parts.append(f'<span class="b {wl}">{ch}</span>')
    return f'<span class="form">{"".join(parts)}</span>'


def compute_streaks(matches: list) -> tuple[int, int, int, str]:
    """Return (longest_win, longest_loss, current, current_kind).
    current_kind is 'W' or 'L' depending on the latest match.

    Tied rubbers count toward W/L based on the games-won tiebreak (see
    match_result) so streak logic stays consistent with the rating engine.
    """
    if not matches:
        return (0, 0, 0, "")
    longest_w = longest_l = cur = 0
    cur_kind = ""
    last = None
    for m in matches:
        won = _row_won(m)
        if won == last:
            cur += 1
        else:
            cur = 1
        last = won
        if won:
            longest_w = max(longest_w, cur)
        else:
            longest_l = max(longest_l, cur)
    cur_kind = "W" if last else "L"
    return (longest_w, longest_l, cur, cur_kind)


def compute_yearly_summary(matches: list) -> list[dict]:
    """Bucket matches by year. Returns list of dicts ordered by year."""
    by_year: dict[str, dict] = {}
    last_mu_by_year: dict[str, list] = {}
    for m in matches:
        played = m[1] or ""
        year = played[:4] if len(played) >= 4 else "?"
        bucket = by_year.setdefault(year, {
            "year": year, "n": 0, "wins": 0, "losses": 0,
            "games_won": 0, "games_lost": 0,
            "mu_first": None, "mu_last": None, "sigma_last": None,
        })
        bucket["n"] += 1
        # Use match_result so tied rubbers go in the correct W/L bucket
        # (games-tiebreak winner counts as a win — matches the rating engine).
        if _row_won(m):
            bucket["wins"] += 1
        else:
            bucket["losses"] += 1
        bucket["games_won"] += (m[13] or 0)
        bucket["games_lost"] += (m[18] or 0)
        if m[20] is not None:
            if bucket["mu_first"] is None:
                bucket["mu_first"] = m[20]
            bucket["mu_last"] = m[20]
            bucket["sigma_last"] = m[21]
    return [by_year[y] for y in sorted(by_year)]


def render_identity_section(
    conn: sqlite3.Connection, pid: int, canonical_name: str
) -> str:
    """Show every alias seen for this player + every merge event that absorbed
    another record into this one. Full transparency on identity decisions.
    """
    import json as _json

    aliases = conn.execute(PLAYER_ALIASES_SQL, (pid,)).fetchall()
    merges = conn.execute(PLAYER_MERGES_IN_SQL, (pid,)).fetchall()

    # Canonical name itself isn't always present as an alias; show it first
    # so the reader sees the "official" name at the top of the alias list.
    alias_rows: list[tuple[str, str, str]] = [
        (canonical_name, "—", "(canonical)")
    ]
    for raw_name, first_seen_at, src in aliases:
        if raw_name == canonical_name:
            continue  # already shown
        alias_rows.append((
            raw_name or "",
            (first_seen_at or "")[:10],
            src or "",
        ))

    alias_html = "".join(
        f'<tr>'
        f'<td>{esc(name)}</td>'
        f'<td class="muted">{esc(seen)}</td>'
        f'<td class="muted">{esc(src)}</td>'
        f'</tr>'
        for (name, seen, src) in alias_rows
    )

    merge_rows_html = ""
    if merges:
        # Re-fetch with audit_log id so each row gets a stable deeplink to the
        # site-wide mapping page (#m-<audit_id>).
        merges_with_id = conn.execute(
            """
            SELECT al.id, al.entity_id, al.ts, al.before_jsonb, al.after_jsonb
            FROM audit_log al
            WHERE al.action = 'player.merged'
              AND al.entity_type = 'players'
              AND json_extract(al.after_jsonb, '$.merged_into_id') = ?
            ORDER BY al.ts
            """,
            (pid,),
        ).fetchall()
        rows = []
        for audit_id, loser_id, merged_at, before_json, after_json in merges_with_id:
            try:
                before = _json.loads(before_json or "{}")
                after = _json.loads(after_json or "{}")
            except _json.JSONDecodeError:
                before, after = {}, {}
            loser_name = before.get("canonical_name", f"#{loser_id}")
            reason = after.get("reason", "—")
            merged_on = (merged_at or "")[:10]
            rows.append(
                f'<tr>'
                f'<td><span class="tag">id #{loser_id}</span> {esc(loser_name)}</td>'
                f'<td class="muted">{esc(merged_on)}</td>'
                f'<td>{esc(reason)}</td>'
                f'<td><a href="../aliases.html#m-{audit_id}" title="Permalink in mapping log">#{audit_id} ↗</a></td>'
                f'</tr>'
            )
        merge_rows_html = "".join(rows)
    else:
        merge_rows_html = (
            '<tr><td colspan="4" class="muted">'
            'No merges — this is a single original record.'
            '</td></tr>'
        )

    return f"""
  <h2 class="section-title">Identity &amp; merge history</h2>
  <p class="muted" style="font-size:12px; margin-top:-4px;">
    Every raw name spelling seen in source data, and every other record that
    was merged into this player. See <a href="../aliases.html">all merges
    site-wide ↗</a> for the full mapping log.
  </p>
  <div class="table-wrap">
  <table>
    <thead><tr>
      <th>Name as recorded</th>
      <th>First seen</th>
      <th>Source file</th>
    </tr></thead>
    <tbody>{alias_html}</tbody>
  </table>
  </div>
  <div class="table-wrap" style="margin-top:8px;">
  <table>
    <thead><tr>
      <th>Merged-in record</th>
      <th>Merged on</th>
      <th>Reason</th>
      <th></th>
    </tr></thead>
    <tbody>{merge_rows_html}</tbody>
  </table>
  </div>
"""


def compute_swings(match_rows_with_delta: list) -> tuple[list, list]:
    """Return (top_3_wins, top_3_losses) as lists of (delta, row_dict).
    Input items: dict with keys delta, played, opps, partner, score, result."""
    sorted_by_delta = sorted(match_rows_with_delta, key=lambda r: r["delta"])
    losses = sorted_by_delta[:3]  # most-negative first
    wins = list(reversed(sorted_by_delta[-3:]))  # most-positive first
    return ([w for w in wins if w["delta"] > 0],
            [l for l in losses if l["delta"] < 0])


def compute_match_impacts(conn: sqlite3.Connection) -> dict:
    """Replay every active match chronologically and snapshot per-player
    rank-bucket position before/after each match.

    Returns: {(match_id, player_id) -> impact dict} with keys:
        side ('A'|'B'), won (bool),
        rank_before (int|None), rank_after (int),
        mu_before (float|None), mu_after (float),
        score_before (float|None), score_after (float),
        mu_delta (float), score_delta (float),
        bypassed (list[int]): pids overtaken on the way up,
        passed_by (list[int]): pids who overtook me on the way down,
        bucket_size_after (int)

    Rank is the player's 1-indexed position in their gender bucket
    (M / F / U), sorted by μ-3σ desc. A player only enters the bucket
    once they have ≥1 rating_history row.
    """
    # Bulk pull, keyed by match_id (NOT model — only one model so far).
    rh_rows = conn.execute(
        "SELECT match_id, player_id, mu_after, sigma_after "
        "FROM rating_history WHERE model_name = ?",
        (MODEL,),
    ).fetchall()
    rh: dict[tuple[int, int], tuple[float, float]] = {
        (mid, pid): (mu, sg) for mid, pid, mu, sg in rh_rows
    }

    genders: dict[int, str] = {
        pid: (g if g in ("M", "F") else "U")
        for pid, g in conn.execute("SELECT id, gender FROM players")
    }

    # Active matches in chronological order, with both sides.
    # games_won is needed to resolve tied rubbers (sets 1-1 → both won=0;
    # the games-tiebreak winner is the "winner" for display purposes,
    # matching the rating engine — see match_result()).
    matches = conn.execute(
        """
        SELECT m.id,
               sa.player1_id, sa.player2_id, sa.won, sa.games_won,
               sb.player1_id, sb.player2_id, sb.won, sb.games_won
        FROM matches m
        JOIN match_sides sa ON sa.match_id = m.id AND sa.side = 'A'
        JOIN match_sides sb ON sb.match_id = m.id AND sb.side = 'B'
        WHERE m.superseded_by_run_id IS NULL
        ORDER BY m.played_on ASC, m.id ASC
        """
    ).fetchall()

    # Current state per player: pid -> (mu, sigma). Players not yet in here
    # haven't played a match yet (pre-system).
    state: dict[int, tuple[float, float]] = {}
    impacts: dict[tuple[int, int], dict] = {}

    for row in matches:
        mid, a1, a2, awon, agw, b1, b2, bwon, bgw = row
        # Resolve "who won this match" via the same convention the rating
        # engine uses: clean win on sets first, otherwise games-tiebreak.
        # match_result returns ('win'|'loss', label, label_long) from one
        # side's perspective; we only need the binary cls here.
        a_cls, _a_label, _ = match_result(awon or 0, bwon or 0,
                                          agw or 0, bgw or 0)
        a_won_eff = (a_cls == "win")
        b_won_eff = not a_won_eff
        sides = [(a1, "A", a_won_eff), (a2, "A", a_won_eff),
                 (b1, "B", b_won_eff), (b2, "B", b_won_eff)]
        participants = [(p, sd, w) for p, sd, w in sides if p is not None]

        # Collect new (mu, sigma) for every participant that has rating_history
        # for this match. Walkovers / no-rating matches will be missing — we
        # still record an impact (rank_before = rank_after, no deltas) so the
        # UI shows the players' standing at the time of the match.
        new_for: dict[int, tuple[float, float]] = {}
        for pid, _sd, _w in participants:
            entry = rh.get((mid, pid))
            if entry is not None:
                new_for[pid] = entry

        # Snapshot bucket BEFORE: per gender, list of (pid, score).
        bucket_before: dict[str, list[tuple[int, float]]] = {"M": [], "F": [], "U": []}
        for pid, (mu, sg) in state.items():
            bucket_before[genders.get(pid, "U")].append((pid, mu - 3 * sg))

        # Project state AFTER (only participants change).
        state_after = dict(state)
        for pid, (mu, sg) in new_for.items():
            state_after[pid] = (mu, sg)
        bucket_after: dict[str, list[tuple[int, float]]] = {"M": [], "F": [], "U": []}
        for pid, (mu, sg) in state_after.items():
            bucket_after[genders.get(pid, "U")].append((pid, mu - 3 * sg))

        # Compute impact per participant.
        for pid, side, won in participants:
            g = genders.get(pid, "U")
            in_before = pid in state

            if in_before:
                mu_b, sg_b = state[pid]
                score_before = mu_b - 3 * sg_b
                # 1-indexed rank: count of others strictly above me + 1
                rank_before = 1 + sum(
                    1 for p, s in bucket_before[g] if p != pid and s > score_before
                )
            else:
                mu_b = sg_b = None
                score_before = None
                rank_before = None

            if pid in new_for:
                mu_a, sg_a = new_for[pid]
            else:
                # No rating change recorded for this match (rare — walkover etc).
                # Carry the old state so rank_after is computed sensibly.
                if not in_before:
                    continue  # nothing to record
                mu_a, sg_a = mu_b, sg_b
            score_after = mu_a - 3 * sg_a
            rank_after = 1 + sum(
                1 for p, s in bucket_after[g] if p != pid and s > score_after
            )

            mu_delta = (mu_a - mu_b) if mu_b is not None else 0.0
            score_delta = (score_after - score_before) if score_before is not None else 0.0

            bypassed: list[int] = []
            passed_by: list[int] = []
            if in_before:
                # Compare against every other bucket member's BEFORE/AFTER.
                # Non-participants didn't move; participants in this match did.
                before_map = {p: s for p, s in bucket_before[g]}
                after_map = {p: s for p, s in bucket_after[g]}
                for other_pid, other_before in before_map.items():
                    if other_pid == pid:
                        continue
                    other_after = after_map.get(other_pid, other_before)
                    if other_before > score_before and other_after <= score_after:
                        bypassed.append(other_pid)
                    elif other_before < score_before and other_after >= score_after:
                        passed_by.append(other_pid)

            impacts[(mid, pid)] = {
                "side": side,
                "won": won,
                "rank_before": rank_before,
                "rank_after": rank_after,
                "mu_before": mu_b,
                "mu_after": mu_a,
                "score_before": score_before,
                "score_after": score_after,
                "mu_delta": mu_delta,
                "score_delta": score_delta,
                "bypassed": bypassed,
                "passed_by": passed_by,
                "bucket_size_after": len(bucket_after[g]),
            }

        # Commit state.
        for pid, (mu, sg) in new_for.items():
            state[pid] = (mu, sg)

    return impacts


def _delta_span(delta: float, decimals: int = 2) -> str:
    """Render a +X.XX / -X.XX / 0 span with appropriate color class."""
    if delta > 0.005:
        cls = "delta-up"
        return f'<span class="{cls}">+{delta:.{decimals}f}</span>'
    if delta < -0.005:
        cls = "delta-dn"
        return f'<span class="{cls}">{delta:.{decimals}f}</span>'
    return f'<span class="delta-z">±0</span>'


def _rank_delta_span(rank_before: int | None, rank_after: int) -> str:
    """Rank delta is inverted (lower number = higher position)."""
    if rank_before is None:
        return ""
    diff = rank_before - rank_after  # positive = moved up
    if diff > 0:
        return f'<span class="delta-up">+{diff}</span>'
    if diff < 0:
        return f'<span class="delta-dn">{diff}</span>'
    return '<span class="delta-z">±0</span>'


def render_match_impact_block(
    mid: int,
    participants: list[tuple],
    impacts: dict,
    name_lookup: dict,
    players_prefix: str = "players/",
) -> str:
    """Render the expanded per-player impact section for ONE match.

    `participants` items are tuples whose first 4 elements are
    (pid, partner_pid_or_None, side ('A'|'B'), won (bool)).
    An optional 5th element is a result label override ('W', 'L',
    'W (g)', or 'L (g)') — supply it for tied rubbers so the card shows
    the games-tiebreak indicator. If omitted, falls back to plain W/L.
    The list is in display order: side A players first, then side B.
    `players_prefix` is the URL prefix for player pages — `players/` from a
    root-level page, or `` (empty) from /players/X.html.
    """
    if not participants:
        return ""

    def _link(p: int) -> str:
        cid, n = name_lookup.get(p, (p, f"#{p}"))
        return f'<a class="player-link" href="{players_prefix}{cid}.html">{esc(n)}</a>'

    def _card(pid: int, side: str, won: bool, label: str | None = None) -> str:
        imp = impacts.get((mid, pid))
        link = _link(pid)
        wl_cls = "win" if won else "loss"
        wl_txt = label if label else ("W" if won else "L")
        if imp is None:
            return (
                f'<div class="impact-player">'
                f'<div class="who">{link}'
                f'<span class="side-tag {wl_cls}">{side} · {wl_txt}</span></div>'
                f'<div class="new-entry">No rating change recorded</div>'
                f'</div>'
            )
        rank_before = imp["rank_before"]
        rank_after = imp["rank_after"]
        score_before = imp["score_before"]
        score_after = imp["score_after"]
        score_delta = imp["score_delta"]
        bypassed = imp["bypassed"]
        passed_by = imp["passed_by"]

        if rank_before is None:
            rank_line = (
                f'<div class="metric"><span class="k">Rank</span>'
                f'<span class="new-entry">new entry → #{rank_after}</span></div>'
            )
        else:
            r_delta = _rank_delta_span(rank_before, rank_after)
            rank_line = (
                f'<div class="metric"><span class="k">Rank</span>'
                f'<span>#{rank_before} {r_delta} = #{rank_after}</span></div>'
            )

        if score_before is None:
            score_line = (
                f'<div class="metric"><span class="k">Score</span>'
                f'<span class="new-entry">starts at {score_after:.1f}</span></div>'
            )
        else:
            sd = _delta_span(score_delta, decimals=2)
            score_line = (
                f'<div class="metric"><span class="k">Score</span>'
                f'<span>{score_before:.1f} {sd} = {score_after:.1f}</span></div>'
            )

        commentary = ""
        if bypassed:
            names = ", ".join(_link(p) for p in bypassed[:5])
            extra = f" + {len(bypassed) - 5} more" if len(bypassed) > 5 else ""
            commentary = (
                f'<div class="commentary">'
                f'<span class="pass-up">↑ bypassed</span> {names}{extra}'
                f'</div>'
            )
        elif passed_by:
            names = ", ".join(_link(p) for p in passed_by[:5])
            extra = f" + {len(passed_by) - 5} more" if len(passed_by) > 5 else ""
            commentary = (
                f'<div class="commentary">'
                f'<span class="pass-dn">↓ passed by</span> {names}{extra}'
                f'</div>'
            )

        return (
            f'<div class="impact-player">'
            f'<div class="who">{link}'
            f'<span class="side-tag {wl_cls}">{side} · {wl_txt}</span></div>'
            f'{rank_line}'
            f'{score_line}'
            f'{commentary}'
            f'</div>'
        )

    # Group cards by side, preserving partner order within each side. The
    # 2-vs-2 layout reads as: [pair A] | VS | [pair B]. participants tuple
    # is (pid, partner, side, won) with optional 5th label override.
    def _label_of(p):
        return p[4] if len(p) >= 5 else None

    side_a_cards = [
        _card(p[0], p[2], p[3], _label_of(p))
        for p in participants if p[2] == "A"
    ]
    side_b_cards = [
        _card(p[0], p[2], p[3], _label_of(p))
        for p in participants if p[2] == "B"
    ]

    return (
        f'<div class="impact-box">'
        f'<div class="impact-side side-A">{"".join(side_a_cards)}</div>'
        f'<div class="impact-vs">VS</div>'
        f'<div class="impact-side side-B">{"".join(side_b_cards)}</div>'
        f'</div>'
    )


def build_player_page(
    conn: sqlite3.Connection,
    pid: int,
    name_lookup: dict[int, str],
    neighbours_by_gender: dict[str, list[dict]] | None = None,
    impacts: dict | None = None,
    predictions: dict[str, dict[int, dict]] | None = None,
) -> str:
    info = conn.execute(PLAYER_INFO_SQL, (pid,)).fetchone()
    if info is None:
        return ""
    _, name, gender = info
    impacts = impacts or {}
    # predictions = {model_name: {match_id: {p_a, actual_a, log_loss, ...}}}
    predictions = predictions or {}

    rating = conn.execute(PLAYER_RATING_SQL, (pid, MODEL)).fetchone()
    mu, sigma = (rating if rating else (None, None))
    mu3s = mu - 3 * sigma if mu is not None else None

    matches = conn.execute(
        PLAYER_MATCHES_SQL, (pid, MODEL, pid, pid)
    ).fetchall()

    class_history = conn.execute(PLAYER_CLASS_HISTORY_SQL, (pid,)).fetchall()

    # Column index map (matches PLAYER_MATCHES_SQL):
    #   0=mid 1=played 2=division 3=round 4=walkover
    #   5=tour_id 6=tour_name 7=year 8=club_name 9=club_slug
    #   10=side 11=my_p1 12=my_p2 13=my_games 14=my_sets 15=my_won
    #   16=opp_p1 17=opp_p2 18=opp_games 19=opp_sets
    #   20=mu_after 21=sigma_after 22=opp_won
    n = len(matches)
    # Tied rubbers (sets 1-1, both won=0) count as a win for the
    # games-tiebreak winner — same convention as the rating engine.
    wins = sum(1 for m in matches if _row_won(m))
    losses = n - wins
    games_won = sum((m[13] or 0) for m in matches)
    games_lost = sum((m[18] or 0) for m in matches)
    win_pct = f"{int(round(wins * 100 / n))}%" if n else "—"

    # Trajectory: only matches where rating_history exists.
    history = [
        (m[1], m[20], m[21]) for m in matches
        if m[20] is not None and m[21] is not None
    ]
    traj = render_trajectory_svg(history)

    # Peers section: 3 above + me + 3 below in the appropriate gender bucket.
    peers_html = ""
    if neighbours_by_gender:
        bucket_key = gender if gender in ("M", "F") else "all"
        bucket = neighbours_by_gender.get(bucket_key, [])
        peers_html = render_neighbours(bucket, pid)

    # Identity block: aliases + merge history (full transparency).
    identity_block = render_identity_section(conn, pid, name)

    # Match rows. Δμ must be computed in chronological order (per match it's
    # mu_after − mu_after_of_prior_match), but the table is rendered newest
    # first. Build the rows then reverse before joining.
    # Per-player calibration: log-loss, accuracy, and best/worst-prediction
    # under each model on this player's actual matches. Computed inline below.
    pred_stats: dict[str, dict] = {
        m: {"n": 0, "log_loss_sum": 0.0, "correct": 0,
            "best_called": None, "worst_called": None}
        for m in predictions
    }

    match_rows = []
    swing_data = []  # for biggest-swings analysis
    last_mu, last_sig = None, None
    for m in matches:
        (mid, played, division, rnd, walkover, _tid, tname, tyear, club_name, club_slug,
         side, my_p1, my_p2, my_games, my_sets, my_won,
         opp_p1, opp_p2, opp_games, opp_sets, mu_after, sigma_after,
         opp_won) = m

        # Predicted P(this player's side wins) under each model. The CSV
        # stores P(side A wins); flip if this player is on side B.
        my_predictions: dict[str, float | None] = {}
        for model_name, pred_map in predictions.items():
            entry = pred_map.get(mid)
            if entry is None:
                my_predictions[model_name] = None
                continue
            p_a = entry["p_a"]
            p_me = p_a if side == "A" else 1.0 - p_a
            my_predictions[model_name] = p_me
            # Update calibration stats for this model. Tied rubbers count
            # the games-tiebreak winner as the actual winner — same
            # convention as the rating engine.
            _my_cls_calib, _, _ = match_result(
                my_won or 0, opp_won or 0, my_games or 0, opp_games or 0
            )
            actual_me = 1 if _my_cls_calib == "win" else 0
            stats = pred_stats[model_name]
            stats["n"] += 1
            stats["log_loss_sum"] += -(
                actual_me * math.log(max(min(p_me, 1 - 1e-9), 1e-9))
                + (1 - actual_me) * math.log(max(min(1 - p_me, 1 - 1e-9), 1e-9))
            )
            if (p_me > 0.5) == bool(actual_me):
                stats["correct"] += 1
            # Track best-called (highest confidence in correct outcome) and
            # worst-called (highest confidence in wrong outcome).
            confidence_in_actual = p_me if actual_me else (1 - p_me)
            if (stats["best_called"] is None
                    or confidence_in_actual > stats["best_called"][0]):
                stats["best_called"] = (confidence_in_actual, mid, played, tname)
            if (stats["worst_called"] is None
                    or confidence_in_actual < stats["worst_called"][0]):
                stats["worst_called"] = (confidence_in_actual, mid, played, tname)
        sets = conn.execute(SET_SCORES_SQL, (mid,)).fetchall()
        partner = render_partner(my_p1, my_p2, pid, name_lookup)
        opps = render_opponents(opp_p1, opp_p2, name_lookup)
        score = render_score(side, sets)
        # Tied rubbers (sets 1-1, both won=0) → the side with more games
        # is the "winner" per the rating engine. match_result() returns
        # 'W (g)' / 'L (g)' to flag the games-tiebreak in the UI.
        result_cls, result_txt, _result_long = match_result(
            my_won or 0, opp_won or 0, my_games or 0, opp_games or 0
        )
        d_mu = ""
        delta_val = None
        d_sig = ""
        if mu_after is not None and last_mu is not None:
            delta_val = mu_after - last_mu
            colour = "win" if delta_val > 0 else ("loss" if delta_val < 0 else "muted")
            d_mu = (
                f'<span class="{colour}">'
                f'{"+" if delta_val>=0 else ""}{delta_val:.2f}</span>'
            )
            if last_sig is not None and sigma_after is not None:
                ds = sigma_after - last_sig
                d_sig = (
                    f'<span class="muted">{"+" if ds>=0 else ""}{ds:.2f}</span>'
                )
        if mu_after is not None:
            last_mu, last_sig = mu_after, sigma_after
        mu_cell = f"{mu_after:.2f}" if mu_after is not None else '<span class="muted">—</span>'
        sig_cell = f"{sigma_after:.2f}" if sigma_after is not None else '<span class="muted">—</span>'
        wo = ' <span class="tag">W/O</span>' if walkover else ""

        # Rank-at-time tag for THIS player. From a /players/X.html page the
        # player links use no prefix; the tag itself doesn't link, just shows
        # the rank.
        my_imp = impacts.get((mid, pid))
        my_rank_tag = (
            f'<span class="rank-tag">#{my_imp["rank_after"]}</span>'
            if my_imp else ""
        )

        # Determine if we have an impact row to render for this match.
        all_pids = [p for p in (my_p1, my_p2, opp_p1, opp_p2) if p is not None]
        has_impact = any((mid, p) in impacts for p in all_pids)
        trigger_cell = (
            f'<td class="expand-trigger" data-mid="{mid}">▶</td>'
            if has_impact
            else '<td class="muted" style="text-align:center;">·</td>'
        )

        # Decay model is the better-calibrated predictor (per backtest);
        # show its prediction inline. Tooltip shows vanilla PL prediction
        # so the user can compare without an extra column.
        pred_decay = my_predictions.get("openskill_pl_decay365")
        pred_pl = my_predictions.get("openskill_pl_vanilla")
        if pred_decay is None:
            pred_cell = '<td class="num muted" data-sort="-1">—</td>'
        else:
            actual_me = 1 if my_won else 0
            # Colour: green if predicted correctly with confidence, red if
            # predicted wrong with confidence, muted if unsure.
            if (pred_decay > 0.5) == bool(actual_me):
                colour = "win" if abs(pred_decay - 0.5) > 0.15 else "muted"
            else:
                colour = "loss" if abs(pred_decay - 0.5) > 0.15 else "muted"
            tooltip = (
                f'Decay-365 said {pred_decay*100:.0f}% chance of winning. '
                f'Vanilla PL said {pred_pl*100:.0f}%.'
                if pred_pl is not None else
                f'Decay-365 said {pred_decay*100:.0f}% chance of winning.'
            )
            pred_cell = (
                f'<td class="num" data-sort="{pred_decay:.4f}" '
                f'title="{esc(tooltip)}">'
                f'<span class="{colour}">{pred_decay*100:.0f}%</span></td>'
            )

        main_row = (
            f'<tr class="match-row" data-mid="{mid}">'
            f'{trigger_cell}'
            f'<td>{esc(played)}</td>'
            f'<td><span class="tag">{esc(club_slug)}</span> {esc(tname)} {esc(tyear)}</td>'
            f'<td class="muted">{esc(division or "")} {esc(rnd or "")}</td>'
            f'<td>{partner}</td>'
            f'<td>{opps}</td>'
            f'<td class="score">{score}{wo}</td>'
            f'<td class="num"><span class="{result_cls}">{result_txt}</span> '
            f'{my_games}-{opp_games}</td>'
            f'{pred_cell}'
            f'<td class="num">{mu_cell}{my_rank_tag}</td>'
            f'<td class="num">{d_mu}</td>'
            f'<td class="num">{sig_cell}</td>'
            f'<td class="num">{d_sig}</td>'
            f'</tr>'
        )
        impact_row = ""
        if has_impact:
            # Read both DB won-flags explicitly. The old code (`opp_won_bool
            # = not my_won_bool`) was wrong even before tied-rubber support
            # — it ignored the actual opp_won value. For tied rubbers (sets
            # 1-1), match_result resolves the winner via games-tiebreak,
            # matching the rating engine. The 5th tuple element is the
            # display label ("W (g)" / "L (g)") so the impact card shows
            # the tiebreak indicator.
            my_cls, my_label, _ = match_result(
                my_won or 0, opp_won or 0, my_games or 0, opp_games or 0
            )
            opp_cls, opp_label, _ = match_result(
                opp_won or 0, my_won or 0, opp_games or 0, my_games or 0
            )
            my_won_bool = (my_cls == "win")
            opp_won_bool = (opp_cls == "win")
            opp_side = "B" if side == "A" else "A"
            participants = [
                (my_p1, my_p2, side, my_won_bool, my_label),
                (my_p2, my_p1, side, my_won_bool, my_label),
                (opp_p1, opp_p2, opp_side, opp_won_bool, opp_label),
                (opp_p2, opp_p1, opp_side, opp_won_bool, opp_label),
            ]
            participants = [p for p in participants if p[0] is not None]
            impact_html = render_match_impact_block(
                mid, participants, impacts, name_lookup, players_prefix=""
            )
            impact_row = (
                f'<tr class="impact-row" data-mid="{mid}" hidden>'
                f'<td colspan="13">{impact_html}</td>'
                f'</tr>'
            )
        # Pair them so reversing keeps the impact row immediately after its
        # match row.
        match_rows.append((main_row, impact_row))
        if delta_val is not None:
            swing_data.append({
                "delta": delta_val,
                "played": played,
                "opps_html": opps,
                "partner_html": partner,
                "score_html": score,
                "result_cls": result_cls,
                "result_txt": result_txt,
                "tname": tname,
                "club_slug": club_slug,
            })

    # Form (last 10), streaks, yearly summary, biggest swings
    form_html = compute_form(matches, last_n=10)
    longest_w, longest_l, cur_streak, cur_kind = compute_streaks(matches)
    yearly = compute_yearly_summary(matches)
    biggest_wins, biggest_losses = compute_swings(swing_data)

    def _yearly_row(y):
        mu_range = (
            f"{y['mu_first']:.2f} → {y['mu_last']:.2f}"
            if y["mu_first"] is not None else "—"
        )
        if y["mu_first"] is not None and y["mu_last"] is not None:
            d = y["mu_last"] - y["mu_first"]
            cls = "win" if d > 0 else ("loss" if d < 0 else "muted")
            sign = "+" if d >= 0 else ""
            d_cell = f'<span class="{cls}">{sign}{d:.2f}</span>'
        else:
            d_cell = "—"
        return (
            f'<tr>'
            f'<td><strong>{esc(y["year"])}</strong></td>'
            f'<td class="num">{y["n"]}</td>'
            f'<td class="num"><span class="win">{y["wins"]}</span>-<span class="loss">{y["losses"]}</span></td>'
            f'<td class="num">{int(round(y["wins"]*100/y["n"]))}%</td>'
            f'<td class="num">{y["games_won"]}-{y["games_lost"]}</td>'
            f'<td class="num">{mu_range}</td>'
            f'<td class="num">{d_cell}</td>'
            f'</tr>'
        )

    yearly_rows = "".join(_yearly_row(y) for y in yearly)

    def render_swing_row(s):
        d = s["delta"]
        cls = "win" if d > 0 else "loss"
        return (
            f'<tr>'
            f'<td>{esc(s["played"])}</td>'
            f'<td><span class="tag">{esc(s["club_slug"])}</span> {esc(s["tname"])}</td>'
            f'<td>{s["partner_html"]}</td>'
            f'<td>{s["opps_html"]}</td>'
            f'<td class="score">{s["score_html"]}</td>'
            f'<td class="num"><span class="{s["result_cls"]}">{s["result_txt"]}</span></td>'
            f'<td class="num"><span class="{cls}">{"+" if d>=0 else ""}{d:.2f}</span></td>'
            f'</tr>'
        )

    swings_html = ""
    if biggest_wins or biggest_losses:
        win_rows = "".join(render_swing_row(s) for s in biggest_wins) or \
            '<tr><td colspan="7" class="muted">No positive swings recorded.</td></tr>'
        loss_rows = "".join(render_swing_row(s) for s in biggest_losses) or \
            '<tr><td colspan="7" class="muted">No negative swings recorded.</td></tr>'
        swings_html = f"""
  <h2 class="section-title">Biggest Δμ swings</h2>
  <p class="muted" style="font-size: 12px; margin-top: -4px;">
    Largest single-match rating jumps (signal: upset wins / surprise losses).
  </p>
  <div class="table-wrap">
  <table>
    <thead><tr><th colspan="7" style="background:#1f3a2a; color: var(--win)">Top wins</th></tr>
    <tr><th>Date</th><th>Tournament</th><th>Partner</th><th>Opponents</th><th>Score</th><th class="num">Result</th><th class="num">Δμ</th></tr>
    </thead>
    <tbody>{win_rows}</tbody>
  </table>
  </div>
  <div class="table-wrap">
  <table>
    <thead><tr><th colspan="7" style="background:#3a1f1f; color: var(--loss)">Worst losses</th></tr>
    <tr><th>Date</th><th>Tournament</th><th>Partner</th><th>Opponents</th><th>Score</th><th class="num">Result</th><th class="num">Δμ</th></tr>
    </thead>
    <tbody>{loss_rows}</tbody>
  </table>
  </div>
"""

    class_rows = "".join(
        f'<tr><td>{esc(yr)}</td><td><span class="tag">{esc(slug)}</span> {esc(tn)}</td>'
        f'<td>Team {esc(tl or "?")}</td><td class="muted">{esc(cap or "")}</td>'
        f'<td class="cls">{esc(cl)}</td></tr>'
        for (yr, tn, _cn, slug, tl, cap, cl) in class_history
    )

    if not class_rows:
        class_rows = (
            '<tr><td colspan="5" class="muted">No team-tournament class assignments.</td></tr>'
        )

    streak_value = (
        f'<span class="{"win" if cur_kind=="W" else "loss"}">{cur_streak}{cur_kind}</span>'
        if cur_streak else "—"
    )

    # Per-player calibration summary: under each model, how well did its
    # predictions track this player's actual results? Useful as both a
    # signal of model quality and a way to spot outliers — players where
    # one model is dramatically more accurate than the other.
    pred_summary_block = ""
    if pred_stats and any(s["n"] > 0 for s in pred_stats.values()):
        cells = []
        MODEL_LABEL = {
            "openskill_pl_vanilla": "Vanilla PL",
            "openskill_pl_decay365": "Decay-365",
        }
        for model_name in ("openskill_pl_vanilla", "openskill_pl_decay365"):
            stats = pred_stats.get(model_name)
            if not stats or stats["n"] == 0:
                continue
            label = MODEL_LABEL.get(model_name, model_name)
            acc = stats["correct"] / stats["n"]
            avg_ll = stats["log_loss_sum"] / stats["n"]
            cells.append(
                f'<div class="stat">'
                f'<div class="label">{esc(label)}</div>'
                f'<div class="value">{acc * 100:.0f}%</div>'
                f'<div class="muted" style="font-size:10px;">'
                f'log-loss {avg_ll:.3f}</div>'
                f'</div>'
            )

        # Biggest model surprise: where the two models' predictions for the
        # same match diverged most. Reveals matches where Lonia-style
        # judgment may be needed.
        max_div = 0.0
        max_div_match: dict | None = None
        decay_preds = predictions.get("openskill_pl_decay365", {})
        vanilla_preds = predictions.get("openskill_pl_vanilla", {})
        for m in matches:
            mid, played, _div, _rnd, _wo, _tid, tname = m[0], m[1], m[2], m[3], m[4], m[5], m[6]
            side = m[10]; my_won = m[15]
            d_entry = decay_preds.get(mid)
            v_entry = vanilla_preds.get(mid)
            if not d_entry or not v_entry:
                continue
            d_p = d_entry["p_a"] if side == "A" else 1 - d_entry["p_a"]
            v_p = v_entry["p_a"] if side == "A" else 1 - v_entry["p_a"]
            div = abs(d_p - v_p)
            if div > max_div:
                max_div = div
                actual_me = 1 if my_won else 0
                max_div_match = {
                    "played": played, "tname": tname, "d_p": d_p,
                    "v_p": v_p, "actual": actual_me, "won": bool(my_won),
                }
        divergence_html = ""
        if max_div_match and max_div >= 0.10:
            d = max_div_match
            divergence_html = (
                f'<p class="muted" style="font-size:12px;margin-top:8px;">'
                f'<strong>Biggest model disagreement</strong> on this player\'s '
                f'matches: {esc(d["played"])} vs {esc(d["tname"])} — Vanilla PL '
                f'said {d["v_p"]*100:.0f}%, Decay said {d["d_p"]*100:.0f}% '
                f'({"WIN" if d["won"] else "LOSS"} actually). The bigger this '
                f'gap, the more useful captain knowledge could be.</p>'
            )

        if cells:
            pred_summary_block = (
                '<h2 class="section-title">Prediction quality on this player\'s matches</h2>'
                '<p class="muted" style="font-size:12px;margin-top:-4px;">'
                'Held-out prediction accuracy (model never saw a match before '
                'predicting it). Lower log-loss is better; perfect = 0, random '
                '= 0.693. See <a href="https://github.com/devkurtc/wks-social-tennis-rankings-malta/blob/main/_ANALYSIS_/model_evaluation/SUMMARY.md">SUMMARY.md</a> for methodology.'
                '</p>'
                f'<div class="stat-grid">{"".join(cells)}</div>'
                f'{divergence_html}'
            )

    rating_block = ""
    if mu is not None:
        rating_block = f"""
<div class="stat-grid">
  <div class="stat"><div class="label">μ skill</div><div class="value">{mu:.2f}</div></div>
  <div class="stat"><div class="label">σ uncertainty</div><div class="value">{sigma:.2f}</div></div>
  <div class="stat"><div class="label">μ-3σ rank</div><div class="value">{mu3s:.2f}</div></div>
  <div class="stat"><div class="label">matches</div><div class="value">{n}</div></div>
  <div class="stat"><div class="label">W-L</div>
    <div class="value"><span class="win">{wins}</span>-<span class="loss">{losses}</span></div></div>
  <div class="stat"><div class="label">win %</div><div class="value">{win_pct}</div></div>
  <div class="stat"><div class="label">games W-L</div><div class="value">{games_won}-{games_lost}</div></div>
  <div class="stat"><div class="label">current streak</div><div class="value">{streak_value}</div></div>
  <div class="stat"><div class="label">longest W</div><div class="value win">{longest_w}</div></div>
  <div class="stat"><div class="label">longest L</div><div class="value loss">{longest_l}</div></div>
</div>
<div style="margin-bottom: 16px;">
  <div class="muted" style="font-size: 11px; text-transform: uppercase; letter-spacing: 0.05em; margin-bottom: 4px;">Form (last 10, newest first)</div>
  {form_html}
</div>
"""

    yearly_block = ""
    if yearly:
        yearly_block = f"""
  <h2 class="section-title">Per-year breakdown</h2>
  <div class="table-wrap">
  <table>
    <thead><tr>
      <th>Year</th><th class="num">n</th><th class="num">W-L</th><th class="num">win%</th>
      <th class="num">games</th><th class="num">μ start → end</th><th class="num">Δμ</th>
    </tr></thead>
    <tbody>{yearly_rows}</tbody>
  </table>
  </div>
"""

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1, viewport-fit=cover">
<meta name="theme-color" content="#0f1115">
<title>{esc(name)} — RallyRank</title>
<link rel="stylesheet" href="../styles.css?v={CSS_VERSION}">
</head>
<body>
{render_nav("../", "")}
<main>
  <div class="profile-head">
    <div>
      <div class="name gender-{esc(gender or "")}">{esc(name)}</div>
      <div class="meta">
        {esc({"M":"Men", "F":"Ladies"}.get(gender, "Unknown gender"))} &middot;
        Player ID #{pid}
      </div>
    </div>
  </div>

  {rating_block}

  {peers_html}

  <h2 class="section-title">Rating trajectory</h2>
  {traj}

  {yearly_block}

  {swings_html}

  {identity_block}

  <h2 class="section-title">Captain class assignments</h2>
  <div class="table-wrap">
  <table>
    <thead><tr><th>Year</th><th>Tournament</th><th>Team</th><th>Captain</th><th>Class</th></tr></thead>
    <tbody>{class_rows}</tbody>
  </table>
  </div>

  {pred_summary_block}

  <h2 id="match-log" class="section-title">Match log ({n})</h2>
  <p class="muted" style="font-size: 12px; margin-top: -4px;">
    Number next to "μ after" is this player's rank in their gender bucket
    immediately AFTER the match. The "Pred" column is the Decay-365 model's
    predicted P(this player's side wins), computed at the time of the match
    — i.e., a true held-out prediction. Click ▶ to see every player's rank/
    score impact for that match.
  </p>
  <div class="table-wrap">
  <table id="player-match-log">
    <thead><tr>
      <th></th>
      <th>Date</th><th>Tournament</th><th>Round</th>
      <th>Partner</th><th>Opponents</th><th>Score</th>
      <th class="num">Result</th>
      <th class="num" title="Decay-365 model's predicted probability that this player's side would win, computed using ratings as of just before the match. Hover any cell for the vanilla PL prediction too. Colour: green = called correctly with &gt;15% confidence; red = called wrong with &gt;15% confidence; muted = uncertain.">Pred</th>
      <th class="num">μ after</th><th class="num">Δμ</th>
      <th class="num">σ after</th><th class="num">Δσ</th>
    </tr></thead>
    <tbody>{''.join(main + imp for main, imp in reversed(match_rows)) if match_rows else '<tr><td colspan="13" class="muted">No matches.</td></tr>'}</tbody>
  </table>
  </div>
</main>
<footer>
  μ after / σ after are the OpenSkill PL values <em>after</em> this match was processed.
  Δμ / Δσ are differences from the prior match (blank for the very first rated match).
</footer>
<script>
(function() {{
  const table = document.getElementById('player-match-log');
  if (!table) return;
  table.addEventListener('click', function(ev) {{
    const cell = ev.target.closest('.expand-trigger');
    if (!cell) return;
    const mid = cell.dataset.mid;
    const row = cell.closest('tr.match-row');
    const ir = table.querySelector('tr.impact-row[data-mid="' + mid + '"]');
    if (!ir) return;
    ir.hidden = !ir.hidden;
    row.classList.toggle('open', !ir.hidden);
  }});
}})();
</script>
</body>
</html>
"""


# --- Tournament roster pages -------------------------------------------------
# A roster page ranks a list of players (read from an .xlsx) against the
# current ratings DB. Useful for pre-tournament seeding when you have the
# entry list before any matches have been played.


def _read_roster_xlsx(path: Path) -> dict[str, list[str]]:
    """Return {'Men': [...], 'Ladies': [...]} from a 2-sheet roster file.

    Each sheet is expected to have header row + two name columns at indices
    (1, 3) — i.e. the standard '(No, Name, No, Name)' layout used by
    `_ANALYSIS_/NewTournamentRanking/Players List.xlsx` and similar.
    """
    import openpyxl  # local import; only needed when a roster is configured

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    out: dict[str, list[str]] = {}
    for sheet_name in ("Men", "Ladies"):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        seen: set[str] = set()
        names: list[str] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            for col in (0, 2):
                if col + 1 >= len(row):
                    continue
                name = row[col + 1]
                if not isinstance(name, str):
                    continue
                name = name.strip()
                if not name or name.lower().startswith("reserve"):
                    continue
                key = name.lower()
                if key in seen:
                    continue
                seen.add(key)
                names.append(name)
        out[sheet_name] = names
    return out


def _name_order_variants(name: str) -> list[str]:
    """Plausible word-order rotations to bridge 'Lastname Firstname' vs
    'Firstname Lastname' parser conventions."""
    parts = name.split()
    if len(parts) < 2:
        return []
    if len(parts) == 2:
        return [f"{parts[1]} {parts[0]}"]
    if len(parts) == 3:
        return [
            f"{parts[2]} {parts[0]} {parts[1]}",
            f"{parts[1]} {parts[2]} {parts[0]}",
        ]
    if len(parts) == 4:
        return [
            f"{parts[3]} {parts[0]} {parts[1]} {parts[2]}",
            f"{parts[2]} {parts[3]} {parts[0]} {parts[1]}",
        ]
    return []


_ROSTER_LOOKUP_SQL = """
SELECT p.id, p.canonical_name, p.gender, r.mu, r.sigma, r.n_matches,
    (
        SELECT pta.class_label
        FROM player_team_assignments pta
        JOIN tournaments t ON t.id = pta.tournament_id
        WHERE pta.player_id = p.id
        ORDER BY t.year DESC, t.id DESC LIMIT 1
    ) AS captain_class,
    (
        SELECT MAX(m.played_on)
        FROM rating_history rh
        JOIN matches m ON m.id = rh.match_id
        WHERE rh.player_id = p.id AND rh.model_name = ?
    ) AS last_played
FROM players p
LEFT JOIN ratings r ON r.player_id = p.id AND r.model_name = ?
WHERE p.merged_into_id IS NULL
"""


def _lookup_roster_player(conn: sqlite3.Connection, name: str) -> dict | None:
    """Resolve a roster name to a rated player.

    Tries: exact canonical match → exact alias match → word-order rotations
    against both canonical and alias tables. Returns None if all attempts
    miss.
    """
    canonical_clause = " AND LOWER(p.canonical_name) = LOWER(?)"
    alias_clause = (
        " AND p.id IN (SELECT player_id FROM player_aliases "
        " WHERE LOWER(raw_name) = LOWER(?)) LIMIT 1"
    )

    def _try(target: str) -> tuple | None:
        row = conn.execute(
            _ROSTER_LOOKUP_SQL + canonical_clause,
            (MODEL, MODEL, target),
        ).fetchone()
        if row:
            return row
        return conn.execute(
            _ROSTER_LOOKUP_SQL + alias_clause,
            (MODEL, MODEL, target),
        ).fetchone()

    row = _try(name)
    if row is None:
        for variant in _name_order_variants(name):
            row = _try(variant)
            if row:
                break
    if row is None:
        return None

    pid, canonical, gender, mu, sigma, n, cls, last = row
    return {
        "roster_name": name,
        "id": pid,
        "canonical": canonical,
        "gender": gender,
        "mu": mu,
        "sigma": sigma,
        "n": n,
        "class": cls,
        "last": last,
    }


def _fuzzy_candidates(conn: sqlite3.Connection, name: str, limit: int = 3) -> list[str]:
    all_names = [
        r[0]
        for r in conn.execute(
            "SELECT canonical_name FROM players WHERE merged_into_id IS NULL"
        ).fetchall()
    ]
    return difflib.get_close_matches(name, all_names, n=limit, cutoff=0.6)


def _proposed_class_label(rank_idx: int, group_size: int = 6, slots_per_tier: int = 4) -> str:
    """Class label (A1, A2, ..., A4, B1, ..., B4, C1, ...) from 0-indexed rank.

    Mirrors the existing player_team_assignments convention: each tier
    (A/B/C/D...) holds `slots_per_tier` slots, and `group_size` players share
    a slot (one per team). With 6 captains and 4 slots per tier, every 6
    positions advance the slot number, and every 24 positions advance the
    tier letter.
    """
    slot_idx = rank_idx // group_size
    tier_idx = slot_idx // slots_per_tier
    within_tier = (slot_idx % slots_per_tier) + 1
    tier_letter = chr(ord("A") + tier_idx)
    return f"{tier_letter}{within_tier}"


def _load_captain_rankings(
    conn: sqlite3.Connection, configs: list[dict]
) -> list[dict]:
    """Load each captain's ranking JSON, resolve names against the DB.

    Returns one dict per captain with:
      - label: short column heading
      - <gender>_by_pid: {player_id: rank}  (1-based; for matched roster players)
      - <gender>_by_name: {lowercased_name: rank}  (covers debutants)
      - <gender>_unresolved: [(rank, name)]  (didn't match any DB player)
    where <gender> is 'men' or 'ladies'.
    """
    out: list[dict] = []
    for cfg in configs:
        path = Path(cfg["json_path"])
        if not path.exists():
            print(
                f"  captain ranking not found: {path}",
                file=sys.stderr,
            )
            continue
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except Exception as e:
            print(f"  captain ranking JSON parse error ({path}): {e}", file=sys.stderr)
            continue
        resolved: dict = {"label": cfg["label"]}
        for sheet_key in ("men", "ladies"):
            names = data.get(sheet_key, []) or []
            by_pid: dict[int, int] = {}
            by_name: dict[str, int] = {}
            unresolved: list[tuple[int, str]] = []
            for i, name in enumerate(names, 1):
                if not isinstance(name, str):
                    continue
                # Index under the captain's spelling AND any word-order
                # rotation so a roster name in the opposite order
                # ("Polidano Kyle" vs "Kyle Polidano") still finds it.
                by_name[name.strip().lower()] = i
                for variant in _name_order_variants(name.strip()):
                    by_name.setdefault(variant.lower(), i)
                row = _lookup_roster_player(conn, name)
                if row is not None:
                    by_pid[row["id"]] = i
                else:
                    unresolved.append((i, name))
            resolved[f"{sheet_key}_by_pid"] = by_pid
            resolved[f"{sheet_key}_by_name"] = by_name
            resolved[f"{sheet_key}_unresolved"] = unresolved
        out.append(resolved)
    return out


def _captain_rank_cell(rank: int | None) -> str:
    """One <td> for a captain rank — blank slot is sortable to the end."""
    if rank is None:
        # data-sort = large sentinel pushes blanks to the end on ascending sort.
        return '<td class="num muted" data-sort="999999">—</td>'
    return f'<td class="num" data-sort="{rank}">{rank}</td>'


def _render_roster_section(
    sheet: str,
    hits: list[dict],
    misses: list[tuple[str, list[str]]],
    captain_rankings: list[dict] | None = None,
    decay_ratings_by_pid: dict[int, tuple[float, float]] | None = None,
) -> str:
    rated = sorted(
        (h for h in hits if h["mu"] is not None),
        key=lambda h: -(h["mu"] - 3 * h["sigma"]),
    )
    unrated = [h for h in hits if h["mu"] is None]
    total = len(rated) + len(unrated) + len(misses)

    # Per-section decay rank: sort the rated pool by decay μ-3σ within
    # this gender section and map player_id → rank. Apples-to-apples
    # comparison within the same field of competitors.
    decay_rank_by_pid: dict[int, int] = {}
    decay_ratings_by_pid = decay_ratings_by_pid or {}
    if decay_ratings_by_pid:
        decay_ranked = [h for h in rated if h["id"] in decay_ratings_by_pid]
        decay_ranked.sort(
            key=lambda h: -(decay_ratings_by_pid[h["id"]][0]
                            - 3 * decay_ratings_by_pid[h["id"]][1])
        )
        for i, h in enumerate(decay_ranked, 1):
            decay_rank_by_pid[h["id"]] = i

    pills = [f'<span class="pill ok">{len(rated)} rated</span>']
    if unrated:
        pills.append(f'<span class="pill warn">{len(unrated)} unrated</span>')
    if misses:
        pills.append(f'<span class="pill miss">{len(misses)} not in DB</span>')
    pills.append(f'<span class="pill">total {total}</span>')

    captain_rankings = captain_rankings or []
    sheet_key = "men" if sheet.lower().startswith("men") else "ladies"

    rows = []
    for i, h in enumerate(rated, 1):
        cons = h["mu"] - 3 * h["sigma"]
        proposed_cls = _proposed_class_label(i - 1)
        prev_cls = h["class"]
        cls_title = (
            f'previous tournament class: {prev_cls}' if prev_cls else
            'no prior captain class on record'
        )
        link = (
            f'<a class="player-link" href="../players/{h["id"]}.html">'
            f'{esc(h["canonical"])}</a>'
        )
        if h["roster_name"].strip().lower() != h["canonical"].strip().lower():
            link += (
                f' <span class="muted" title="roster name">'
                f'({esc(h["roster_name"])})</span>'
            )
        captain_cells = "".join(
            _captain_rank_cell(c[f"{sheet_key}_by_pid"].get(h["id"]))
            for c in captain_rankings
        )
        decay_cell = (
            _captain_rank_cell(decay_rank_by_pid.get(h["id"]))
            if decay_ratings_by_pid else ""
        )
        rows.append(
            f'<tr>'
            f'<td class="num" data-sort="{i}">{i}</td>'
            f'<td class="cls" title="{esc(cls_title)}">{esc(proposed_cls)}</td>'
            f'<td>{link}</td>'
            f'<td class="num" data-sort="{cons:.4f}"><strong>{cons:.2f}</strong></td>'
            f'{decay_cell}'
            f'<td class="num" data-sort="{h["mu"]:.4f}">{h["mu"]:.2f}</td>'
            f'<td class="num" data-sort="{h["sigma"]:.4f}">{h["sigma"]:.2f}</td>'
            f'<td class="num" data-sort="{h["n"]}">{h["n"]}</td>'
            f'<td data-sort="{esc(h["last"] or "")}">{esc(h["last"] or "?")}</td>'
            f'{captain_cells}'
            f'</tr>'
        )

    captain_headers = "".join(
        f'<th class="num" title="{esc(c["label"])}\'s pre-tournament rank">'
        f'{esc(c["label"])}</th>'
        for c in captain_rankings
    )
    decay_header = (
        '<th class="num" title="Time-decay challenger model rank '
        '(τ=365d). Recency-weighted: old matches contribute exponentially '
        'less. Backtest log-loss is 5.8% better than vanilla; see '
        '_ANALYSIS_/model_evaluation/SUMMARY.md.">Decay #</th>'
        if decay_ratings_by_pid else ""
    )
    rated_table = (
        f'<div class="table-wrap"><table class="sortable">'
        f'<thead><tr>'
        f'<th class="num">#</th><th>Class</th><th>Player</th>'
        f'<th class="num">μ-3σ</th>'
        f'{decay_header}'
        f'<th class="num">μ</th><th class="num">σ</th>'
        f'<th class="num">n</th><th>Last played</th>'
        f'{captain_headers}'
        f'</tr></thead>'
        f'<tbody>{"".join(rows)}</tbody>'
        f'</table></div>'
    ) if rated else ""

    miss_html = ""
    if misses:
        items = []
        for n, suggestions in misses:
            captain_chips = " ".join(
                f'<span class="muted" style="margin-left:6px;">'
                f'{esc(c["label"])} #{c[f"{sheet_key}_by_name"][n.strip().lower()]}'
                f'</span>'
                for c in captain_rankings
                if n.strip().lower() in c[f"{sheet_key}_by_name"]
            )
            sugg = (
                ", ".join(esc(s) for s in suggestions)
                if suggestions
                else '<span class="muted">no close match</span>'
            )
            items.append(
                f'<li>{esc(n)}{captain_chips} '
                f'<span class="muted">— suggest: {sugg}</span></li>'
            )
        miss_html = (
            f'<h3 class="section-title">Not yet in our system ({len(misses)})</h3>'
            f'<ul class="plain">{"".join(items)}</ul>'
        )

    unrated_html = ""
    if unrated:
        items = "".join(f"<li>{esc(h['canonical'])}</li>" for h in unrated)
        unrated_html = (
            f'<h3 class="section-title">In DB but unrated ({len(unrated)})</h3>'
            f'<ul class="plain">{items}</ul>'
        )

    return (
        f'<section><div style="display:flex;align-items:baseline;gap:12px;'
        f'flex-wrap:wrap;margin:24px 0 8px;">'
        f'<h2 style="margin:0;">{esc(sheet)}</h2>'
        f'<span>{" ".join(pills)}</span>'
        f'</div>{rated_table}{unrated_html}{miss_html}</section>'
    )


_ROSTER_PAGE_CSS = """
.pill { display: inline-block; background: var(--card); border: 1px solid var(--border);
        padding: 2px 8px; border-radius: 999px; font-size: 12px; color: var(--muted); margin-right:4px;}
.pill.ok { color: var(--win); border-color: rgba(70,194,129,0.4); }
.pill.warn { color: #d6b46a; border-color: rgba(214,180,106,0.4); }
.pill.miss { color: var(--loss); border-color: rgba(224,122,122,0.4); }
ul.plain { list-style: none; padding: 0; margin: 0; columns: 2; column-gap: 24px; }
ul.plain li { padding: 4px 0; border-bottom: 1px solid var(--border); break-inside: avoid; }
table.sortable th { cursor: pointer; user-select: none; }
table.sortable th:hover { color: var(--fg); }
table.sortable th[aria-sort="ascending"]::after { content: " ▲"; opacity: 0.6; }
table.sortable th[aria-sort="descending"]::after { content: " ▼"; opacity: 0.6; }
@media (max-width: 700px) { ul.plain { columns: 1; } }
"""

_ROSTER_PAGE_JS = r"""
<script>
// Click-to-sort for any <table class="sortable">.
// Sort key: th-cell's data-sort attribute when present (numeric or string),
// otherwise textContent. Three-state: asc → desc → original.
document.addEventListener('DOMContentLoaded', () => {
  document.querySelectorAll('table.sortable').forEach((tbl) => {
    const tbody = tbl.querySelector('tbody');
    const originalRows = Array.from(tbody.querySelectorAll('tr'));
    tbl.querySelectorAll('thead th').forEach((th, idx) => {
      th.addEventListener('click', () => {
        const cur = th.getAttribute('aria-sort');
        // Reset all other headers
        tbl.querySelectorAll('thead th').forEach((o) => {
          if (o !== th) o.removeAttribute('aria-sort');
        });
        let dir;
        if (cur === 'ascending') dir = 'descending';
        else if (cur === 'descending') dir = 'none';
        else dir = 'ascending';
        if (dir === 'none') {
          th.removeAttribute('aria-sort');
          originalRows.forEach((r) => tbody.appendChild(r));
          return;
        }
        th.setAttribute('aria-sort', dir);
        const rows = Array.from(tbody.querySelectorAll('tr'));
        const numeric = th.classList.contains('num');
        const get = (row) => {
          const cell = row.cells[idx];
          if (cell.dataset.sort !== undefined) return cell.dataset.sort;
          return cell.textContent.trim();
        };
        rows.sort((a, b) => {
          const av = get(a);
          const bv = get(b);
          if (numeric) {
            const an = parseFloat(av);
            const bn = parseFloat(bv);
            if (!isNaN(an) && !isNaN(bn)) {
              return dir === 'ascending' ? an - bn : bn - an;
            }
          }
          return dir === 'ascending'
            ? String(av).localeCompare(String(bv))
            : String(bv).localeCompare(String(av));
        });
        rows.forEach((r) => tbody.appendChild(r));
      });
    });
  });
});
</script>
"""


def build_matches_page(
    conn: sqlite3.Connection,
    name_lookup: dict,
    impacts: dict | None = None,
) -> str:
    """Chronological feed of every active match — newest first.

    Single page with client-side year/club/text filter. Match sides are linked
    to player profiles. Winning side is bolded. Each player name carries a
    rank-at-the-time tag, and each row expands to show every player's
    rank/score change for that match.
    """
    impacts = impacts or {}
    rows = conn.execute(
        """
        SELECT m.id, m.played_on, m.division, m.round, m.walkover,
               t.name AS tour_name, t.year AS tour_year,
               c.slug AS club_slug,
               sa.player1_id, sa.player2_id, sa.games_won, sa.won,
               sb.player1_id, sb.player2_id, sb.games_won, sb.won
        FROM matches m
        JOIN tournaments t ON t.id = m.tournament_id
        JOIN clubs c ON c.id = t.club_id
        JOIN match_sides sa ON sa.match_id = m.id AND sa.side = 'A'
        JOIN match_sides sb ON sb.match_id = m.id AND sb.side = 'B'
        WHERE m.superseded_by_run_id IS NULL
        ORDER BY m.played_on DESC, m.id DESC
        """
    ).fetchall()

    # Bulk-fetch set scores once. Per-match SQL would be 5K queries.
    set_scores_by_mid: dict[int, list[tuple]] = {}
    for mid, sn, a, b, tb in conn.execute(
        "SELECT match_id, set_number, side_a_games, side_b_games, was_tiebreak "
        "FROM match_set_scores ORDER BY match_id, set_number"
    ):
        set_scores_by_mid.setdefault(mid, []).append((sn, a, b, tb))

    def _name_with_rank(pid: int, mid: int) -> str:
        """Linked player name + rank-at-time tag (from impacts)."""
        if pid is None:
            return ""
        cid, n = name_lookup.get(pid, (pid, f"#{pid}"))
        link = f'<a class="player-link" href="players/{cid}.html">{esc(n)}</a>'
        imp = impacts.get((mid, pid))
        if imp is None:
            return link
        return f'{link}<span class="rank-tag">#{imp["rank_after"]}</span>'

    def _render_set_score(side: str, sets: list) -> str:
        """Render '6-2, 6-1' from THIS side's perspective."""
        if not sets:
            return ""
        out = []
        for _sn, a, b, tb in sets:
            my, opp = (a, b) if side == "A" else (b, a)
            mark = " <span class='muted'>(TB)</span>" if tb else ""
            out.append(f"{my}-{opp}{mark}")
        return ", ".join(out)

    body_rows = []
    years: set = set()
    clubs: set = set()
    for r in rows:
        (mid, played, division, rnd, walkover,
         tour_name, tour_year, club_slug,
         a1, a2, agw, awon,
         b1, b2, bgw, bwon) = r
        years.add(tour_year)
        clubs.add(club_slug)

        sets = set_scores_by_mid.get(mid, [])
        score_a = _render_set_score("A", sets)
        score_b = _render_set_score("B", sets)
        # Resolve who won via match_result — handles tied rubbers (sets 1-1,
        # both won=0) by deferring to the games-won tiebreak, same as the
        # rating engine. a_won_eff is True iff side A is the effective winner.
        a_cls_eff, _, _ = match_result(awon or 0, bwon or 0, agw or 0, bgw or 0)
        a_won_eff = (a_cls_eff == "win")
        if score_a:
            score_cell = (
                f'<span class="score">'
                f'<strong class="{"win" if a_won_eff else "loss"}">{score_a}</strong>'
                f' / <span class="muted">{score_b}</span>'
                f'</span>'
            )
        else:
            score_cell = f'<span class="score muted">{agw or 0}-{bgw or 0}</span>'
        wo = ' <span class="tag">W/O</span>' if walkover else ""

        side_a_names = " / ".join(
            x for x in (_name_with_rank(a1, mid), _name_with_rank(a2, mid)) if x
        ) or "—"
        side_b_names = " / ".join(
            x for x in (_name_with_rank(b1, mid), _name_with_rank(b2, mid)) if x
        ) or "—"
        if a_won_eff:
            side_a_html = f'<strong class="win">{side_a_names}</strong>'
            side_b_html = f'<span class="muted">{side_b_names}</span>'
        else:
            side_b_html = f'<strong class="win">{side_b_names}</strong>'
            side_a_html = f'<span class="muted">{side_a_names}</span>'

        has_impact = any(
            (mid, p) in impacts for p in (a1, a2, b1, b2) if p is not None
        )
        trigger_cell = (
            f'<td class="expand-trigger" data-mid="{mid}">▶</td>'
            if has_impact
            else '<td class="muted" style="text-align:center;">·</td>'
        )

        body_rows.append(
            f'<tr class="match-row" data-mid="{mid}" '
            f'data-year="{esc(tour_year)}" data-club="{esc(club_slug)}">'
            f'{trigger_cell}'
            f'<td>{esc(played)}</td>'
            f'<td><span class="tag">{esc(club_slug)}</span> {esc(tour_name)}</td>'
            f'<td class="muted">{esc(division or "")} {esc(rnd or "")}</td>'
            f'<td>{side_a_html}</td>'
            f'<td>{side_b_html}</td>'
            f'<td class="score">{score_cell}{wo}</td>'
            f'</tr>'
        )

        if has_impact:
            # Pass label overrides so tied rubbers display "W (g)" / "L (g)"
            # in the per-player impact card. Resolution mirrors the rating
            # engine (see match_result).
            a_cls, a_label, _ = match_result(
                awon or 0, bwon or 0, agw or 0, bgw or 0
            )
            b_cls, b_label, _ = match_result(
                bwon or 0, awon or 0, bgw or 0, agw or 0
            )
            a_won_b = (a_cls == "win")
            b_won_b = (b_cls == "win")
            participants = [
                (a1, a2, "A", a_won_b, a_label),
                (a2, a1, "A", a_won_b, a_label),
                (b1, b2, "B", b_won_b, b_label),
                (b2, b1, "B", b_won_b, b_label),
            ]
            participants = [p for p in participants if p[0] is not None]
            impact_html = render_match_impact_block(
                mid, participants, impacts, name_lookup
            )
            body_rows.append(
                f'<tr class="impact-row" data-mid="{mid}" '
                f'data-year="{esc(tour_year)}" data-club="{esc(club_slug)}" hidden>'
                f'<td colspan="7">{impact_html}</td>'
                f'</tr>'
            )

    year_options = "".join(
        f'<option value="{y}">{y}</option>' for y in sorted(years, reverse=True)
    )
    club_options = "".join(
        f'<option value="{esc(c)}">{esc(c)}</option>' for c in sorted(clubs)
    )

    js = """
    <script>
    function applyFilters() {
      const y = document.getElementById('f-year').value;
      const c = document.getElementById('f-club').value;
      const q = document.getElementById('f-search').value.toLowerCase();
      const matchRows = document.querySelectorAll('tr.match-row');
      let visible = 0;
      matchRows.forEach(row => {
        const ok = (!y || row.dataset.year === y)
                && (!c || row.dataset.club === c)
                && (!q || row.textContent.toLowerCase().includes(q));
        row.style.display = ok ? '' : 'none';
        const mid = row.dataset.mid;
        const ir = document.querySelector('tr.impact-row[data-mid="' + mid + '"]');
        if (ir && !ok) {
          ir.hidden = true;
          row.classList.remove('open');
        }
        if (ok) visible++;
      });
      document.getElementById('count').textContent = visible.toLocaleString() + ' matches';
    }
    function toggleImpact(ev) {
      const cell = ev.target.closest('.expand-trigger');
      if (!cell) return;
      const mid = cell.dataset.mid;
      const row = cell.closest('tr.match-row');
      const ir = document.querySelector('tr.impact-row[data-mid="' + mid + '"]');
      if (!ir) return;
      ir.hidden = !ir.hidden;
      row.classList.toggle('open', !ir.hidden);
    }
    document.addEventListener('DOMContentLoaded', () => {
      ['f-year','f-club','f-search'].forEach(id =>
        document.getElementById(id).addEventListener('input', applyFilters));
      document.querySelector('table').addEventListener('click', toggleImpact);
      applyFilters();
    });
    </script>
    """

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1, viewport-fit=cover">
<meta name="theme-color" content="#0f1115">
<title>All matches — RallyRank</title>
<link rel="stylesheet" href="styles.css?v={CSS_VERSION}">
</head>
<body>
<header>
  <h1>All matches</h1>
  <p>Every active match across every loaded source file, newest first.
     Winning side is bolded. Numbers next to names are each player's rank
     in their gender bucket immediately AFTER that match. Click ▶ to see
     each player's rank/score change.</p>
</header>
{render_nav("", "matches")}
<main>
  <div class="controls">
    <input id="f-search" type="search" placeholder="Search player / tournament / division ...">
    <select id="f-year">
      <option value="">All years</option>
      {year_options}
    </select>
    <select id="f-club">
      <option value="">All clubs</option>
      {club_options}
    </select>
    <span id="count" class="muted"></span>
  </div>
  <div class="table-wrap">
  <table>
    <thead><tr>
      <th></th>
      <th>Date</th><th>Tournament</th><th>Round</th>
      <th>Side A</th><th>Side B</th><th>Score</th>
    </tr></thead>
    <tbody>{''.join(body_rows)}</tbody>
  </table>
  </div>
</main>
<footer>
  Total: {len(rows):,} active match(es). Superseded matches (from re-ingested files) are excluded.
</footer>
{js}
</body>
</html>
"""


def build_disagreements_page(
    conn: sqlite3.Connection,
    name_lookup: dict,
    predictions: dict[str, dict[int, dict]],
    impacts: dict | None = None,
    *,
    min_gap: float = 0.10,
    max_rows: int = 300,
) -> str:
    """Feed of matches where the two models disagreed most about who'd win."""
    impacts = impacts or {}
    decay = predictions.get("openskill_pl_decay365") or {}
    pl = predictions.get("openskill_pl_vanilla") or {}
    common_mids = set(decay.keys()) & set(pl.keys())
    if not common_mids:
        return ""

    placeholders = ",".join("?" * len(common_mids))
    rows = conn.execute(
        f"""
        SELECT m.id, m.played_on, m.division, m.round, m.walkover,
               t.name AS tour_name, t.year AS tour_year,
               c.slug AS club_slug,
               sa.player1_id, sa.player2_id, sa.games_won, sa.won,
               sb.player1_id, sb.player2_id, sb.games_won, sb.won
        FROM matches m
        JOIN tournaments t ON t.id = m.tournament_id
        JOIN clubs c ON c.id = t.club_id
        JOIN match_sides sa ON sa.match_id = m.id AND sa.side = 'A'
        JOIN match_sides sb ON sb.match_id = m.id AND sb.side = 'B'
        WHERE m.id IN ({placeholders})
          AND m.superseded_by_run_id IS NULL
        """,
        tuple(common_mids),
    ).fetchall()

    set_scores_by_mid: dict[int, list[tuple]] = {}
    for mid, sn, a, b, tb in conn.execute(
        "SELECT match_id, set_number, side_a_games, side_b_games, was_tiebreak "
        "FROM match_set_scores ORDER BY match_id, set_number"
    ):
        set_scores_by_mid.setdefault(mid, []).append((sn, a, b, tb))

    def _name_with_rank(pid: int, mid: int) -> str:
        if pid is None:
            return ""
        cid, n = name_lookup.get(pid, (pid, f"#{pid}"))
        link = f'<a class="player-link" href="players/{cid}.html">{esc(n)}</a>'
        imp = impacts.get((mid, pid))
        if imp is None:
            return link
        return f'{link}<span class="rank-tag">#{imp["rank_after"]}</span>'

    def _set_score(side: str, sets: list) -> str:
        if not sets:
            return ""
        out = []
        for _sn, a, b, tb in sets:
            my, opp = (a, b) if side == "A" else (b, a)
            mark = " <span class='muted'>(TB)</span>" if tb else ""
            out.append(f"{my}-{opp}{mark}")
        return ", ".join(out)

    ordered: list[tuple[float, tuple]] = []
    for r in rows:
        mid = r[0]
        gap = abs(decay[mid]["p_a"] - pl[mid]["p_a"])
        if gap >= min_gap:
            ordered.append((gap, r))
    ordered.sort(key=lambda x: -x[0])
    ordered = ordered[:max_rows]

    body_rows: list[str] = []
    years: set = set()
    clubs: set = set()
    n_decay_better = n_pl_better = n_both_right = n_both_wrong = 0
    for gap, r in ordered:
        (mid, played, division, rnd, walkover,
         tour_name, tour_year, club_slug,
         a1, a2, agw, awon,
         b1, b2, bgw, bwon) = r
        years.add(tour_year)
        clubs.add(club_slug)
        d_p = decay[mid]["p_a"]
        v_p = pl[mid]["p_a"]
        # Ground-truth verdict for model calibration. Tied rubbers (sets 1-1,
        # both won=0) are decided by the games-won tiebreak — same convention
        # as the rating engine. Without this, every tied match was scored as
        # "Side B won" because awon=0, which silently broke the model-gap
        # accuracy measurement.
        a_cls_dis, _, _ = match_result(
            awon or 0, bwon or 0, agw or 0, bgw or 0
        )
        actual_a = 1 if a_cls_dis == "win" else 0
        d_correct = (d_p > 0.5) == bool(actual_a)
        v_correct = (v_p > 0.5) == bool(actual_a)
        if d_correct and v_correct:
            n_both_right += 1
            verdict = '<span class="muted">both right</span>'
        elif d_correct:
            n_decay_better += 1
            verdict = '<span class="win"><strong>Decay</strong></span>'
        elif v_correct:
            n_pl_better += 1
            verdict = '<span class="win"><strong>PL</strong></span>'
        else:
            n_both_wrong += 1
            verdict = '<span class="loss">both wrong</span>'

        sets = set_scores_by_mid.get(mid, [])
        score_a = _set_score("A", sets)
        score_b = _set_score("B", sets)
        # actual_a (above) already reflects the games-tiebreak for tied
        # rubbers; reuse it so the bolded side matches the verdict column.
        if score_a:
            score_cell = (
                f'<span class="score">'
                f'<strong class="{"win" if actual_a else "loss"}">{score_a}</strong>'
                f' / <span class="muted">{score_b}</span>'
                f'</span>'
            )
        else:
            score_cell = f'<span class="score muted">{agw or 0}-{bgw or 0}</span>'
        wo = ' <span class="tag">W/O</span>' if walkover else ""

        side_a_names = " / ".join(
            x for x in (_name_with_rank(a1, mid), _name_with_rank(a2, mid)) if x
        ) or "—"
        side_b_names = " / ".join(
            x for x in (_name_with_rank(b1, mid), _name_with_rank(b2, mid)) if x
        ) or "—"
        body_rows.append(
            f'<tr data-year="{esc(str(tour_year))}" data-club="{esc(club_slug)}">'
            f'<td>{esc(played)}</td>'
            f'<td><span class="tag">{esc(club_slug)}</span> {esc(tour_name)} '
            f'{esc(str(tour_year))}</td>'
            f'<td>{side_a_names}</td>'
            f'<td>{side_b_names}</td>'
            f'<td class="score">{score_cell}{wo}</td>'
            f'<td class="num" data-sort="{v_p:.4f}">{v_p*100:.0f}%</td>'
            f'<td class="num" data-sort="{d_p:.4f}">{d_p*100:.0f}%</td>'
            f'<td class="num" data-sort="{gap:.4f}"><strong>{gap*100:.0f}%</strong></td>'
            f'<td class="num">{verdict}</td>'
            f'</tr>'
        )

    year_options = "".join(
        f'<option value="{y}">{y}</option>' for y in sorted(years, reverse=True)
    )
    club_options = "".join(
        f'<option value="{esc(c)}">{esc(c)}</option>' for c in sorted(clubs)
    )

    js = """
<script>
(function(){
  function applyFilters() {
    const q = (document.getElementById('f-search').value || '').toLowerCase();
    const y = document.getElementById('f-year').value;
    const c = document.getElementById('f-club').value;
    let visible = 0;
    document.querySelectorAll('tbody tr').forEach((row) => {
      const okY = !y || row.dataset.year === y;
      const okC = !c || row.dataset.club === c;
      const okQ = !q || row.textContent.toLowerCase().includes(q);
      const show = okY && okC && okQ;
      row.style.display = show ? '' : 'none';
      if (show) visible++;
    });
    document.getElementById('count').textContent = visible + ' match(es)';
  }
  document.addEventListener('DOMContentLoaded', () => {
    ['f-search','f-year','f-club'].forEach((id) => {
      document.getElementById(id).addEventListener('input', applyFilters);
    });
    document.querySelectorAll('thead th.num').forEach((th) => {
      th.style.cursor = 'pointer';
      th.addEventListener('click', () => {
        const tbody = document.querySelector('tbody');
        const rows = Array.from(tbody.querySelectorAll('tr'));
        const headers = Array.from(document.querySelectorAll('thead th'));
        const colIdx = headers.indexOf(th);
        const cur = th.getAttribute('aria-sort');
        const dir = cur === 'descending' ? 'ascending' : 'descending';
        headers.forEach((o) => { if (o !== th) o.removeAttribute('aria-sort'); });
        th.setAttribute('aria-sort', dir);
        rows.sort((a, b) => {
          const av = parseFloat(a.cells[colIdx].dataset.sort || '0');
          const bv = parseFloat(b.cells[colIdx].dataset.sort || '0');
          return dir === 'descending' ? bv - av : av - bv;
        });
        rows.forEach((r) => tbody.appendChild(r));
      });
    });
    applyFilters();
  });
})();
</script>
"""

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1, viewport-fit=cover">
<meta name="theme-color" content="#0f1115">
<title>Model gaps — RallyRank</title>
<link rel="stylesheet" href="styles.css?v={CSS_VERSION}">
</head>
<body>
<header>
  <h1>Model gaps — where the two models disagree</h1>
  <p>Matches sorted by how differently the two production rating models
     predicted them. Gap = |Decay% − PL%|. These are the matches where
     captain / human input most adds information the rating system
     can't reach on its own.</p>
</header>
{render_nav("", "disagreements")}
<main>
  <p class="muted" style="font-size:13px;">
    Top {len(body_rows):,} matches with gap ≥ {int(min_gap*100)}% (out of
    {len(common_mids):,} total predicted matches).
    Verdict tally on this filtered set:
    <strong>{n_decay_better}</strong> decay-only correct ·
    <strong>{n_pl_better}</strong> PL-only correct ·
    <strong>{n_both_right}</strong> both right ·
    <strong>{n_both_wrong}</strong> both wrong.
    Click any numeric column header to sort.
  </p>
  <div class="controls">
    <input id="f-search" type="search" placeholder="Search player / tournament ...">
    <select id="f-year">
      <option value="">All years</option>
      {year_options}
    </select>
    <select id="f-club">
      <option value="">All clubs</option>
      {club_options}
    </select>
    <span id="count" class="muted"></span>
  </div>
  <div class="table-wrap">
  <table>
    <thead><tr>
      <th>Date</th><th>Tournament</th>
      <th>Side A</th><th>Side B</th><th>Score</th>
      <th class="num" title="Vanilla PL probability that side A wins">PL %</th>
      <th class="num" title="Decay-365 probability that side A wins">Decay %</th>
      <th class="num" title="Disagreement = |Decay% − PL%|">|Δ|</th>
      <th class="num" title="Which model called this match correctly">Verdict</th>
    </tr></thead>
    <tbody>{''.join(body_rows)}</tbody>
  </table>
  </div>
</main>
<footer>
  Generated from {len(common_mids):,} matches with held-out predictions in
  both <code>openskill_pl_vanilla</code> and <code>openskill_pl_decay365</code>.
  See <a href="https://github.com/devkurtc/wks-social-tennis-rankings-malta/blob/main/_ANALYSIS_/model_evaluation/SUMMARY.md">SUMMARY.md</a>
  for backtest methodology.
</footer>
{js}
</body>
</html>
"""


def build_tournament_roster_page(conn: sqlite3.Connection, config: dict) -> str:
    """Render one tournament roster ranking page."""
    roster_path = Path(config["roster_xlsx"])
    if not roster_path.exists():
        return ""  # silently skip — caller logs the miss
    roster = _read_roster_xlsx(roster_path)

    captain_rankings = _load_captain_rankings(
        conn, config.get("captain_rankings", []) or []
    )
    # Surface unresolved captain names so the user can fix typos in the JSON.
    for c in captain_rankings:
        for sheet_key in ("men", "ladies"):
            for rank, name in c.get(f"{sheet_key}_unresolved", []):
                print(
                    f"  captain '{c['label']}' {sheet_key} #{rank}: "
                    f"{name!r} did not resolve to a DB player"
                    " (will still display by name in 'Not yet in system' if "
                    "the roster lists this person).",
                    file=sys.stderr,
                )

    # Load the time-decay challenger ratings if available. The recompute
    # for this model is run separately; if no rows exist yet, the column
    # is just omitted from the page.
    decay_ratings_by_pid: dict[int, tuple[float, float]] = {
        row[0]: (row[1], row[2])
        for row in conn.execute(
            "SELECT player_id, mu, sigma FROM ratings "
            "WHERE model_name = 'openskill_pl_decay365'"
        ).fetchall()
    }

    sections: list[str] = []
    totals = {"rated": 0, "unrated": 0, "miss": 0, "all": 0}
    for sheet, names in roster.items():
        hits: list[dict] = []
        misses: list[tuple[str, list[str]]] = []
        for n in names:
            r = _lookup_roster_player(conn, n)
            if r is None:
                misses.append((n, _fuzzy_candidates(conn, n)))
            else:
                hits.append(r)
        sections.append(
            _render_roster_section(
                sheet, hits, misses, captain_rankings,
                decay_ratings_by_pid=decay_ratings_by_pid,
            )
        )
        totals["rated"] += sum(1 for h in hits if h["mu"] is not None)
        totals["unrated"] += sum(1 for h in hits if h["mu"] is None)
        totals["miss"] += len(misses)
        totals["all"] += len(names)

    # DB summary for the legend
    last = conn.execute(
        "SELECT MAX(played_on) FROM matches WHERE superseded_by_run_id IS NULL"
    ).fetchone()[0]
    n_matches = conn.execute(
        "SELECT COUNT(*) FROM matches WHERE superseded_by_run_id IS NULL"
    ).fetchone()[0]
    n_tour = conn.execute(
        "SELECT COUNT(DISTINCT tournament_id) FROM matches WHERE superseded_by_run_id IS NULL"
    ).fetchone()[0]

    nav = render_nav("../", f"tournament:{config['slug']}")
    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1, viewport-fit=cover">
<meta name="theme-color" content="#0f1115">
<title>{esc(config['title'])} — RallyRank</title>
<link rel="stylesheet" href="../styles.css?v={CSS_VERSION}">
<style>{_ROSTER_PAGE_CSS}</style>
</head>
<body>
<header>
  <h1>{esc(config['title'])}</h1>
  <p>{esc(config.get('subtitle', ''))}</p>
</header>
{nav}
<main>
  <p class="muted" style="font-size:13px;">
    Source: <code>{esc(config['roster_xlsx'])}</code> &middot;
    rated against <strong>{n_matches}</strong> matches across
    <strong>{n_tour}</strong> tournaments
    (last match: <strong>{esc(last or '?')}</strong>) &middot;
    model <code>{esc(MODEL)}</code> &middot;
    coverage: <strong>{totals['rated']}/{totals['all']}</strong> rated,
    <strong>{totals['miss']}</strong> not yet in system.
  </p>
  {''.join(sections)}
</main>
<footer>
  Click any name to open that player's profile (μ trajectory, match history, peers).
  Click any column header to sort. Players in &ldquo;Not yet in our system&rdquo;
  will appear after their first rated match is loaded.
</footer>
{_ROSTER_PAGE_JS}
</body>
</html>
"""


# --- Changelog page ----------------------------------------------------------


CHANGELOG_JSON = PROJECT_ROOT / "scripts" / "phase0" / "changelog_entries.json"

GITHUB_REPO_URL = "https://github.com/devkurtc/wks-social-tennis-rankings-malta"


_MONTHS_LONG = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]
_MONTHS_SHORT = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                 "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def _parse_iso_ts(ts: str):
    """Parse an ISO 8601 timestamp like '2026-04-26T12:27:00+02:00'.

    Returns a timezone-aware `datetime`, or `None` on parse failure.
    """
    from datetime import datetime
    if not ts:
        return None
    try:
        return datetime.fromisoformat(ts)
    except ValueError:
        return None


def _format_month_from_ts(ts: str) -> str:
    """'2026-04-26T...' -> 'April 2026'."""
    dt = _parse_iso_ts(ts)
    if dt is None:
        return ts or ""
    return f"{_MONTHS_LONG[dt.month - 1]} {dt.year}"


def _format_date_time_malta(ts: str) -> str:
    """'2026-04-26T12:27:00+02:00' -> '26 Apr 2026 · 12:27'.

    The timestamp string already encodes the Malta offset; we render the
    wall-clock time as written (no further conversion needed).
    """
    dt = _parse_iso_ts(ts)
    if dt is None:
        return ts or ""
    return f"{dt.day} {_MONTHS_SHORT[dt.month - 1]} {dt.year} · {dt.hour:02d}:{dt.minute:02d}"


def build_changelog_page() -> str | None:
    """Render site/changelog.html from changelog_entries.json.

    Returns None (and prints a warning) if the JSON is missing or malformed.
    """
    if not CHANGELOG_JSON.exists():
        print(f"  skipped changelog: {CHANGELOG_JSON} not found", file=sys.stderr)
        return None
    try:
        data = json.loads(CHANGELOG_JSON.read_text(encoding="utf-8"))
    except json.JSONDecodeError as e:
        print(f"  skipped changelog: invalid JSON — {e}", file=sys.stderr)
        return None

    entries = data.get("entries", [])
    if not entries:
        print("  skipped changelog: no entries in JSON", file=sys.stderr)
        return None

    # Validate: warn on missing/unparseable timestamps (they sort to bottom).
    for e in entries:
        if _parse_iso_ts(e.get("timestamp", "")) is None:
            print(
                f"  changelog warning: entry '{e.get('id','?')}' has missing or "
                f"unparseable timestamp '{e.get('timestamp','')}'",
                file=sys.stderr,
            )

    # Newest first. Stable sort + reverse=True preserves JSON insertion order
    # for entries with identical timestamps (e.g. multiple ELI5 entries derived
    # from the same commit), so the JSON file controls within-tie ordering.
    from datetime import datetime, timezone
    far_past = datetime(1970, 1, 1, tzinfo=timezone.utc)
    entries_sorted = sorted(
        entries,
        key=lambda e: _parse_iso_ts(e.get("timestamp", "")) or far_past,
        reverse=True,
    )

    # Group by month-of-timestamp.
    grouped: list[tuple[str, list[dict]]] = []
    current_month: str | None = None
    current_bucket: list[dict] = []
    for e in entries_sorted:
        m = _format_month_from_ts(e.get("timestamp", ""))
        if m != current_month:
            if current_bucket:
                grouped.append((current_month or "", current_bucket))
            current_month = m
            current_bucket = []
        current_bucket.append(e)
    if current_bucket:
        grouped.append((current_month or "", current_bucket))

    parts: list[str] = []
    for month, bucket in grouped:
        parts.append(f'<h2 class="month-heading">{esc(month)}</h2>')
        for e in bucket:
            kind = (e.get("kind") or "new").lower()
            audience = (e.get("audience") or "all").lower()
            kind_label = {"new": "New", "improved": "Improved", "fixed": "Fixed"}.get(
                kind, kind.title()
            )
            head_pills = [
                f'<span class="pill kind-{esc(kind)}">{esc(kind_label)}</span>',
            ]
            if audience in ("captains", "admins"):
                aud_label = "For captains" if audience == "captains" else "For admins"
                head_pills.append(
                    f'<span class="pill aud-{esc(audience)}">{esc(aud_label)}</span>'
                )
            head_pills.append(
                f'<span class="entry-date">{esc(_format_date_time_malta(e.get("timestamp","")))}</span>'
            )

            details_html = ""
            if e.get("details"):
                details_html = (
                    "<details>"
                    "<summary>More detail</summary>"
                    f"<p>{esc(e['details'])}</p>"
                    "</details>"
                )

            trace_bits: list[str] = []
            for cm in e.get("commits") or []:
                short = esc(str(cm)[:7])
                trace_bits.append(
                    f'<a href="{GITHUB_REPO_URL}/commit/{esc(cm)}" '
                    f'target="_blank" rel="noopener">{short}</a>'
                )
            for tk in e.get("tasks") or []:
                trace_bits.append(esc(tk))
            trace_html = ""
            if trace_bits:
                trace_html = (
                    f'<div class="entry-trace">tech ref: '
                    + " · ".join(trace_bits)
                    + "</div>"
                )

            parts.append(
                f'<article class="entry" data-kind="{esc(kind)}" '
                f'data-audience="{esc(audience)}">'
                f'<div class="entry-head">{"".join(head_pills)}</div>'
                f'<h3 class="entry-title">{esc(e.get("title",""))}</h3>'
                f'<p class="entry-summary">{esc(e.get("summary",""))}</p>'
                f'{details_html}'
                f'{trace_html}'
                f'</article>'
            )

    cards_html = "\n".join(parts)
    nav_html = render_nav("", "changelog")

    return f"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>What's new — RallyRank</title>
<link rel="stylesheet" href="styles.css?v={CSS_VERSION}">
</head>
<body>
{nav_html}
<header>
  <h1>What’s new on RallyRank</h1>
  <p>Recent additions, improvements, and fixes — in plain English. Use the filters below to narrow the list.</p>
</header>

<main class="changelog">
  <p class="lead">Most recent at the top. All times shown in Malta local time. Click <em>More detail</em> on any entry for the longer story.</p>

  <div class="filters" id="changelog-filters">
    <span class="filter active" data-filter-kind="all">All</span>
    <span class="filter" data-filter-kind="new">New features</span>
    <span class="filter" data-filter-kind="improved">Improvements</span>
    <span class="filter" data-filter-kind="fixed">Fixes</span>
  </div>

  <div id="changelog-list">
{cards_html}
  </div>
  <div id="changelog-empty" class="empty" style="display:none;">
    Nothing to show with the current filter.
  </div>
</main>

<script>
(function() {{
  const list = document.getElementById('changelog-list');
  const emptyMsg = document.getElementById('changelog-empty');
  const filterEls = document.querySelectorAll('#changelog-filters .filter');
  let activeKind = 'all';

  function apply() {{
    const cards = list.querySelectorAll('.entry');
    let visible = 0;
    let lastVisibleHeading = null;
    list.querySelectorAll('.month-heading').forEach((h) => {{
      h.dataset.hasVisible = '0';
    }});
    cards.forEach((c) => {{
      const k = c.dataset.kind;
      const ok = activeKind === 'all' || k === activeKind;
      c.classList.toggle('hidden', !ok);
      if (ok) {{
        visible++;
        // mark previous-sibling month heading as having visible content
        let p = c.previousElementSibling;
        while (p) {{
          if (p.classList && p.classList.contains('month-heading')) {{
            p.dataset.hasVisible = '1';
            break;
          }}
          p = p.previousElementSibling;
        }}
      }}
    }});
    list.querySelectorAll('.month-heading').forEach((h) => {{
      h.style.display = h.dataset.hasVisible === '1' ? '' : 'none';
    }});
    emptyMsg.style.display = visible === 0 ? '' : 'none';
  }}

  filterEls.forEach((el) => {{
    el.addEventListener('click', () => {{
      filterEls.forEach((x) => x.classList.remove('active'));
      el.classList.add('active');
      activeKind = el.dataset.filterKind;
      apply();
    }});
  }});

  apply();
}})();
</script>
</body>
</html>
"""


# --- Mapping transparency page ----------------------------------------------


def _merge_kind_from_reason(reason: str) -> str:
    """Bucket a merge's `reason` string into a stable kind label.

    The reason text is human-written (set by the merger that fired) and
    starts with a small set of fixed prefixes — we match on those. Used for
    the filter pills + count summary on the Mapping page.
    """
    r = (reason or "").lower()
    if r.startswith("manual alias"):
        return "manual"
    if r.startswith("typo auto-merge"):
        return "typo"
    if r.startswith("token-equivalent"):
        return "token"
    if r.startswith("case-only"):
        return "case"
    return "other"


# All merges, newest first. Joins the audit_log row to the surviving
# (winner) record so we can render a working link straight to its page.
ALL_MERGES_SQL = """
SELECT
    al.id              AS audit_id,
    al.ts              AS merged_at,
    al.entity_id       AS loser_id,
    al.before_jsonb    AS before_json,
    al.after_jsonb     AS after_json,
    json_extract(al.after_jsonb, '$.merged_into_id')   AS winner_id,
    json_extract(al.after_jsonb, '$.reason')            AS reason
FROM audit_log al
WHERE al.action = 'player.merged' AND al.entity_type = 'players'
ORDER BY al.ts DESC, al.id DESC
"""


def build_aliases_page(conn: sqlite3.Connection, name_lookup: dict) -> str:
    """Render site/aliases.html — full identity-resolution transparency.

    Three sections:
      1. Stats — totals + per-kind breakdown.
      2. Merge log — every player.merged audit row, deep-linked by audit id.
      3. Pending suggestions — live snapshot of fuzzy candidates that the
         automated rules left for human review.
    """
    import json as _json
    # Local import — players is in the same dir as this script; sys.path
    # is configured by the CLI but not by direct script invocation, so be
    # defensive.
    import sys as _sys
    _sys.path.insert(0, str(Path(__file__).parent))
    import players as _players  # noqa: E402

    # ---- Section 1: Stats ----
    n_players_active = conn.execute(
        "SELECT COUNT(*) FROM players WHERE merged_into_id IS NULL"
    ).fetchone()[0]
    n_players_merged = conn.execute(
        "SELECT COUNT(*) FROM players WHERE merged_into_id IS NOT NULL"
    ).fetchone()[0]
    n_aliases = conn.execute("SELECT COUNT(*) FROM player_aliases").fetchone()[0]

    merge_rows = conn.execute(ALL_MERGES_SQL).fetchall()
    n_merges = len(merge_rows)
    by_kind: dict[str, int] = {"case": 0, "token": 0, "typo": 0, "manual": 0, "other": 0}
    for r in merge_rows:
        by_kind[_merge_kind_from_reason(r[6] or "")] += 1

    stats_html = f"""
  <div class="stat-grid" style="margin-bottom:18px;">
    <div class="stat"><div class="stat-label">Active players</div>
      <div class="stat-value">{n_players_active:,}</div></div>
    <div class="stat"><div class="stat-label">Records merged</div>
      <div class="stat-value">{n_players_merged:,}</div></div>
    <div class="stat"><div class="stat-label">Total merges</div>
      <div class="stat-value">{n_merges:,}</div></div>
    <div class="stat"><div class="stat-label">Raw alias forms</div>
      <div class="stat-value">{n_aliases:,}</div></div>
  </div>
  <div class="muted" style="font-size:12px; margin: -10px 0 18px 0;">
    By kind:
    <span class="tag" data-kind-tally="case">case · {by_kind['case']}</span>
    <span class="tag" data-kind-tally="token">token · {by_kind['token']}</span>
    <span class="tag" data-kind-tally="typo">typo · {by_kind['typo']}</span>
    <span class="tag" data-kind-tally="manual">manual · {by_kind['manual']}</span>
    {"<span class='tag' data-kind-tally='other'>other · " + str(by_kind['other']) + "</span>" if by_kind['other'] else ""}
  </div>
"""

    # ---- Section 2: Merge log ----
    merge_log_rows: list[str] = []
    for audit_id, merged_at, loser_id, before_json, after_json, winner_id, reason in merge_rows:
        try:
            before = _json.loads(before_json or "{}")
            after = _json.loads(after_json or "{}")
        except _json.JSONDecodeError:
            before, after = {}, {}
        loser_name = before.get("canonical_name", f"#{loser_id}")
        winner_name = after.get("winner_canonical_name", f"#{winner_id}")
        kind = _merge_kind_from_reason(reason or "")
        merged_on = (merged_at or "")[:10]
        # Winner page link (loser pages aren't generated; the loser is
        # a "ghost" whose history now lives under the winner).
        winner_link = (
            f'<a class="player-link" href="players/{int(winner_id)}.html">{esc(winner_name)}</a>'
            if winner_id is not None else esc(winner_name)
        )
        merge_log_rows.append(
            f'<tr id="m-{audit_id}" data-kind="{kind}" '
            f'data-search="{esc((loser_name + " " + winner_name + " " + (reason or "")).lower())}">'
            f'<td class="muted" style="font-variant-numeric:tabular-nums;">'
            f'<a href="#m-{audit_id}" class="muted" title="Permalink">#{audit_id}</a></td>'
            f'<td class="muted">{esc(merged_on)}</td>'
            f'<td><span class="kind-pill kind-{kind}">{kind}</span></td>'
            f'<td><span class="tag">id #{loser_id}</span> {esc(loser_name)}</td>'
            f'<td>→</td>'
            f'<td>{winner_link} <span class="tag">id #{winner_id or "?"}</span></td>'
            f'<td class="muted">{esc(reason or "")}</td>'
            f'</tr>'
        )

    merge_log_html = "".join(merge_log_rows) or (
        '<tr><td colspan="7" class="muted">No merges yet.</td></tr>'
    )

    # ---- Section 3: Pending fuzzy suggestions ----
    # Snapshot at build time. Use the same threshold/gates as the CLI's
    # default `suggest-merges` so the site mirrors what an admin would see.
    # Filter out pairs already ruled "different people" via the review tools.
    kd_path = Path(__file__).resolve().parent / "known_distinct.json"
    suggestions = _players.suggest_fuzzy_matches(
        conn,
        threshold=0.85,
        same_gender_only=True,
        min_matches=1,
        known_distinct=_players.load_known_distinct(str(kd_path)),
    )

    BUCKETS = [
        ("very-high", "VERY HIGH", 0.95, 1.01,
         "Auto-merge candidates — usually safe to add to manual_aliases.json"),
        ("high",      "HIGH",      0.88, 0.95,
         "Almost certainly the same person — quick glance"),
        ("medium",    "MEDIUM",    0.78, 0.88,
         "Needs human review — not obvious"),
        ("low",       "LOW",       0.00, 0.78,
         "Probably different — but flagged"),
    ]

    bucket_rows: dict[str, list[str]] = {b[0]: [] for b in BUCKETS}
    for s in suggestions:
        c = s["confidence"]
        bucket_key = next(
            (k for (k, _, lo, hi, _) in BUCKETS if lo <= c < hi),
            "low",
        )
        a = s["a"]; b = s["b"]
        signals = " · ".join(s.get("reasons") or [])
        a_link = f'<a class="player-link" href="players/{a["id"]}.html">{esc(a["name"])}</a>'
        b_link = f'<a class="player-link" href="players/{b["id"]}.html">{esc(b["name"])}</a>'
        bucket_rows[bucket_key].append(
            f'<tr data-search="{esc((a["name"] + " " + b["name"]).lower())}">'
            f'<td class="num">{c:.2f}</td>'
            f'<td>{a_link} <span class="muted">({a.get("n", 0)}m'
            + (f", {a.get('latest_class')}" if a.get("latest_class") else "")
            + f')</span></td>'
            f'<td class="muted">vs</td>'
            f'<td>{b_link} <span class="muted">({b.get("n", 0)}m'
            + (f", {b.get('latest_class')}" if b.get("latest_class") else "")
            + f')</span></td>'
            f'<td class="muted" style="white-space:normal; font-size:11px;">{esc(signals)}</td>'
            f'</tr>'
        )

    suggestion_sections: list[str] = []
    for key, label, lo, hi, hint in BUCKETS:
        rows = bucket_rows[key]
        rng = f"{lo:.2f}+" if hi > 1.0 else f"{lo:.2f}–{hi:.2f}"
        body = "".join(rows) if rows else (
            f'<tr><td colspan="5" class="muted">No pairs in this bucket.</td></tr>'
        )
        suggestion_sections.append(f"""
  <h3 class="bucket-heading bucket-{key}">{label}
    <span class="muted">({len(rows)} pair{'s' if len(rows) != 1 else ''}, conf {rng})</span></h3>
  <p class="muted" style="font-size:12px; margin-top:-4px;">{esc(hint)}</p>
  <div class="table-wrap">
  <table>
    <thead><tr>
      <th class="num">Conf</th>
      <th>Player A</th>
      <th></th>
      <th>Player B</th>
      <th>Signals</th>
    </tr></thead>
    <tbody>{body}</tbody>
  </table>
  </div>
""")

    nav_html = render_nav("", "aliases")

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1, viewport-fit=cover">
<meta name="theme-color" content="#0f1115">
<title>Mapping &amp; merges — RallyRank</title>
<link rel="stylesheet" href="styles.css?v={CSS_VERSION}">
<style>
  .kind-pill {{
    display: inline-block; padding: 1px 6px; border-radius: 4px;
    font-size: 11px; font-weight: 600; text-transform: uppercase;
    letter-spacing: 0.04em;
  }}
  .kind-case   {{ background: #25344f; color: #9bc1ff; }}
  .kind-token  {{ background: #2c3e2c; color: #9fd8a4; }}
  .kind-typo   {{ background: #3e3520; color: #ffd58a; }}
  .kind-manual {{ background: #3a2a3e; color: #d99bd0; }}
  .kind-other  {{ background: #2a3242; color: #8b96a8; }}
  .bucket-heading {{
    margin: 18px 0 4px 0; font-size: 14px; font-weight: 600;
  }}
  .bucket-very-high {{ color: #ffd58a; }}
  .bucket-high      {{ color: #9fd8a4; }}
  .bucket-medium    {{ color: #9bc1ff; }}
  .bucket-low       {{ color: #8b96a8; }}
  .filter-pills {{
    display: flex; gap: 6px; flex-wrap: wrap; margin: 8px 0;
  }}
  .filter-pills .pill {{
    cursor: pointer; user-select: none;
    padding: 4px 10px; border-radius: 6px; font-size: 12px;
    background: var(--card); color: var(--muted); border: 1px solid var(--border);
  }}
  .filter-pills .pill.active {{
    background: var(--accent); color: var(--bg); border-color: var(--accent);
  }}
  /* Target highlight when arriving via #m-<id> deep link */
  tr:target {{ background: #2a3a55 !important; outline: 2px solid var(--accent); }}
</style>
</head>
<body>
<header>
  <h1>Mapping &amp; merges</h1>
  <p>
    Every identity-resolution decision RallyRank has made — what got merged
    into whom, why, and what's still pending review. Permalink any merge by
    clicking its <code>#id</code>.
  </p>
</header>
{nav_html}
<main>
{stats_html}

  <h2 class="section-title">Merge log</h2>
  <p class="muted" style="font-size:12px; margin-top:-4px;">
    Newest first. Each merge captures: which losing record was absorbed, into
    which surviving record, on what date, and the rule that fired.
  </p>
  <div class="filter-pills" id="merge-filters">
    <span class="pill active" data-kind="all">All ({n_merges})</span>
    <span class="pill" data-kind="case">case ({by_kind['case']})</span>
    <span class="pill" data-kind="token">token ({by_kind['token']})</span>
    <span class="pill" data-kind="typo">typo ({by_kind['typo']})</span>
    <span class="pill" data-kind="manual">manual ({by_kind['manual']})</span>
    <input id="merge-search" type="search" placeholder="Search names / reason..." style="margin-left:auto; flex:1 1 200px;">
  </div>
  <div class="table-wrap">
  <table id="merge-log">
    <thead><tr>
      <th>#</th>
      <th>Date</th>
      <th>Kind</th>
      <th>Loser (absorbed)</th>
      <th></th>
      <th>Winner (surviving)</th>
      <th>Reason</th>
    </tr></thead>
    <tbody>{merge_log_html}</tbody>
  </table>
  </div>
  <div id="merge-empty" class="muted" style="display:none; padding:8px;">
    No merges match the current filter.
  </div>

  <h2 class="section-title" style="margin-top:32px;">Pending fuzzy suggestions</h2>
  <p class="muted" style="font-size:12px; margin-top:-4px;">
    Snapshot at build time. Pairs the automated rules left for human review.
    VERY HIGH = obvious typos the auto-merger declined (usually because both
    look "established"). LOW = probably different people, just flagged.
  </p>
{''.join(suggestion_sections)}
</main>
<footer>
  Want a merge undone or added? <strong>Open an issue on GitHub</strong> with
  the <code>#id</code> of the merge (or the names of the suggested pair).
  Manual aliases live in <code>scripts/phase0/manual_aliases.json</code>.
</footer>
<script>
(function() {{
  const tbody = document.querySelector('#merge-log tbody');
  if (!tbody) return;
  const rows = Array.from(tbody.querySelectorAll('tr'));
  const filterEls = document.querySelectorAll('#merge-filters .pill');
  const searchEl = document.getElementById('merge-search');
  const emptyEl = document.getElementById('merge-empty');
  let activeKind = 'all';
  let q = '';

  function apply() {{
    let visible = 0;
    rows.forEach((r) => {{
      const k = r.dataset.kind;
      const s = r.dataset.search || '';
      const okKind = activeKind === 'all' || k === activeKind;
      const okSearch = !q || s.includes(q);
      const show = okKind && okSearch;
      r.style.display = show ? '' : 'none';
      if (show) visible++;
    }});
    emptyEl.style.display = visible === 0 ? '' : 'none';
  }}

  filterEls.forEach((el) => {{
    el.addEventListener('click', () => {{
      filterEls.forEach((x) => x.classList.remove('active'));
      el.classList.add('active');
      activeKind = el.dataset.kind;
      apply();
    }});
  }});
  searchEl.addEventListener('input', () => {{
    q = searchEl.value.trim().toLowerCase();
    apply();
  }});

  // If the URL has a #m-<id> hash, scroll into view + flash the row.
  if (location.hash && location.hash.startsWith('#m-')) {{
    const target = document.getElementById(location.hash.slice(1));
    if (target) {{
      // If the targeted row was hidden by an active filter, reset to "all".
      target.style.display = '';
      filterEls.forEach((x) => x.classList.toggle('active', x.dataset.kind === 'all'));
      activeKind = 'all';
      apply();
      target.scrollIntoView({{behavior: 'smooth', block: 'center'}});
    }}
  }}
}})();
</script>
</body>
</html>
"""


# --- How it works page -------------------------------------------------------


HOW_IT_WORKS_MODEL = "openskill_pl_decay365"


def build_how_it_works_page(conn: sqlite3.Connection) -> str:
    """ELI5 explainer + bell-curve diagram + match-prediction calculator.

    Calculator data is the production rating set (decay-weighted) embedded
    as JSON at build time — no API call, no stale numbers.
    """
    rows = conn.execute(
        """
        SELECT p.id, p.canonical_name, p.gender, r.mu, r.sigma, r.n_matches
        FROM players p
        JOIN ratings r ON r.player_id = p.id
        WHERE p.merged_into_id IS NULL
          AND r.model_name = ?
        """,
        (HOW_IT_WORKS_MODEL,),
    ).fetchall()
    # Strip players who never actually played (no rating-history rows ⇒
    # their μ is the seed value and a calculator pick on them is misleading).
    players_data = [
        {
            "id": pid,
            "name": name,
            "g": gender or "",
            "mu": round(mu, 4),
            "sigma": round(sigma, 4),
            "n": n,
        }
        for pid, name, gender, mu, sigma, n in rows
        if n and n > 0
    ]
    players_json = json.dumps(players_data)

    # Inline bell-curve SVG: two players overlapping. Player A is experienced
    # (narrow + tall = low σ); Player B is newer (wide + low = high σ but
    # similar μ). Visually conveys "the more matches you play, the sharper
    # the system's estimate becomes."
    bell_svg = """
<svg viewBox="0 0 600 240" xmlns="http://www.w3.org/2000/svg" role="img"
     aria-label="Two bell curves showing two players' rating distributions.">
  <defs>
    <linearGradient id="hwr-a" x1="0" x2="0" y1="0" y2="1">
      <stop offset="0%" stop-color="#46c281" stop-opacity="0.55"/>
      <stop offset="100%" stop-color="#46c281" stop-opacity="0.05"/>
    </linearGradient>
    <linearGradient id="hwr-b" x1="0" x2="0" y1="0" y2="1">
      <stop offset="0%" stop-color="#4ea1ff" stop-opacity="0.45"/>
      <stop offset="100%" stop-color="#4ea1ff" stop-opacity="0.05"/>
    </linearGradient>
  </defs>
  <!-- baseline axis -->
  <line x1="40" y1="200" x2="560" y2="200" stroke="#2a3242" stroke-width="1"/>
  <text x="40" y="220" fill="#8b96a8" font-size="11" font-family="system-ui">weaker</text>
  <text x="525" y="220" fill="#8b96a8" font-size="11" font-family="system-ui">stronger</text>
  <text x="295" y="220" fill="#8b96a8" font-size="11" font-family="system-ui" text-anchor="middle">skill →</text>

  <!-- Player A: narrow, tall (experienced, low σ) -->
  <!-- Approximate Gaussian with a smooth Bezier "hill" -->
  <path d="M 200,200 C 240,200 260,30 300,30 C 340,30 360,200 400,200 Z"
        fill="url(#hwr-a)" stroke="#46c281" stroke-width="1.5"/>
  <!-- mu marker -->
  <line x1="300" y1="30" x2="300" y2="200" stroke="#46c281" stroke-dasharray="3 3" stroke-width="1"/>
  <text x="300" y="22" fill="#46c281" font-size="11" font-family="system-ui" text-anchor="middle" font-weight="600">μ (Avery — many matches)</text>

  <!-- Player B: wide, lower peak (newer, high σ) -->
  <path d="M 100,200 C 170,200 220,80 320,80 C 420,80 470,200 540,200 Z"
        fill="url(#hwr-b)" stroke="#4ea1ff" stroke-width="1.5"/>
  <line x1="320" y1="80" x2="320" y2="200" stroke="#4ea1ff" stroke-dasharray="3 3" stroke-width="1"/>
  <text x="320" y="72" fill="#4ea1ff" font-size="11" font-family="system-ui" text-anchor="middle" font-weight="600">μ (Blake — new player)</text>

  <!-- σ width brackets -->
  <line x1="200" y1="208" x2="400" y2="208" stroke="#46c281" stroke-width="1"/>
  <text x="300" y="180" fill="#46c281" font-size="10" font-family="system-ui" text-anchor="middle">narrow σ ⇒ confident</text>
  <line x1="100" y1="216" x2="540" y2="216" stroke="#4ea1ff" stroke-width="1" opacity="0.7"/>
  <text x="320" y="135" fill="#4ea1ff" font-size="10" font-family="system-ui" text-anchor="middle">wide σ ⇒ unsure</text>
</svg>
"""

    body = f"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>How it works — RallyRank</title>
<link rel="stylesheet" href="styles.css?v={CSS_VERSION}">
</head>
<body>
{render_nav("", "how-it-works")}
<header>
  <h1>How RallyRank works</h1>
  <p>Plain English. No jargon. With a calculator at the bottom you can poke at.</p>
</header>

<main class="hwr">
  <p class="lead">Reading time: about 90 seconds. Skip to the calculator if you just want to play.</p>

  <section>
    <h2>The big idea</h2>
    <p>We don't just count wins. After every match, the system updates <em>both</em> a guess of how good you are <em>and</em> how confident it is in that guess. Two numbers per player, refined every time you step on court.</p>
  </section>

  <section>
    <h2>Two numbers: μ (skill) and σ (uncertainty)</h2>
    <p>Every player has two numbers behind their name:</p>
    <ul>
      <li><span class="code">μ</span> ("mu") — the system's best guess of your skill. Higher = stronger.</li>
      <li><span class="code">σ</span> ("sigma") — how unsure the system is. Higher = less confident. Shrinks as you play more matches.</li>
    </ul>
    <p>You can picture each player as a bell curve: μ is the peak, σ is the width. The wider the curve, the less the system trusts the number yet.</p>

    <div class="diagram">
      {bell_svg}
      <div class="diagram-caption">Avery has played a lot — narrow bell, system is confident. Blake is new — wide bell, system is hedging until they play more.</div>
    </div>
  </section>

  <section>
    <h2>Why the leaderboard sorts by μ − 3σ</h2>
    <p>If we ranked by μ alone, a brand-new player who happened to win their first match would shoot to the top. So we use a <em>conservative</em> number: <span class="code">μ − 3σ</span>. It's roughly "the lowest skill we're 99% sure this player is at least as strong as."</p>
    <p>Practical effect: new players start near the bottom and climb as their σ shrinks. Once you've played ~15+ matches, σ has shrunk enough that μ − 3σ is close to your real skill estimate.</p>
  </section>

  <section>
    <h2>What changes a rating after a match</h2>
    <p>Win or lose, both teams' ratings move. The base move is whatever OpenSkill's math says (bigger upset ⇒ bigger move). On top of that, four multipliers stack to scale the change up or down:</p>
    <ul>
      <li><strong>Game volume</strong> — a 26-game battle is more informative than a 6-0 6-0 blowout, so it changes ratings more.</li>
      <li><strong>Division K</strong> — a Div 1 win moves your rating more than a Div 4 win (higher division = stronger field = stronger signal).</li>
      <li><strong>Time decay</strong> — matches from two years ago count for less than matches from last month. Recent form weighs heavier.</li>
      <li><strong>Partner weighting</strong> — within a pair, the stronger partner's rating moves a bit more (they probably contributed more). Net team change is preserved.</li>
    </ul>
    <p>Walkovers count, but at half weight — the win was real, the play wasn't.</p>

    <p class="diagram-caption" style="margin-top:18px;">Interactive: slide the multipliers below to see how they stack on a hypothetical winning player whose baseline μ change is +1.00:</p>
    <div class="stack-diagram" id="hwr-stack">
      <div class="stack-row">
        <div class="stack-label">Baseline (OpenSkill)<span class="stack-mult" id="stk-base-mult">×1.00</span></div>
        <div class="stack-bar-track"><div class="stack-bar-fill baseline" id="stk-base-bar" style="width:50%;"></div></div>
        <div class="stack-value" id="stk-base-val">+1.00</div>
      </div>
      <div class="stack-row">
        <div class="stack-label">
          Game volume
          <span class="stack-mult" id="stk-vol-mult">×1.00</span>
          <input type="range" id="stk-vol" min="6" max="36" step="1" value="18"
                 style="width:100%;margin-top:4px;" aria-label="Total games in match">
          <span style="font-size:10px;color:var(--muted);" id="stk-vol-label">18 games (typical)</span>
        </div>
        <div class="stack-bar-track"><div class="stack-bar-fill" id="stk-vol-bar" style="width:50%;"></div></div>
        <div class="stack-value" id="stk-vol-val">+1.00</div>
      </div>
      <div class="stack-row">
        <div class="stack-label">
          Division K
          <span class="stack-mult" id="stk-div-mult">×0.85</span>
          <select id="stk-div" style="width:100%;margin-top:4px;background:var(--bg);color:var(--fg);border:1px solid var(--border);border-radius:4px;padding:3px 6px;font-size:11px;">
            <option value="1.00">Div 1</option>
            <option value="0.85" selected>Div 2</option>
            <option value="0.70">Div 3</option>
            <option value="0.55">Div 4</option>
          </select>
        </div>
        <div class="stack-bar-track"><div class="stack-bar-fill" id="stk-div-bar" style="width:42%;"></div></div>
        <div class="stack-value" id="stk-div-val">+0.85</div>
      </div>
      <div class="stack-row">
        <div class="stack-label">
          Time decay
          <span class="stack-mult" id="stk-time-mult">×1.00</span>
          <input type="range" id="stk-time" min="0" max="1095" step="30" value="0"
                 style="width:100%;margin-top:4px;" aria-label="Match age in days">
          <span style="font-size:10px;color:var(--muted);" id="stk-time-label">today</span>
        </div>
        <div class="stack-bar-track"><div class="stack-bar-fill" id="stk-time-bar" style="width:42%;"></div></div>
        <div class="stack-value" id="stk-time-val">+0.85</div>
      </div>
      <div class="stack-row">
        <div class="stack-label">
          Partner share
          <span class="stack-mult" id="stk-pw-mult">×1.00</span>
          <input type="range" id="stk-pw" min="35" max="65" step="1" value="50"
                 style="width:100%;margin-top:4px;" aria-label="This player's share of pair strength, percent">
          <span style="font-size:10px;color:var(--muted);" id="stk-pw-label">50% (equal partners)</span>
        </div>
        <div class="stack-bar-track"><div class="stack-bar-fill" id="stk-pw-bar" style="width:42%;"></div></div>
        <div class="stack-value" id="stk-pw-val">+0.85</div>
      </div>
      <div class="stack-row stack-final-row">
        <div class="stack-label">Final μ change</div>
        <div class="stack-bar-track"><div class="stack-bar-fill up" id="stk-final-bar" style="width:42%;"></div></div>
        <div class="stack-value" id="stk-final-val">+0.85</div>
      </div>
    </div>
    <p class="diagram-caption">Try: slide the time-decay slider out to 2 years and watch the final move shrink. Or push the partner-share slider to 65% to see the stronger partner's bump.</p>
  </section>

  <section>
    <h2>Captain's say sits on top</h2>
    <p>For tournaments where captains pre-assign players to A/B/C/D teams, the leaderboard groups by that captain choice <em>first</em>, then ranks by μ − 3σ <em>within</em> the group. Math knows numbers; captains know things math can't see (chemistry, injury, attitude). Both views are visible — the default respects captains, and a "raw" sort shows pure math.</p>
  </section>

  <section class="calc-section">
    <h2>Try it: predict the next match</h2>
    <p>Pick four players (start typing — the box autocompletes from real RallyRank players). The calculator shows who's favoured and by how much, using the production decay-weighted ratings.</p>

    <div class="calc">
      <div class="calc-pair">
        <span class="pair-label">Pair A</span>
        <input type="text" id="hwr-a1" list="hwr-players" placeholder="Player 1 name" autocomplete="off">
        <input type="text" id="hwr-a2" list="hwr-players" placeholder="Player 2 name" autocomplete="off">
      </div>
      <div class="calc-vs">vs</div>
      <div class="calc-pair">
        <span class="pair-label">Pair B</span>
        <input type="text" id="hwr-b1" list="hwr-players" placeholder="Player 3 name" autocomplete="off">
        <input type="text" id="hwr-b2" list="hwr-players" placeholder="Player 4 name" autocomplete="off">
      </div>
      <datalist id="hwr-players"></datalist>
      <div id="hwr-out" class="calc-out placeholder">
        Pick four different players to see a prediction.
      </div>
    </div>

    <div class="callout">
      <strong>Heads up:</strong> the prediction assumes both pairs play their typical level on the day. Tennis is a contact-with-a-ball sport — surface, weather, fatigue, and a bad decision in a tiebreak all swing real outcomes. Treat this as "what should happen on average," not a guaranteed result.
    </div>
  </section>

  <section>
    <h2>Want more depth?</h2>
    <p>The full algorithm is <a href="https://openskill.me/" target="_blank" rel="noopener">OpenSkill (Plackett-Luce)</a> with a time-decay weight, partner-weighted updates, and division-aware K. Engineering details live in the project's <a href="https://github.com/devkurtc/wks-social-tennis-rankings-malta/blob/main/PLAN.md" target="_blank" rel="noopener">PLAN.md</a> on GitHub.</p>
  </section>
</main>

<script>
(function() {{
  const PLAYERS = {players_json};
  // BETA controls the spread of skill->probability. 25/6 is the OpenSkill
  // default; we keep it for consistency with the engine's predict_win.
  const BETA = 25 / 6;

  // Index by lowercase canonical name for typeahead lookup.
  const byName = new Map();
  for (const p of PLAYERS) byName.set(p.name.toLowerCase(), p);

  // Populate the shared <datalist>.
  const dl = document.getElementById('hwr-players');
  // Sort alphabetically for the dropdown.
  const sorted = PLAYERS.slice().sort((a, b) => a.name.localeCompare(b.name));
  for (const p of sorted) {{
    const opt = document.createElement('option');
    opt.value = p.name;
    dl.appendChild(opt);
  }}

  // Abramowitz-Stegun erf approximation (max error ~1.5e-7).
  function erf(x) {{
    const sign = x < 0 ? -1 : 1;
    x = Math.abs(x);
    const a1 = 0.254829592, a2 = -0.284496736, a3 = 1.421413741;
    const a4 = -1.453152027, a5 = 1.061405429, p = 0.3275911;
    const t = 1.0 / (1.0 + p * x);
    const y = 1.0 - (((((a5*t + a4)*t) + a3)*t + a2)*t + a1)*t * Math.exp(-x*x);
    return sign * y;
  }}
  function normalCdf(z) {{ return 0.5 * (1 + erf(z / Math.SQRT2)); }}

  function lookup(inputId) {{
    const el = document.getElementById(inputId);
    const v = el.value.trim().toLowerCase();
    if (!v) {{ el.classList.remove('invalid'); return null; }}
    const p = byName.get(v);
    el.classList.toggle('invalid', !p);
    return p || null;
  }}

  // Real OpenSkill Plackett-Luce update for 2 teams of 2 players each.
  // Direct port of openskill.models.weng_lin.plackett_luce._compute (default
  // gamma = sqrt(team_sigma_squared)/c, no scores/weights/balance), validated
  // against the Python engine to ~4 decimal places (rest is float-order noise).
  const KAPPA = 0.0001;
  function plUpdate(teamA, teamB, winnerSide) {{
    const ranks = winnerSide === 'A' ? [0, 1] : [1, 0];
    function agg(team) {{
      let mu = 0, varSum = 0;
      for (const p of team) {{ mu += p.mu; varSum += p.sigma * p.sigma; }}
      return {{ mu, varSum }};
    }}
    const aggA = agg(teamA), aggB = agg(teamB);
    const c = Math.sqrt(aggA.varSum + aggB.varSum + 2 * BETA * BETA);
    const expA = Math.exp(aggA.mu / c);
    const expB = Math.exp(aggB.mu / c);
    const teams = [
      {{ mu: aggA.mu, varSum: aggA.varSum, rank: ranks[0], expMu: expA }},
      {{ mu: aggB.mu, varSum: aggB.varSum, rank: ranks[1], expMu: expB }},
    ];
    // sum_q[q] = sum over teams i where rank_i >= rank_q of exp(team_mu_i / c)
    const sum_q = [0, 0];
    for (const ti of teams) {{
      for (let q = 0; q < 2; q++) {{
        if (ti.rank >= teams[q].rank) sum_q[q] += ti.expMu;
      }}
    }}
    const a = [1, 1]; // teams have distinct ranks ⇒ one team per rank
    const updated = [[], []];
    for (let iIdx = 0; iIdx < 2; iIdx++) {{
      const ti = teams[iIdx];
      let omega = 0, delta = 0;
      for (let q = 0; q < 2; q++) {{
        if (teams[q].rank <= ti.rank) {{
          const p = ti.expMu / sum_q[q];
          delta += p * (1 - p) / a[q];
          if (q === iIdx) omega += (1 - p) / a[q];
          else omega -= p / a[q];
        }}
      }}
      omega *= ti.varSum / c;
      delta *= ti.varSum / (c * c);
      const gammaVal = Math.sqrt(ti.varSum) / c;
      delta *= gammaVal;
      const team = iIdx === 0 ? teamA : teamB;
      for (const pl of team) {{
        const sigSq = pl.sigma * pl.sigma;
        const ratio = sigSq / ti.varSum;
        const newMu = pl.mu + ratio * omega;
        const newSigma = pl.sigma * Math.sqrt(Math.max(1 - ratio * delta, KAPPA));
        updated[iIdx].push({{
          name: pl.name,
          mu: newMu,
          sigma: newSigma,
          deltaMu: newMu - pl.mu,
          deltaSigma: newSigma - pl.sigma,
        }});
      }}
    }}
    return {{ A: updated[0], B: updated[1] }};
  }}

  // State for the "if A wins" / "if B wins" toggle. Defaults to whichever
  // pair is the favourite — flips automatically when the favourite changes.
  let pickedWinner = null; // 'A' | 'B' | null (auto = favourite)
  let lastFav = null;

  function fmt2(x) {{
    const v = x.toFixed(2);
    return x >= 0 ? '+' + v : v; // signed for deltas; non-negative numbers get a + prefix where used
  }}
  function fmtSigned(x) {{ return (x >= 0 ? '+' : '') + x.toFixed(2); }}
  function deltaClass(d) {{
    if (Math.abs(d) < 0.005) return 'delta-zero';
    return d > 0 ? 'delta-pos' : 'delta-neg';
  }}

  function compute() {{
    const a1 = lookup('hwr-a1');
    const a2 = lookup('hwr-a2');
    const b1 = lookup('hwr-b1');
    const b2 = lookup('hwr-b2');
    const out = document.getElementById('hwr-out');

    const all = [a1, a2, b1, b2].filter(Boolean);
    if (all.length < 4) {{
      out.classList.add('placeholder');
      out.innerHTML = 'Pick four different players to see a prediction.';
      return;
    }}
    const ids = all.map(p => p.id);
    if (new Set(ids).size !== 4) {{
      out.classList.add('placeholder');
      out.innerHTML = 'Pair A and Pair B must be four different players.';
      return;
    }}

    // Win probability — Phi((muA - muB) / sqrt(2β² + varA + varB)).
    const muA = a1.mu + a2.mu, muB = b1.mu + b2.mu;
    const varA = a1.sigma**2 + a2.sigma**2;
    const varB = b1.sigma**2 + b2.sigma**2;
    const c = Math.sqrt(2 * BETA * BETA + varA + varB);
    const pA = normalCdf((muA - muB) / c);
    const pB = 1 - pA;
    const aPct = Math.round(pA * 100);
    const bPct = 100 - aPct;

    const fav = pA >= pB ? 'A' : 'B';
    const favPct = Math.max(pA, pB);
    let verdict;
    if (favPct >= 0.85) verdict = 'a heavy favourite — an upset would be a real surprise.';
    else if (favPct >= 0.65) verdict = 'the favourite, but the underdog has a real chance.';
    else if (favPct >= 0.55) verdict = 'a slight edge — close to a coin flip.';
    else verdict = 'about a coin flip on paper.';
    const favText = fav === 'A'
      ? `Pair A (${{a1.name}} + ${{a2.name}})`
      : `Pair B (${{b1.name}} + ${{b2.name}})`;

    // If favourite flipped (e.g. user swapped players), reset auto-pick to the new favourite.
    if (lastFav !== fav) {{ pickedWinner = null; lastFav = fav; }}
    const winner = pickedWinner || fav;

    // Compute post-match ratings under the chosen scenario.
    const teamA = [{{name: a1.name, mu: a1.mu, sigma: a1.sigma}},
                   {{name: a2.name, mu: a2.mu, sigma: a2.sigma}}];
    const teamB = [{{name: b1.name, mu: b1.mu, sigma: b1.sigma}},
                   {{name: b2.name, mu: b2.mu, sigma: b2.sigma}}];
    const after = plUpdate(teamA, teamB, winner);

    function rowFor(before, side, idx) {{
      const post = after[side][idx];
      const conf = before.sigma < 4 ? 'high'
                 : before.sigma < 6 ? 'medium' : 'low';
      const dMu = post.deltaMu;
      const dSigma = post.deltaSigma;
      return `
        <div>${{before.name}} <span style="color: var(--muted);">(${{side}})</span></div>
        <div class="num">${{before.mu.toFixed(2)}}</div>
        <div class="num">${{before.sigma.toFixed(2)}}</div>
        <div class="num hide-mobile" style="color: var(--muted);">${{conf}} · ${{before.n}}m</div>
        <div class="num ${{deltaClass(dMu)}}">${{fmtSigned(dMu)}} → ${{post.mu.toFixed(2)}}</div>
        <div class="num ${{deltaClass(dSigma)}} hide-mobile">${{fmtSigned(dSigma)}} → ${{post.sigma.toFixed(2)}}</div>
      `;
    }}

    const isUpset = (winner !== fav);
    const upsetTag = isUpset
      ? ` <span style="color: var(--loss); font-weight:600;">(upset!)</span>`
      : '';

    out.classList.remove('placeholder');
    out.innerHTML = `
      <p><strong>${{favText}}</strong> is ${{verdict}}</p>
      <div class="calc-bar">
        <div class="a" style="width: ${{Math.max(aPct, 6)}}%;">A · ${{aPct}}%</div>
        <div class="b" style="width: ${{Math.max(bPct, 6)}}%;">B · ${{bPct}}%</div>
      </div>

      <div class="calc-section-head">If the match is played</div>
      <div class="calc-toggle" role="tablist">
        <button data-pick="A" class="${{winner === 'A' ? 'active' : ''}}">Pair A wins</button>
        <button data-pick="B" class="${{winner === 'B' ? 'active' : ''}}">Pair B wins</button>
      </div>
      <p style="margin: 8px 0 0 0; color: var(--muted); font-size: 12.5px;">Showing rating moves if <strong style="color: var(--fg);">${{winner === 'A' ? 'Pair A' : 'Pair B'}}</strong> wins${{upsetTag}}. Toggle to see the other outcome.</p>

      <div class="calc-table">
        <div class="calc-th">Player</div>
        <div class="calc-th num">μ before</div>
        <div class="calc-th num">σ before</div>
        <div class="calc-th num hide-mobile">Conf · n</div>
        <div class="calc-th num">Δ μ → after</div>
        <div class="calc-th num hide-mobile">Δ σ → after</div>
        ${{rowFor(a1, 'A', 0)}}
        ${{rowFor(a2, 'A', 1)}}
        ${{rowFor(b1, 'B', 0)}}
        ${{rowFor(b2, 'B', 1)}}
      </div>
    `;

    // Wire the toggle buttons inside the freshly-rendered output.
    out.querySelectorAll('.calc-toggle button').forEach(btn => {{
      btn.addEventListener('click', () => {{
        pickedWinner = btn.dataset.pick;
        compute();
      }});
    }});
  }}

  ['hwr-a1','hwr-a2','hwr-b1','hwr-b2'].forEach(id => {{
    document.getElementById(id).addEventListener('input', compute);
    document.getElementById(id).addEventListener('change', compute);
  }});

  // --- Multiplier-stack diagram (interactive) ---
  // Visualizes how the four post-match multipliers scale a baseline μ
  // change of +1.00 for one player on the winning side. Pure illustration —
  // numbers come from the production constants in scripts/phase0/rating.py.
  function clamp(x, lo, hi) {{ return Math.max(lo, Math.min(hi, x)); }}

  function volMult(games) {{
    // rating.py: max(0.5, min(1.5, games / 18)) — clamped at [0.5, 1.5].
    return clamp(games / 18, 0.5, 1.5);
  }}
  function timeDecayMult(days) {{
    // openskill_pl_decay365: w = exp(-age / 365).
    return Math.exp(-days / 365);
  }}
  function partnerMult(sharePct) {{
    // Stronger partner gets larger share of pair's net Δμ. We model this
    // as a linear amplifier: 50% share = ×1.00 (equal split), 65% = ×1.30,
    // 35% = ×0.70. (rating.apply_partner_weighting in code; visualization
    // keeps the linear shape since absolute size depends on partner's μ.)
    return sharePct / 50;
  }}

  function updateStack() {{
    const baseline = 1.00;
    const games = parseFloat(document.getElementById('stk-vol').value);
    const div = parseFloat(document.getElementById('stk-div').value);
    const days = parseFloat(document.getElementById('stk-time').value);
    const share = parseFloat(document.getElementById('stk-pw').value);

    const mVol = volMult(games);
    const mDiv = div;
    const mTime = timeDecayMult(days);
    const mPw = partnerMult(share);

    const vBase = baseline;
    const vVol = vBase * mVol;
    const vDiv = vVol * mDiv;
    const vTime = vDiv * mTime;
    const vPw = vTime * mPw;

    // Bar widths normalized so the largest of the running values uses 100%.
    const ref = Math.max(2.0, Math.abs(vBase), Math.abs(vVol), Math.abs(vDiv), Math.abs(vTime), Math.abs(vPw));
    const pct = (v) => Math.max(0, Math.min(100, Math.abs(v) / ref * 100));

    document.getElementById('stk-base-bar').style.width = pct(vBase) + '%';
    document.getElementById('stk-base-val').textContent = fmtSigned(vBase);

    document.getElementById('stk-vol-mult').textContent = '×' + mVol.toFixed(2);
    document.getElementById('stk-vol-bar').style.width = pct(vVol) + '%';
    document.getElementById('stk-vol-val').textContent = fmtSigned(vVol);
    const volLabel = games < 14 ? `${{games}} games (blowout)`
                  : games > 22 ? `${{games}} games (long battle)`
                  : `${{games}} games (typical)`;
    document.getElementById('stk-vol-label').textContent = volLabel;

    document.getElementById('stk-div-mult').textContent = '×' + mDiv.toFixed(2);
    document.getElementById('stk-div-bar').style.width = pct(vDiv) + '%';
    document.getElementById('stk-div-val').textContent = fmtSigned(vDiv);

    document.getElementById('stk-time-mult').textContent = '×' + mTime.toFixed(2);
    document.getElementById('stk-time-bar').style.width = pct(vTime) + '%';
    document.getElementById('stk-time-val').textContent = fmtSigned(vTime);
    let tLabel;
    if (days === 0) tLabel = 'today';
    else if (days < 60) tLabel = `${{Math.round(days)}} days ago`;
    else if (days < 730) tLabel = `${{Math.round(days/30)}} months ago`;
    else tLabel = `${{(days/365).toFixed(1)}} years ago`;
    document.getElementById('stk-time-label').textContent = tLabel;

    document.getElementById('stk-pw-mult').textContent = '×' + mPw.toFixed(2);
    document.getElementById('stk-pw-bar').style.width = pct(vPw) + '%';
    document.getElementById('stk-pw-val').textContent = fmtSigned(vPw);
    let pwLabel;
    if (share < 45) pwLabel = `${{share}}% (weaker half of pair)`;
    else if (share > 55) pwLabel = `${{share}}% (stronger half of pair)`;
    else pwLabel = `${{share}}% (equal partners)`;
    document.getElementById('stk-pw-label').textContent = pwLabel;

    const finalBar = document.getElementById('stk-final-bar');
    finalBar.style.width = pct(vPw) + '%';
    finalBar.classList.toggle('up', vPw >= 0);
    finalBar.classList.toggle('down', vPw < 0);
    document.getElementById('stk-final-val').textContent = fmtSigned(vPw);
  }}

  ['stk-vol', 'stk-div', 'stk-time', 'stk-pw'].forEach(id => {{
    document.getElementById(id).addEventListener('input', updateStack);
    document.getElementById(id).addEventListener('change', updateStack);
  }});
  updateStack();
}})();
</script>
</body>
</html>
"""
    return body


# --- Main --------------------------------------------------------------------


def _load_predictions(csv_path: Path) -> dict[int, dict]:
    """Load per-match predictions from a backtest CSV. Returns
    {match_id: {p_a, actual_a, log_loss, brier, correct}}.

    File is generated by `scripts/phase0/backtest.py --out <path>`; if the
    file doesn't exist the result is an empty dict (no Pred column rendered).
    """
    import csv as _csv
    out: dict[int, dict] = {}
    if not csv_path.exists():
        return out
    with csv_path.open() as f:
        reader = _csv.DictReader(f)
        for row in reader:
            try:
                out[int(row["match_id"])] = {
                    "p_a": float(row["p_a"]),
                    "actual_a": int(row["actual_a"]),
                    "log_loss": float(row["log_loss"]),
                    "brier": float(row["brier"]),
                    "correct": int(row["correct"]),
                }
            except (KeyError, ValueError):
                continue
    return out


def main() -> int:
    if not os.path.exists(DB_PATH):
        print(f"DB not found: {DB_PATH}", file=sys.stderr)
        return 1

    OUT_DIR.mkdir(exist_ok=True)
    write(OUT_DIR / "styles.css", CSS)
    # Tell GitHub Pages not to run Jekyll — it would otherwise hide files that
    # start with `_` and may rewrite paths. The site is fully pre-rendered.
    write(OUT_DIR / ".nojekyll", "")

    conn = sqlite3.connect(DB_PATH)
    try:
        name_lookup = fetch_player_lookup(conn)
        neighbours_by_gender = fetch_neighbour_index(conn)

        # Replay rating history once to derive per-(match, player) rank/score
        # impact (rank_before/after, deltas, bypassed/passed-by). Reused by
        # both the All Matches feed and per-player pages.
        print("Computing per-match impacts ...")
        match_impacts = compute_match_impacts(conn)
        print(f"  {len(match_impacts):,} (match, player) impact rows")

        # Load held-out predictions for per-player calibration. Source files
        # are produced by `scripts/phase0/backtest.py --out`. If they don't
        # exist, the Pred column / summary block silently disappears.
        repo_root = Path(__file__).resolve().parent.parent.parent
        pred_dir = repo_root / "_ANALYSIS_" / "model_evaluation" / "predictions"
        predictions: dict[str, dict[int, dict]] = {
            "openskill_pl_vanilla": _load_predictions(
                pred_dir / "openskill_pl_vanilla.csv"
            ),
            "openskill_pl_decay365": _load_predictions(
                pred_dir / "openskill_pl_decay365.csv"
            ),
        }
        n_pred = sum(len(p) for p in predictions.values())
        print(f"  Loaded {n_pred:,} prediction rows across "
              f"{len(predictions)} models")

        # Index
        write(OUT_DIR / "index.html", build_index(conn))
        print(f"Wrote {OUT_DIR / 'index.html'}")

        # All-matches feed (chronological)
        write(
            OUT_DIR / "matches.html",
            build_matches_page(conn, name_lookup, impacts=match_impacts),
        )
        print(f"Wrote {OUT_DIR / 'matches.html'}")

        # Model-disagreement feed: matches where the two production models
        # predicted most differently. Only renders if predictions exist for
        # both engines.
        disagreements_html = build_disagreements_page(
            conn, name_lookup, predictions, impacts=match_impacts,
        )
        if disagreements_html:
            write(OUT_DIR / "disagreements.html", disagreements_html)
            print(f"Wrote {OUT_DIR / 'disagreements.html'}")

        # What's new (changelog)
        changelog_html = build_changelog_page()
        if changelog_html:
            write(OUT_DIR / "changelog.html", changelog_html)
            print(f"Wrote {OUT_DIR / 'changelog.html'}")

        # How it works (ELI5 explainer + bell-curve + calculator)
        write(OUT_DIR / "how-it-works.html", build_how_it_works_page(conn))
        print(f"Wrote {OUT_DIR / 'how-it-works.html'}")

        # Mapping & merges (full identity-resolution transparency)
        write(OUT_DIR / "aliases.html", build_aliases_page(conn, name_lookup))
        print(f"Wrote {OUT_DIR / 'aliases.html'}")

        # Per-player pages: only for unmerged players with at least 1 active match.
        eligible = conn.execute("""
            SELECT DISTINCT p.id
            FROM players p
            JOIN match_sides ms ON ms.player1_id = p.id OR ms.player2_id = p.id
            JOIN matches m ON m.id = ms.match_id
            WHERE p.merged_into_id IS NULL AND m.superseded_by_run_id IS NULL
        """).fetchall()
        eligible_ids = [r[0] for r in eligible]
        for pid in eligible_ids:
            html_text = build_player_page(
                conn, pid, name_lookup, neighbours_by_gender,
                impacts=match_impacts,
                predictions=predictions,
            )
            if html_text:
                write(OUT_DIR / player_filename(pid), html_text)
        print(f"Wrote {len(eligible_ids)} player pages under {OUT_DIR/'players'}/")

        # Tournament roster pages.
        for cfg in TOURNAMENT_ROSTERS:
            roster_path = Path(cfg["roster_xlsx"])
            if not roster_path.exists():
                print(
                    f"  skipped tournament '{cfg['slug']}': "
                    f"roster file not found at {roster_path}",
                    file=sys.stderr,
                )
                continue
            page = build_tournament_roster_page(conn, cfg)
            if page:
                out = OUT_DIR / "tournaments" / f"{cfg['slug']}.html"
                write(out, page)
                print(f"Wrote {out}")
    finally:
        conn.close()
    print(f"\nDone. Open: file://{(OUT_DIR/'index.html').resolve()}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
