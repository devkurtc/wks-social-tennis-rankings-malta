"""Parser for `Draws and Results Elektra Mixed Doubles 2022.xlsx` (T-P0-014).

Spec: `scripts/phase0/parser_spec_elektra_2022.md`.

Unlike every other Phase 0 file, this one is a CROSS-TAB MATRIX:
  - row 4: column-header numbers `1.0..N.0`
  - rows 5..(4+N): one per pair (col 1 = rank, col 2 = pair string,
    cols 3..(2+N) = result string of "this pair vs opponent of rank col-2")
  - diagonal cells empty (no self-match)
  - upper triangle and lower triangle are reciprocal — the same match recorded
    from each side's perspective. Walking ONLY the upper triangle is what
    keeps each match inserted exactly once.
  - Div 5 A and Div 5 B share a single cross-group Final string in row 15
    col 2; we extract it from `Div 5 A` only to dedupe.

Score strings come in many flavors — see the `parse_score_string` docstring.

Public API:
    parse(xlsx_path: str, db_conn: sqlite3.Connection) -> int
        Returns the new ingestion_run_id.
"""

from __future__ import annotations

import json
import os
import re
import sqlite3
import sys
from typing import Optional

import openpyxl

# Path-relative imports — same pattern as the other parsers in this dir.
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import players as players_mod  # noqa: E402

# Reuse helpers / insertion logic from sibling parsers.
from parsers import sports_experience_2025 as se  # noqa: E402
from parsers import mixed_doubles as md  # noqa: E402

AGENT_VERSION = "phase0-elektra-2022-parser-1.0"
TOURNAMENT_FORMAT = "doubles_division"
DEFAULT_CLUB_NAME = "VLTC"
DEFAULT_CLUB_SLUG = "vltc"

# Matrix layout constants
HEADER_ROW = 4         # row holding column-header numbers (1.0, 2.0, ...)
RANK_COL = 1           # col holding pair-rank (1.0..N.0)
PAIR_COL = 2           # col holding pair string
DATA_COL_OFFSET = 2    # col index for opponent rank R is `2 + R`
FIRST_PAIR_ROW = 5     # row index for pair rank R is `4 + R`

# Final row (only on Div 5 A / Div 5 B) — observed at row 15, col 2.
FINAL_LABEL_PREFIX = "final:"


# ─────────────────────────────────────────────────────────────────────────────
# Score-string parsing
# ─────────────────────────────────────────────────────────────────────────────

_SET_RE = re.compile(r"(\d{1,2})\s*-\s*(\d{1,2})")
_TB_LABEL_RE = re.compile(r"\b[Tt]\s*/?\s*[Bb]\b")
_WALKOVER_RE = re.compile(r"\bw\s*/\s*o\b", re.IGNORECASE)


def parse_score_string(raw: str) -> Optional[dict]:
    """Parse a free-text score string into a structured dict.

    Returns None for unparseable / empty input. The returned dict has:
        sets: list[(set_no:int, side_a_games:int, side_b_games:int, was_tiebreak:int)]
        walkover: bool
        super_tiebreak: Optional[(side_a:int, side_b:int)]   # convenience copy

    Accepted formats (see spec for full list):
      '7-5, 6-0'                  → 2 regular sets, no super-tb
      '6-2; 4-6 TB 10-8'          → 2 regular sets + super-tb (semicolon variant)
      '2-6, 6-4 TB 8-10'          → 2 regular sets + super-tb
      '2-6, 6-7 T/B 9-11'         → 'T/B' variant
      '6-4, 3-6 4-10'             → bare super-tb (no TB label)
      '6-3, 4-6\\nTB 10-6'         → embedded newline tolerated
      '1-6, 6-4 TB 10-4\\n'        → trailing newline tolerated
      '6-0, 6-0 w/o'              → walkover marker; sets still recorded
    """
    if raw is None:
        return None
    if not isinstance(raw, str):
        return None
    s = raw.replace("\xa0", " ").strip()
    if not s:
        return None

    walkover = bool(_WALKOVER_RE.search(s))

    # Find every "<digits>-<digits>" pair (preserves order).
    matches = list(_SET_RE.finditer(s))
    if len(matches) < 2:
        # Need at least 2 sets to constitute a tennis match in this format.
        return None
    if len(matches) > 3:
        # Unexpected — log & bail out so caller can skip.
        print(
            f"[elektra_2022] unparseable score (>3 score pairs found): {raw!r}",
            file=sys.stderr,
        )
        return None

    # First set, second set are always the first two score pairs.
    s1a, s1b = int(matches[0].group(1)), int(matches[0].group(2))
    s2a, s2b = int(matches[1].group(1)), int(matches[1].group(2))

    # Did a third pair (super-tb) appear?
    super_tb: Optional[tuple[int, int]] = None
    if len(matches) == 3:
        third = matches[2]
        # Sanity: super-tb usually has at least one side >= 7 (10-pt).
        # We accept it whether or not 'TB'/'T/B' label is present, since
        # the file shows both forms ('6-4, 3-6 4-10' has no label).
        ta, tb = int(third.group(1)), int(third.group(2))
        super_tb = (ta, tb)

    sets: list[tuple[int, int, int, int]] = []
    # was_tiebreak for a regular set is set when either side scored 7 (tennis
    # tiebreak set convention). Mirrors `mixed_doubles._insert_match`.
    was_tb_1 = 1 if (s1a == 7 or s1b == 7) else 0
    was_tb_2 = 1 if (s2a == 7 or s2b == 7) else 0
    sets.append((1, s1a, s1b, was_tb_1))
    sets.append((2, s2a, s2b, was_tb_2))
    if super_tb is not None:
        sets.append((3, super_tb[0], super_tb[1], 1))  # was_tiebreak=1 for super-tb

    return {
        "sets": sets,
        "walkover": walkover,
        "super_tiebreak": super_tb,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Match insertion
# ─────────────────────────────────────────────────────────────────────────────

def _insert_match_from_parsed(
    conn: sqlite3.Connection,
    tournament_id: int,
    ingestion_run_id: int,
    division: str,
    round_label: Optional[str],
    pair_a_str: str,
    pair_b_str: str,
    parsed: dict,
    source_file_id: int,
    placeholder_date: str,
) -> int:
    """Insert a match given pre-parsed pair strings and a parsed score dict.

    Mirrors the side-aggregation + winner-decision logic from the sibling
    Mixed parser (which itself mirrors sports_experience_2025) so the rating
    engine sees identical conventions across all three.
    """
    cur = conn.execute(
        "INSERT INTO matches (tournament_id, played_on, match_type, division, round, "
        "ingestion_run_id, walkover) "
        "VALUES (?, ?, 'doubles', ?, ?, ?, ?)",
        (
            tournament_id,
            placeholder_date,
            division,
            round_label,
            ingestion_run_id,
            1 if parsed["walkover"] else 0,
        ),
    )
    match_id = cur.lastrowid

    a1_raw, a2_raw = se._split_pair(pair_a_str)
    b1_raw, b2_raw = se._split_pair(pair_b_str)
    a1_id = players_mod.get_or_create_player(conn, a1_raw, source_file_id)
    a2_id = players_mod.get_or_create_player(conn, a2_raw, source_file_id)
    b1_id = players_mod.get_or_create_player(conn, b1_raw, source_file_id)
    b2_id = players_mod.get_or_create_player(conn, b2_raw, source_file_id)
    # Mixed doubles → don't set players.gender (NULL allowed).

    # Insert per-set rows
    for set_no, ga, gb, tb in parsed["sets"]:
        conn.execute(
            "INSERT INTO match_set_scores (match_id, set_number, side_a_games, side_b_games, was_tiebreak) "
            "VALUES (?, ?, ?, ?, ?)",
            (match_id, set_no, ga, gb, tb),
        )

    # Aggregate per-side: regular sets only count toward sets_won + games_won.
    sets_won_a = sets_won_b = games_won_a = games_won_b = 0
    has_super_tb = False
    super_tb_a = super_tb_b = 0
    for set_no, ga, gb, _tb in parsed["sets"]:
        if set_no <= 2:
            games_won_a += ga
            games_won_b += gb
            if ga > gb:
                sets_won_a += 1
            elif gb > ga:
                sets_won_b += 1
        else:
            has_super_tb = True
            super_tb_a, super_tb_b = ga, gb

    if sets_won_a > sets_won_b:
        won_a, won_b = 1, 0
    elif sets_won_b > sets_won_a:
        won_a, won_b = 0, 1
    elif has_super_tb:
        if super_tb_a > super_tb_b:
            won_a, won_b = 1, 0
        elif super_tb_b > super_tb_a:
            won_a, won_b = 0, 1
        else:
            won_a, won_b = 0, 0
    else:
        won_a, won_b = 0, 0

    conn.execute(
        "INSERT INTO match_sides (match_id, side, player1_id, player2_id, sets_won, games_won, won) "
        "VALUES (?, 'A', ?, ?, ?, ?, ?)",
        (match_id, a1_id, a2_id, sets_won_a, games_won_a, won_a),
    )
    conn.execute(
        "INSERT INTO match_sides (match_id, side, player1_id, player2_id, sets_won, games_won, won) "
        "VALUES (?, 'B', ?, ?, ?, ?, ?)",
        (match_id, b1_id, b2_id, sets_won_b, games_won_b, won_b),
    )
    return match_id


# ─────────────────────────────────────────────────────────────────────────────
# Sheet walkers
# ─────────────────────────────────────────────────────────────────────────────

def _read_division_label(ws) -> str:
    """Pull the division label from cell [3,1]; fall back to sheet title."""
    v = ws.cell(3, 1).value
    if isinstance(v, str) and v.strip():
        return v.strip()
    return ws.title


def _discover_pair_count(ws) -> int:
    """Count the number of pairs by scanning row HEADER_ROW for numeric headers.

    Stops at the first non-numeric column (usually `'TOTAL POINTS'`).
    Returns 0 if no numeric headers found.
    """
    n = 0
    c = DATA_COL_OFFSET + 1  # first data col
    max_col = ws.max_column or 30
    while c <= max_col:
        v = ws.cell(HEADER_ROW, c).value
        if isinstance(v, (int, float)):
            n += 1
            c += 1
        else:
            break
    return n


def _read_pair_string(ws, rank: int) -> Optional[str]:
    """Return the pair string for player-rank `rank` (None if absent)."""
    v = ws.cell(FIRST_PAIR_ROW + rank - 1, PAIR_COL).value
    if not isinstance(v, str):
        return None
    s = v.strip()
    return s or None


def _walk_matrix_matches(
    ws,
    n_pairs: int,
    on_match,
    on_skip,
):
    """Walk the upper triangle of the matrix and call `on_match` per cell.

    `on_match(rank_a, rank_b, pair_a_str, pair_b_str, raw_score, parsed)`
    is called once per upper-triangle cell with a successfully-parsed score.

    `on_skip(rank_a, rank_b, raw_score, reason)` for cells that have content
    but couldn't be parsed.

    Empty cells are silently ignored (they're either the diagonal or a
    not-yet-played match — both fine to skip without noise).
    """
    for ra in range(1, n_pairs + 1):
        pair_a_str = _read_pair_string(ws, ra)
        if pair_a_str is None or "/" not in pair_a_str:
            # Defensive: the file should always have a pair string per rank,
            # but if a row is missing we'd rather skip than crash.
            continue
        for rb in range(ra + 1, n_pairs + 1):
            pair_b_str = _read_pair_string(ws, rb)
            if pair_b_str is None or "/" not in pair_b_str:
                continue
            cell_value = ws.cell(FIRST_PAIR_ROW + ra - 1, DATA_COL_OFFSET + rb).value
            if cell_value is None:
                # Not played yet, or genuinely empty — skip silently.
                continue
            if not isinstance(cell_value, str) or not cell_value.strip():
                continue
            parsed = parse_score_string(cell_value)
            if parsed is None:
                on_skip(ra, rb, cell_value, "could not parse score string")
                continue
            on_match(ra, rb, pair_a_str, pair_b_str, cell_value, parsed)


# ─────────────────────────────────────────────────────────────────────────────
# Cross-group Final (Div 5 A only)
# ─────────────────────────────────────────────────────────────────────────────

_FINAL_RE = re.compile(
    r"^\s*Final\s*:\s*(?P<pair_a>.+?/.+?)\s+vs\s+(?P<pair_b>.+?/.+?)\s+(?P<score>\d.+)$",
    re.IGNORECASE,
)


def _extract_div5_final(ws) -> Optional[tuple[str, str, dict]]:
    """Look for the Div 5 cross-group Final string in row 15 col 2.

    Returns (pair_a_str, pair_b_str, parsed_score_dict) or None if absent /
    unparseable.
    """
    max_r = ws.max_row or 30
    for r in range(11, max_r + 1):
        v = ws.cell(r, 2).value
        if isinstance(v, str) and v.strip().lower().startswith(FINAL_LABEL_PREFIX):
            m = _FINAL_RE.match(v.strip())
            if not m:
                print(
                    f"[elektra_2022] Final string did not match expected format: {v!r}",
                    file=sys.stderr,
                )
                return None
            pair_a = m.group("pair_a").strip()
            pair_b = m.group("pair_b").strip()
            score_str = m.group("score").strip()
            parsed = parse_score_string(score_str)
            if parsed is None:
                print(
                    f"[elektra_2022] Final score unparseable: {score_str!r}",
                    file=sys.stderr,
                )
                return None
            return (pair_a, pair_b, parsed)
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Top-level parse
# ─────────────────────────────────────────────────────────────────────────────

def _ensure_default_club(conn: sqlite3.Connection) -> int:
    row = conn.execute("SELECT id FROM clubs WHERE slug = ?", (DEFAULT_CLUB_SLUG,)).fetchone()
    if row is not None:
        return row[0]
    cur = conn.execute(
        "INSERT INTO clubs (name, slug) VALUES (?, ?)",
        (DEFAULT_CLUB_NAME, DEFAULT_CLUB_SLUG),
    )
    return cur.lastrowid


def _read_tournament_name(wb: openpyxl.Workbook) -> Optional[str]:
    """Tournament title from cell [1,1] of any 'Div *' sheet."""
    for sn in wb.sheetnames:
        if sn.lower().startswith("div"):
            ws = wb[sn]
            v = ws.cell(1, 1).value
            if isinstance(v, str) and v.strip():
                return v.replace("\xa0", " ").strip()
    return None


def parse(xlsx_path: str, db_conn: sqlite3.Connection) -> int:
    """Parse the Elektra 2022 cross-tab matrix file. Returns ingestion_run_id."""
    sha256 = se._sha256_of_file(xlsx_path)
    filename = os.path.basename(xlsx_path)
    year = md._extract_year_from_filename(filename) or 2022
    placeholder_date = f"{year}-01-01"

    with db_conn:
        club_id = _ensure_default_club(db_conn)

        existing = db_conn.execute(
            "SELECT id FROM source_files WHERE original_filename = ? AND sha256 = ? "
            "ORDER BY id DESC LIMIT 1",
            (filename, sha256),
        ).fetchone()
        if existing is not None:
            source_file_id = existing[0]
        else:
            cur = db_conn.execute(
                "INSERT INTO source_files (club_id, original_filename, sha256) VALUES (?, ?, ?)",
                (club_id, filename, sha256),
            )
            source_file_id = cur.lastrowid

        cur = db_conn.execute(
            "INSERT INTO ingestion_runs (source_file_id, status, agent_version) "
            "VALUES (?, 'running', ?)",
            (source_file_id, AGENT_VERSION),
        )
        ingestion_run_id = cur.lastrowid

        prior_run_id = se._supersede_prior_runs_for_file(db_conn, source_file_id, ingestion_run_id)
        if prior_run_id is not None:
            db_conn.execute(
                "UPDATE ingestion_runs SET supersedes_run_id = ? WHERE id = ?",
                (prior_run_id, ingestion_run_id),
            )

        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        try:
            tournament_name = _read_tournament_name(wb) or filename.rsplit(".", 1)[0]

            cur = db_conn.execute(
                "INSERT INTO tournaments (club_id, name, year, format, source_file_id) "
                "VALUES (?, ?, ?, ?, ?)",
                (club_id, tournament_name, year, TOURNAMENT_FORMAT, source_file_id),
            )
            tournament_id = cur.lastrowid

            n_matches_inserted = 0
            unparsed_or_empty: list[dict] = []
            walkovers: list[dict] = []
            finals_inserted: list[dict] = []

            # Walk every Division sheet in the workbook (sheet name starts with 'Div').
            for sheet_name in wb.sheetnames:
                if not sheet_name.lower().startswith("div"):
                    continue
                ws = wb[sheet_name]
                division = _read_division_label(ws)
                n_pairs = _discover_pair_count(ws)
                if n_pairs < 2:
                    print(
                        f"[elektra_2022] sheet {sheet_name!r}: no pair-count detected, skipping",
                        file=sys.stderr,
                    )
                    continue

                def on_match(ra, rb, pa, pb, raw, parsed, _div=division, _sheet=sheet_name):
                    nonlocal n_matches_inserted
                    _insert_match_from_parsed(
                        db_conn,
                        tournament_id,
                        ingestion_run_id,
                        _div,
                        None,
                        pa,
                        pb,
                        parsed,
                        source_file_id,
                        placeholder_date,
                    )
                    n_matches_inserted += 1
                    if parsed["walkover"]:
                        walkovers.append({
                            "sheet": _sheet,
                            "ranks": (ra, rb),
                            "raw": raw,
                        })

                def on_skip(ra, rb, raw, reason, _div=division, _sheet=sheet_name):
                    print(
                        f"[elektra_2022] {_sheet!r} ranks {ra} vs {rb}: "
                        f"skipped {raw!r} ({reason})",
                        file=sys.stderr,
                    )
                    unparsed_or_empty.append({
                        "sheet": _sheet,
                        "ranks": (ra, rb),
                        "raw": raw,
                        "reason": reason,
                    })

                _walk_matrix_matches(ws, n_pairs, on_match, on_skip)

                # Extract the cross-group Final (only from Div 5 A; the Div 5 B
                # copy is a duplicate string and would double-insert).
                if sheet_name.lower() == "div 5 a":
                    final = _extract_div5_final(ws)
                    if final is not None:
                        pair_a, pair_b, parsed = final
                        # For the final, drop the '- Group A' suffix so the
                        # division reads as 'Division 5'.
                        final_division = re.sub(
                            r"\s*-\s*(Group|Round)\s+.*$",
                            "",
                            division,
                            flags=re.IGNORECASE,
                        ).strip()
                        _insert_match_from_parsed(
                            db_conn,
                            tournament_id,
                            ingestion_run_id,
                            final_division,
                            "final",
                            pair_a,
                            pair_b,
                            parsed,
                            source_file_id,
                            placeholder_date,
                        )
                        n_matches_inserted += 1
                        finals_inserted.append({
                            "division": final_division,
                            "pair_a": pair_a,
                            "pair_b": pair_b,
                        })

            quality_report = {
                "n_matches_inserted": n_matches_inserted,
                "unparsed_or_empty": unparsed_or_empty,
                "walkovers": walkovers,
                "finals_inserted": finals_inserted,
                "placeholder_date_used": placeholder_date,
                "notes": [
                    "Cross-tab matrix file; only the upper triangle is processed "
                    "to avoid double-counting reciprocal cells.",
                    "Mixed doubles — players.gender left NULL.",
                    "File contains no per-match dates; played_on is a placeholder.",
                    "Div 5 A and Div 5 B both carry the cross-group Final string "
                    "in row 15 col 2; only Div 5 A's copy is processed.",
                ],
            }

            db_conn.execute(
                "UPDATE ingestion_runs "
                "SET status = 'completed', completed_at = datetime('now'), quality_report_jsonb = ? "
                "WHERE id = ?",
                (json.dumps(quality_report), ingestion_run_id),
            )
        finally:
            wb.close()

    return ingestion_run_id
