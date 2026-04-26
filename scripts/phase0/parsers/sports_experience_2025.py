"""Parser for `Sports Experience Chosen Doubles 2025 result sheet.xlsx` (T-P0-004).

Spec: `scripts/phase0/parser_spec_sports_experience_2025.md` (276 lines).
This parser is faithful to the file's structure 1:1 — no cleverness here.

Phase 0 limitations / pre-resolved decisions (per T-P0-004 progress log):
  1. The file has no per-match dates anywhere. All matches receive
     `played_on = '2025-01-01'` (placeholder). Within-tournament chronological
     order falls back to `match.id` insertion order, which mirrors the file's
     row layout (round-robin round order).
  2. Final-block layout differs across sheets. Men Div 3 splits player names
     across two rows; Men Div 4 / Lad Div 3 use single-row pair strings.
     Detected by `'/' in cell_value`.
  3. Lad Div 1 has unplayed matches with blank score cells — these are SKIPPED
     (not inserted) and logged to stderr. A literal `0.0` in BOTH score columns
     is a legitimate 6-0 6-0 bagel and IS recorded.
  4. `vs.` and `vs` (Men Div 2 quirk) are both accepted as the divider.
  5. `Players Ladies` rows 10/11 both rank 5.0 — both are valid pairs,
     no dedupe by rank.

Idempotency / re-process: re-loading the same file (matched by sha256) creates
a new `ingestion_runs` row, sets `supersedes_run_id` to the prior latest run
for that file, and marks all prior matches `superseded_by_run_id = <new>`.

Public API:
    parse(xlsx_path: str, db_conn: sqlite3.Connection) -> int
        Returns the new ingestion_run_id.
"""

from __future__ import annotations

import hashlib
import json
import os
import re
import sqlite3
import sys
from typing import Iterable, Optional

import openpyxl

# Path-relative import — works for both `python -m scripts.phase0.parsers.<x>`
# and module-level `import` from cli.py / tests.
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import players as players_mod  # noqa: E402

AGENT_VERSION = "phase0-manual-parser-1.0"
PLACEHOLDER_DATE = "2025-01-01"
TOURNAMENT_FORMAT = "doubles_division"
DEFAULT_CLUB_NAME = "VLTC"
DEFAULT_CLUB_SLUG = "vltc"

# Sheet → (division-label, gender, group-row-pairs)
# group-row-pairs: list of (group_label_row, group_a_or_b, matches_start_row)
# where matches_start_row is where the round-robin begins for that group.
# For single-group sheets, only one entry.
SHEET_DEFS: dict[str, dict] = {
    "Men Div 1": {
        "division": "Men Division 1",
        "gender": "M",
        "groups": [("A", 9)],  # single group — start at row 9
        "final_anchor_row": None,
    },
    "Men Div 2": {
        "division": "Men Division 2",
        "gender": "M",
        "groups": [("A", 9)],
        "final_anchor_row": None,
    },
    "Men Div 3": {
        "division": "Men Division 3",
        "gender": "M",
        "groups": [("A", 9), ("B", 39)],  # group B label at row 35, matches start row 39
        "final_anchor_row": 67,           # 'Final' label cell row in col 16
    },
    "Men Div 4": {
        "division": "Men Division 4",
        "gender": "M",
        "groups": [("A", 9), ("B", 39)],
        "final_anchor_row": 67,
    },
    "Lad Div 1": {
        "division": "Ladies Division 1",
        "gender": "F",
        "groups": [("A", 9)],
        "final_anchor_row": None,
    },
    "Lad Div 2": {
        "division": "Ladies Division 2",
        "gender": "F",
        "groups": [("A", 9)],
        "final_anchor_row": None,
    },
    "Lad Div 3": {
        "division": "Ladies Division 3",
        "gender": "F",
        "groups": [("A", 9), ("B", 47)],   # group B label at row 43, matches start row 47
        "final_anchor_row": 82,
    },
}

# Column anchors (1-indexed openpyxl). Left and right blocks share structure.
LEFT_BLOCK = {
    "name_col": 1,
    "set1_col": 2,
    "set2_col": 5,
    "tie_col": 8,
}
RIGHT_BLOCK = {
    "name_col": 16,
    "set1_col": 17,
    "set2_col": 20,
    "tie_col": 23,
}


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _sha256_of_file(path: str) -> str:
    """Compute SHA-256 hex digest of a file."""
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _is_vs_divider(value) -> bool:
    """Return True if the cell value is a `vs.` or `vs` divider (tolerantly)."""
    if not isinstance(value, str):
        return False
    return value.strip().rstrip(".").lower() == "vs"


def _is_pair_string(value) -> bool:
    """A pair string is a non-empty string containing exactly one '/'."""
    if not isinstance(value, str):
        return False
    s = value.strip()
    if not s:
        return False
    if s.lower().rstrip(".") == "vs":
        return False
    return "/" in s


def _split_pair(pair_str: str) -> tuple[str, str]:
    """Split `'First Last/First Last'` into two raw player names.

    Tolerant of multiple consecutive separators: `'Foo / Bar'`, `'Foo//Bar'`,
    `'Foo / / Bar'` all split correctly into `('Foo', 'Bar')`. Some clubs
    (e.g. TCK Mixed Doubles 2026) use `//` as the pair separator.

    Strips leading/trailing whitespace per half; preserves internal whitespace
    (last names like `Treeby Ward` are two-word). Does NOT normalize — that
    is `players.get_or_create_player`'s job.
    """
    # Split on one-or-more slashes, drop empty fragments
    parts = [p.strip() for p in pair_str.split("/") if p.strip()]
    if len(parts) != 2:
        raise ValueError(
            f"pair string did not split into exactly 2 names: {pair_str!r}"
        )
    return parts[0], parts[1]


def _coerce_score(value) -> Optional[int]:
    """Coerce an Excel score cell to int. Returns None for blank cells.

    Excel reads these as floats (`6.0`); we want clean ints. `0.0` is preserved
    as `0` (legitimate bagel score), not converted to None.
    """
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return int(value)
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        try:
            return int(float(s))
        except ValueError:
            return None
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Match extraction (single block)
# ─────────────────────────────────────────────────────────────────────────────

def _extract_block_match(ws, row: int, block: dict) -> Optional[dict]:
    """Extract one match from a block (left or right) anchored at `row`.

    Returns a dict with `pair_a`, `pair_b`, `set1`, `set2`, `tiebreak` or None
    if the row band does not contain a valid match (pair/vs/pair pattern broken,
    OR scores are completely blank — i.e. an unplayed match per edge case 5).

    On unplayed-match detection, prints a stderr log line.
    """
    name_col = block["name_col"]
    set1_col = block["set1_col"]
    set2_col = block["set2_col"]
    tie_col = block["tie_col"]

    pair_a_cell = ws.cell(row, name_col).value
    divider_cell = ws.cell(row + 1, name_col).value
    pair_b_cell = ws.cell(row + 2, name_col).value

    if not _is_pair_string(pair_a_cell):
        return None
    if not _is_vs_divider(divider_cell):
        return None
    if not _is_pair_string(pair_b_cell):
        return None

    s1a = _coerce_score(ws.cell(row, set1_col).value)
    s1b = _coerce_score(ws.cell(row + 2, set1_col).value)
    s2a = _coerce_score(ws.cell(row, set2_col).value)
    s2b = _coerce_score(ws.cell(row + 2, set2_col).value)
    tba = _coerce_score(ws.cell(row, tie_col).value)
    tbb = _coerce_score(ws.cell(row + 2, tie_col).value)

    # Unplayed match detection (edge case 5): if both set1 cells AND both set2
    # cells AND both tiebreak cells are None, this is an unplayed match.
    # `0.0` is a real bagel score and is preserved by _coerce_score.
    if s1a is None and s1b is None and s2a is None and s2b is None and tba is None and tbb is None:
        print(
            f"[parser] skipping unplayed match: "
            f"{ws.title!r} row {row} block_col={name_col} "
            f"pair_A={pair_a_cell!r} pair_B={pair_b_cell!r}",
            file=sys.stderr,
        )
        return None

    # Partial-blank tolerance: if set 1 is fully present but set 2 blank for
    # both sides (or vice-versa), the parser still records what's there.
    # The spec only flags fully-blank score rows as unplayed.
    return {
        "pair_a": pair_a_cell.strip(),
        "pair_b": pair_b_cell.strip(),
        "set1_a": s1a,
        "set1_b": s1b,
        "set2_a": s2a,
        "set2_b": s2b,
        "tb_a": tba,
        "tb_b": tbb,
    }


def _iter_group_matches(ws, start_row: int, max_iterations: int = 50) -> Iterable[tuple[int, dict, dict]]:
    """Iterate matches in a group, walking the (r, r+4, r+8, …) anchor pattern.

    Yields (row, left_match_or_None, right_match_or_None) tuples until the
    pattern breaks (no pair string at col 1 of `row`).
    """
    r = start_row
    for _ in range(max_iterations):
        # Stop when the LEFT block runs out — by that point the right block
        # has long since stopped too.
        left_pair = ws.cell(r, LEFT_BLOCK["name_col"]).value
        if not _is_pair_string(left_pair):
            return
        left = _extract_block_match(ws, r, LEFT_BLOCK)
        right = _extract_block_match(ws, r, RIGHT_BLOCK)
        yield r, left, right
        r += 4


def _find_score_row(ws, name_row: int, set1_col: int, set2_col: int, tie_col: int) -> int:
    """Locate the row holding numeric score cells for a Final-block entry.

    The Final block has up to two adjacent rows associated with one pair
    (when names are split). Scores live on whichever of `name_row` or
    `name_row + 1` has any numeric value in any of the score columns.
    Falls back to `name_row` if neither has numbers (defensive default).
    """
    for candidate in (name_row, name_row + 1):
        for col in (set1_col, set2_col, tie_col):
            v = ws.cell(candidate, col).value
            if isinstance(v, (int, float)):
                return candidate
    return name_row


# ─────────────────────────────────────────────────────────────────────────────
# Final-block extraction
# ─────────────────────────────────────────────────────────────────────────────

def _extract_final(ws, anchor_row: int) -> Optional[dict]:
    """Extract the Final-block match from a two-group sheet.

    Three known layouts (all anchored at `[anchor_row, 16]` = 'Final'):

    (a) Men Div 3 — names split across two rows per side:
        [70,16]='Dunstan Vella', [71,16]='Cyril Lastimosa', scores on row 71;
        [73,16]='Neville Sciriha', [74,16]='Matthias Sciriha', scores on row 73.

    (b) Men Div 4 — pair A is full-string at row 70 with scores on row 71;
        pair B is full-string at row 73 with scores on row 73.

    (c) Lad Div 3 — both pairs are full-string single-row at rows 86 and 88
        with scores on those same rows. Anchor `Final` at row 82, so pair A
        is at row 86 (anchor + 4) and pair B at row 88 (anchor + 6).

    The detection rule (per pre-resolved decision 2): use `'/' in cell_value`
    to distinguish "split-name" rows (no '/') from "pair-string" rows.
    """
    name_col = RIGHT_BLOCK["name_col"]
    set1_col = RIGHT_BLOCK["set1_col"]
    set2_col = RIGHT_BLOCK["set2_col"]
    tie_col = RIGHT_BLOCK["tie_col"]

    label = ws.cell(anchor_row, name_col).value
    if not isinstance(label, str) or label.strip().lower() != "final":
        return None

    # Search a small range below the anchor for the two pair-name rows. We
    # look for either a pair-string (contains '/') or a split-name (just a
    # bare name).
    # Layout (a) has Pair A at offset+3 (split, e.g. row 70 if anchor=67),
    # layout (b)/(c) has Pair A at offset+3 or +4 as a single string.
    # We scan offsets 3..8 in order and treat the first 'name-bearing' row
    # as pair A, then look for a 'vs.' divider, then pair B name.
    pair_a_first_row = None
    for off in range(3, 9):
        v = ws.cell(anchor_row + off, name_col).value
        if isinstance(v, str) and v.strip() and not _is_vs_divider(v):
            pair_a_first_row = anchor_row + off
            break
    if pair_a_first_row is None:
        return None

    pair_a_first = ws.cell(pair_a_first_row, name_col).value
    if "/" in pair_a_first:
        pair_a_str = pair_a_first.strip()
        # Single-row pair: pair name and (potentially) scores live on the same
        # row, but for Men Div 4 scores are on the next row. Choose by sniffing
        # where the set1 score actually lives (a numeric cell). If neither row
        # has a number, default to the same row.
        pair_a_score_row = _find_score_row(ws, pair_a_first_row, set1_col, set2_col, tie_col)
    else:
        # Split layout (Men Div 3): name-row + (name-row+1) is the second name,
        # scores live on whichever of the two has the numeric set1 cell. Per
        # observed file: Pair A's scores live on the SECOND name row (row 71),
        # Pair B's scores live on the FIRST name row (row 73). Use score-sniff.
        second = ws.cell(pair_a_first_row + 1, name_col).value
        if not isinstance(second, str) or not second.strip():
            return None
        pair_a_str = f"{pair_a_first.strip()}/{second.strip()}"
        pair_a_score_row = _find_score_row(ws, pair_a_first_row, set1_col, set2_col, tie_col)

    # Now find the 'vs.' divider below pair A's last row.
    vs_row = None
    search_start = pair_a_score_row + 1
    for r in range(search_start, search_start + 4):
        v = ws.cell(r, name_col).value
        if _is_vs_divider(v):
            vs_row = r
            break
    if vs_row is None:
        return None

    # Pair B's name starts on vs_row + 1.
    pair_b_first_row = vs_row + 1
    pair_b_first = ws.cell(pair_b_first_row, name_col).value
    if not isinstance(pair_b_first, str) or not pair_b_first.strip():
        return None

    if "/" in pair_b_first:
        pair_b_str = pair_b_first.strip()
        pair_b_score_row = _find_score_row(ws, pair_b_first_row, set1_col, set2_col, tie_col)
    else:
        # Split layout — Pair B name continues on next row.
        second = ws.cell(pair_b_first_row + 1, name_col).value
        if not isinstance(second, str) or not second.strip():
            return None
        pair_b_str = f"{pair_b_first.strip()}/{second.strip()}"
        pair_b_score_row = _find_score_row(ws, pair_b_first_row, set1_col, set2_col, tie_col)

    s1a = _coerce_score(ws.cell(pair_a_score_row, set1_col).value)
    s1b = _coerce_score(ws.cell(pair_b_score_row, set1_col).value)
    s2a = _coerce_score(ws.cell(pair_a_score_row, set2_col).value)
    s2b = _coerce_score(ws.cell(pair_b_score_row, set2_col).value)
    tba = _coerce_score(ws.cell(pair_a_score_row, tie_col).value)
    tbb = _coerce_score(ws.cell(pair_b_score_row, tie_col).value)

    if s1a is None and s1b is None and s2a is None and s2b is None and tba is None and tbb is None:
        return None  # nothing to record

    return {
        "pair_a": pair_a_str,
        "pair_b": pair_b_str,
        "set1_a": s1a,
        "set1_b": s1b,
        "set2_a": s2a,
        "set2_b": s2b,
        "tb_a": tba,
        "tb_b": tbb,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Match → DB row insertion
# ─────────────────────────────────────────────────────────────────────────────

def _insert_match(
    conn: sqlite3.Connection,
    tournament_id: int,
    ingestion_run_id: int,
    division: str,
    round_label: Optional[str],
    match_data: dict,
    source_file_id: int,
    gender: str,
) -> int:
    """Insert a match + its sides + per-set scores. Returns the match_id."""
    cur = conn.execute(
        "INSERT INTO matches (tournament_id, played_on, match_type, division, round, ingestion_run_id) "
        "VALUES (?, ?, 'doubles', ?, ?, ?)",
        (tournament_id, PLACEHOLDER_DATE, division, round_label, ingestion_run_id),
    )
    match_id = cur.lastrowid

    # Players for both sides — pass raw names verbatim.
    a1_raw, a2_raw = _split_pair(match_data["pair_a"])
    b1_raw, b2_raw = _split_pair(match_data["pair_b"])
    a1_id = players_mod.get_or_create_player(conn, a1_raw, source_file_id)
    a2_id = players_mod.get_or_create_player(conn, a2_raw, source_file_id)
    b1_id = players_mod.get_or_create_player(conn, b1_raw, source_file_id)
    b2_id = players_mod.get_or_create_player(conn, b2_raw, source_file_id)

    # Set gender on first sight if not already set. Phase 0 only — multi-club
    # cross-check is a Phase 1 problem.
    for pid in (a1_id, a2_id, b1_id, b2_id):
        conn.execute(
            "UPDATE players SET gender = ? WHERE id = ? AND gender IS NULL",
            (gender, pid),
        )

    # Build the per-set rows. set 1 and set 2 are required if at least one
    # side has a score; the match-tiebreak is only inserted if either tb cell
    # is non-None.
    set_rows: list[tuple[int, int, int, int]] = []  # (set_number, side_a_games, side_b_games, was_tiebreak)
    s1a, s1b = match_data["set1_a"], match_data["set1_b"]
    s2a, s2b = match_data["set2_a"], match_data["set2_b"]
    tba, tbb = match_data["tb_a"], match_data["tb_b"]

    if s1a is not None or s1b is not None:
        ga = s1a if s1a is not None else 0
        gb = s1b if s1b is not None else 0
        was_tb = (ga == 7 or gb == 7)
        set_rows.append((1, ga, gb, int(was_tb)))
    if s2a is not None or s2b is not None:
        ga = s2a if s2a is not None else 0
        gb = s2b if s2b is not None else 0
        was_tb = (ga == 7 or gb == 7)
        set_rows.append((2, ga, gb, int(was_tb)))
    if tba is not None or tbb is not None:
        ga = tba if tba is not None else 0
        gb = tbb if tbb is not None else 0
        # Match-deciding super-tiebreak; was_tiebreak = TRUE always.
        next_set_no = 3 if len(set_rows) == 2 else (max(r[0] for r in set_rows) + 1 if set_rows else 1)
        set_rows.append((next_set_no, ga, gb, 1))

    for set_no, ga, gb, tb in set_rows:
        conn.execute(
            "INSERT INTO match_set_scores (match_id, set_number, side_a_games, side_b_games, was_tiebreak) "
            "VALUES (?, ?, ?, ?, ?)",
            (match_id, set_no, ga, gb, tb),
        )

    # Per-side aggregates: sets_won counts only first-2 sets (regular sets);
    # PLAN.md §5.2 says match-tiebreak doesn't count toward games_won either.
    sets_won_a = 0
    sets_won_b = 0
    games_won_a = 0
    games_won_b = 0
    has_super_tb = False
    for set_no, ga, gb, tb in set_rows:
        if set_no <= 2:
            games_won_a += ga
            games_won_b += gb
            if ga > gb:
                sets_won_a += 1
            elif gb > ga:
                sets_won_b += 1
        else:
            has_super_tb = True
            super_tb_a = ga
            super_tb_b = gb

    # Determine the winning side. Regular sets resolve it normally; if tied at
    # 1-1 in regular sets and a super-tiebreak was played, the super-tb decides.
    if sets_won_a > sets_won_b:
        won_a, won_b = 1, 0
    elif sets_won_b > sets_won_a:
        won_a, won_b = 0, 1
    elif has_super_tb:
        if super_tb_a > super_tb_b:
            won_a, won_b = 1, 0
        elif super_tb_b > super_tb_a:
            won_a, won_b = 0, 1
        else:
            # Pathological tie — leave both as 0; the rating engine should
            # treat this as undecided and probably skip.
            won_a, won_b = 0, 0
    else:
        won_a, won_b = 0, 0

    conn.execute(
        "INSERT INTO match_sides (match_id, side, player1_id, player2_id, sets_won, games_won, won) "
        "VALUES (?, 'A', ?, ?, ?, ?, ?)",
        (match_id, a1_id, a2_id, sets_won_a, games_won_a, won_a),
    )
    conn.execute(
        "INSERT INTO match_sides (match_id, side, player1_id, player2_id, sets_won, games_won, won) "
        "VALUES (?, 'B', ?, ?, ?, ?, ?)",
        (match_id, b1_id, b2_id, sets_won_b, games_won_b, won_b),
    )

    return match_id


# ─────────────────────────────────────────────────────────────────────────────
# Top-level parse
# ─────────────────────────────────────────────────────────────────────────────

def _ensure_default_club(conn: sqlite3.Connection) -> int:
    """Ensure a default club row exists; return its id."""
    row = conn.execute("SELECT id FROM clubs WHERE slug = ?", (DEFAULT_CLUB_SLUG,)).fetchone()
    if row is not None:
        return row[0]
    cur = conn.execute(
        "INSERT INTO clubs (name, slug) VALUES (?, ?)",
        (DEFAULT_CLUB_NAME, DEFAULT_CLUB_SLUG),
    )
    return cur.lastrowid


def _supersede_prior_runs_for_file(conn: sqlite3.Connection, source_file_id: int, new_run_id: int) -> Optional[int]:
    """Mark all prior matches loaded from this source_file as superseded.

    Returns the latest prior-run id (for `supersedes_run_id` linkage) or None
    if this is the first run for the file.
    """
    row = conn.execute(
        "SELECT id FROM ingestion_runs "
        "WHERE source_file_id = ? AND id != ? "
        "ORDER BY id DESC LIMIT 1",
        (source_file_id, new_run_id),
    ).fetchone()
    if row is None:
        return None
    prior_latest_id = row[0]

    # Mark every match from any prior run of this file as superseded by the new run.
    conn.execute(
        "UPDATE matches "
        "SET superseded_by_run_id = ? "
        "WHERE ingestion_run_id IN ("
        "    SELECT id FROM ingestion_runs WHERE source_file_id = ? AND id != ?"
        ") AND superseded_by_run_id IS NULL",
        (new_run_id, source_file_id, new_run_id),
    )

    # Mark the prior runs themselves as 'superseded'.
    conn.execute(
        "UPDATE ingestion_runs "
        "SET status = 'superseded' "
        "WHERE source_file_id = ? AND id != ? AND status != 'superseded'",
        (source_file_id, new_run_id),
    )
    return prior_latest_id


def parse(xlsx_path: str, db_conn: sqlite3.Connection) -> int:
    """Parse the Sports Experience Chosen Doubles 2025 file into the DB.

    Returns the new ingestion_run_id.
    """
    sha256 = _sha256_of_file(xlsx_path)
    filename = os.path.basename(xlsx_path)

    # Single transaction wrapper: everything below either commits or rolls
    # back atomically. The `with conn:` context commits on normal exit and
    # rolls back on exception.
    with db_conn:
        club_id = _ensure_default_club(db_conn)

        # source_files row — always insert a NEW row even if sha matches a
        # prior one. The supersede semantics use ingestion_runs, not source_files.
        # However, to support the file-level supersede correctly we MUST detect
        # the SAME logical file. We match by (filename, sha256) — if both match
        # an existing row, reuse that source_file_id; otherwise create new.
        existing = db_conn.execute(
            "SELECT id FROM source_files WHERE original_filename = ? AND sha256 = ? "
            "ORDER BY id DESC LIMIT 1",
            (filename, sha256),
        ).fetchone()
        if existing is not None:
            source_file_id = existing[0]
        else:
            cur = db_conn.execute(
                "INSERT INTO source_files (club_id, original_filename, sha256) VALUES (?, ?, ?)",
                (club_id, filename, sha256),
            )
            source_file_id = cur.lastrowid

        # ingestion_runs row — running, will set completed at end.
        cur = db_conn.execute(
            "INSERT INTO ingestion_runs (source_file_id, status, agent_version) "
            "VALUES (?, 'running', ?)",
            (source_file_id, AGENT_VERSION),
        )
        ingestion_run_id = cur.lastrowid

        # Supersede prior runs for this file (if any) — Phase 0 file-level.
        prior_run_id = _supersede_prior_runs_for_file(db_conn, source_file_id, ingestion_run_id)
        if prior_run_id is not None:
            db_conn.execute(
                "UPDATE ingestion_runs SET supersedes_run_id = ? WHERE id = ?",
                (prior_run_id, ingestion_run_id),
            )

        # tournaments row — always insert fresh per ingestion (ties to source file).
        # Tournament name is the second line of cell [1,1] of any match sheet
        # plus the year. We open the workbook once for everything below.
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
        # Collect cell [1,1] from the first match sheet (skip roster sheets).
        tournament_name_raw = None
        for sn in wb.sheetnames:
            if sn in SHEET_DEFS:
                ws_first = wb[sn]
                # In read_only mode, iterate row 1 to get cell [1,1].
                row1 = next(ws_first.iter_rows(min_row=1, max_row=1, values_only=True))
                if row1 and isinstance(row1[0], str):
                    tournament_name_raw = row1[0]
                break

        # Strip non-breaking spaces and split on newline; take the second line.
        if tournament_name_raw:
            cleaned = tournament_name_raw.replace("\xa0", " ")
            lines = [ln.strip() for ln in cleaned.split("\n") if ln.strip()]
            tournament_name = lines[1] if len(lines) >= 2 else lines[0]
        else:
            tournament_name = "Sports Experience Chosen Doubles 2025"

        cur = db_conn.execute(
            "INSERT INTO tournaments (club_id, name, year, format, source_file_id) "
            "VALUES (?, ?, ?, ?, ?)",
            (club_id, tournament_name, 2025, TOURNAMENT_FORMAT, source_file_id),
        )
        tournament_id = cur.lastrowid

        # Walk every match sheet.
        skipped_unplayed: list[dict] = []
        n_matches_inserted = 0

        for sheet_name, sheet_def in SHEET_DEFS.items():
            if sheet_name not in wb.sheetnames:
                continue
            # Need a non-read-only worksheet for random cell access. Re-open
            # the workbook in normal mode just for this. (read_only mode doesn't
            # support .cell(r,c) random access.)
            pass

        wb.close()

        # Re-open the workbook in normal mode for cell-random-access walks.
        wb_full = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=False)
        try:
            for sheet_name, sheet_def in SHEET_DEFS.items():
                if sheet_name not in wb_full.sheetnames:
                    continue
                ws = wb_full[sheet_name]
                base_division = sheet_def["division"]
                gender = sheet_def["gender"]
                groups = sheet_def["groups"]
                final_anchor = sheet_def["final_anchor_row"]

                for group_letter, start_row in groups:
                    if len(groups) > 1:
                        division_str = f"{base_division} - Group {group_letter}"
                    else:
                        division_str = base_division
                    for row, left_match, right_match in _iter_group_matches(ws, start_row):
                        if left_match is not None:
                            _insert_match(
                                db_conn,
                                tournament_id,
                                ingestion_run_id,
                                division_str,
                                None,
                                left_match,
                                source_file_id,
                                gender,
                            )
                            n_matches_inserted += 1
                        else:
                            # Could be unplayed; check why.
                            left_pair = ws.cell(row, LEFT_BLOCK["name_col"]).value
                            divider = ws.cell(row + 1, LEFT_BLOCK["name_col"]).value
                            pair_b = ws.cell(row + 2, LEFT_BLOCK["name_col"]).value
                            if (
                                _is_pair_string(left_pair)
                                and _is_vs_divider(divider)
                                and _is_pair_string(pair_b)
                            ):
                                skipped_unplayed.append({
                                    "sheet": sheet_name,
                                    "row": row,
                                    "block": "left",
                                    "division": base_division,
                                    "pair_A": str(left_pair).strip(),
                                    "pair_B": str(pair_b).strip(),
                                    "reason": "no scores recorded",
                                })
                        if right_match is not None:
                            _insert_match(
                                db_conn,
                                tournament_id,
                                ingestion_run_id,
                                division_str,
                                None,
                                right_match,
                                source_file_id,
                                gender,
                            )
                            n_matches_inserted += 1
                        else:
                            right_pair = ws.cell(row, RIGHT_BLOCK["name_col"]).value
                            divider = ws.cell(row + 1, RIGHT_BLOCK["name_col"]).value
                            pair_b = ws.cell(row + 2, RIGHT_BLOCK["name_col"]).value
                            if (
                                _is_pair_string(right_pair)
                                and _is_vs_divider(divider)
                                and _is_pair_string(pair_b)
                            ):
                                skipped_unplayed.append({
                                    "sheet": sheet_name,
                                    "row": row,
                                    "block": "right",
                                    "division": base_division,
                                    "pair_A": str(right_pair).strip(),
                                    "pair_B": str(pair_b).strip(),
                                    "reason": "no scores recorded",
                                })

                # Final block — separate division string (no group suffix).
                if final_anchor is not None:
                    final_match = _extract_final(ws, final_anchor)
                    if final_match is not None:
                        _insert_match(
                            db_conn,
                            tournament_id,
                            ingestion_run_id,
                            base_division,
                            "final",
                            final_match,
                            source_file_id,
                            gender,
                        )
                        n_matches_inserted += 1
        finally:
            wb_full.close()

        quality_report = {
            "n_matches_inserted": n_matches_inserted,
            "skipped_unplayed_matches": skipped_unplayed,
            "placeholder_date_used": PLACEHOLDER_DATE,
            "notes": [
                "File contains no per-match dates; played_on is a placeholder.",
                "Lad Div 1 has unplayed matches (blank score cells) — skipped.",
            ],
        }

        db_conn.execute(
            "UPDATE ingestion_runs "
            "SET status = 'completed', completed_at = datetime('now'), quality_report_jsonb = ? "
            "WHERE id = ?",
            (json.dumps(quality_report), ingestion_run_id),
        )

    return ingestion_run_id
