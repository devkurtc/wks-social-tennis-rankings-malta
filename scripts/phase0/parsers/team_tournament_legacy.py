"""Parser for the LEGACY VLTC Team-Tournament template family (T-P0-014).

Spec: `scripts/phase0/parser_spec_team_tournament_legacy.md`.

Files this parser handles (older "DAY N" template, distinct from the modern
`team_tournament.py` parser):
    - PKF  Team Tournament 2023.xlsx
    -  PKF  Team Tournament 2024.xlsx (leading space)
    - TENNIS TRADE  Team Tournament 2023.xlsx
    - SAN MICHEL TEAM TOURNAMENT 2023.xlsx (single MATCH RESULTS sheet)
    - SAN MICHEL TEAM TOURNAMENT 2025.xlsx
    -  Team Tournament 2024.xlsx (San Michel 2024 — leading space)

Layout: each "encounter" is a 21-row block. First row carries the encounter
label (`DAY N` / `FINAL` / `SEMI FINAL`). Row +1 has team captains; rows +4
through +19 hold up to 8 rubber blocks of 2 rows each. Rubber row r has the
date in col 2, CAT in col 3, side-A player 1 in col 6, side-B player 1 in
col 8; row r+1 has side-A player 2 in col 6, side-B player 2 in col 8. Set 1
games at [r,10] / [r,11]; set 2 games at [r+1,10] / [r+1,11].

Public API:
    parse(xlsx_path: str, db_conn: sqlite3.Connection) -> int
        Returns the new ingestion_run_id.

Idempotency: re-loading the same file (matched by filename + sha256) creates
a new ingestion_runs row and supersedes prior matches. Same convention as
sports_experience_2025.py / team_tournament.py.
"""

from __future__ import annotations

import datetime as _dt
import hashlib
import json
import os
import re
import sqlite3
import sys
from typing import Optional

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import players as players_mod  # noqa: E402

AGENT_VERSION = "phase0-team-tournament-legacy-parser-1.0"
TOURNAMENT_FORMAT = "doubles_team"
DEFAULT_CLUB_NAME = "VLTC"
DEFAULT_CLUB_SLUG = "vltc"

# Column anchors (1-indexed openpyxl)
COL_DAYNAME = 1
COL_DATE = 2
COL_CAT = 3
COL_CRT = 4
COL_TIME = 5
COL_SIDE_A_PLAYER = 6
COL_VS = 7
COL_SIDE_B_PLAYER = 8
COL_SET_LABEL = 9
COL_GAMES_A = 10
COL_GAMES_B = 11
COL_SETS_A = 12
COL_SETS_B = 13
COL_NOTE = 14

# Encounter layout offsets relative to the encounter anchor row
ENC_TEAM_OFFSET = 1   # row +1: TEAM A vs TEAM B header
ENC_HEADER_OFFSET = 2  # row +2: DAY/DATE/CAT/CRT/TIME/PLAYERS headers
ENC_SUBHEADER_OFFSET = 3  # row +3: GAMES/SETS sub-headers
ENC_FIRST_RUBBER_OFFSET = 4  # row +4: first rubber's SET 1 row
ENC_RUBBERS_PER_BLOCK = 8
ENC_NOTES_OFFSET = 20  # row +20: NOTES row
ENC_TOTALS_OFFSET = 21  # row +21: team totals
ENC_BLOCK_HEIGHT = 22  # min rows from one encounter anchor to next

# Year-in-filename detection
_YEAR_IN_FILENAME_RE = re.compile(r"\b(20\d{2})\b")
_YEAR_IN_TITLE_RE = re.compile(r"\b(19|20)\d{2}\b")

# Walkover marker
_WALKOVER_RE = re.compile(r"walk\s*over|\bw/o\b", re.IGNORECASE)

# Day-of-week label (helps confirm a row is a rubber row when DATE is missing)
_WEEKDAY_RE = re.compile(
    r"^(mon|tue|wed|thu|fri|sat|sun)(day|s)?$", re.IGNORECASE
)


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _sha256_of_file(path: str) -> str:
    """Compute SHA-256 hex digest of a file."""
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _coerce_score(value) -> Optional[int]:
    """Coerce an Excel score cell to int. Returns None for blank cells."""
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return int(value)
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        try:
            return int(float(s))
        except ValueError:
            return None
    return None


def _strip_str(value) -> Optional[str]:
    """Return stripped string or None for non-strings/empty values.

    Strips trailing NBSP (\\xa0) and trailing apostrophe-followed-by-space
    artifacts often seen in the legacy files.
    """
    if value is None:
        return None
    if isinstance(value, str):
        s = value.strip()
        # Python's str.strip handles NBSP (\xa0) since it's whitespace.
        return s if s else None
    if isinstance(value, (int, float)):
        return str(value).strip()
    return None


def _coerce_date(value) -> Optional[str]:
    """Coerce a cell value to ISO 8601 date string."""
    if value is None:
        return None
    if isinstance(value, _dt.datetime):
        return value.date().isoformat()
    if isinstance(value, _dt.date):
        return value.isoformat()
    if isinstance(value, str):
        # Try common string forms (DD/MM/YYYY).
        m = re.search(r"(\d{1,2})[/.\-](\d{1,2})[/.\-](\d{2,4})", value)
        if m:
            d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
            if y < 100:
                y += 2000
            try:
                return _dt.date(y, mo, d).isoformat()
            except ValueError:
                return None
    return None


def _is_set1_marker(value) -> bool:
    """Return True if cell value is the 'SET 1' marker."""
    if not isinstance(value, str):
        return False
    s = value.strip().upper().replace(" ", "")
    return s in ("SET1", "SET#1", "1STSET")


def _normalize_division(raw: str) -> str:
    """Map a raw CAT label (legacy file form) to the canonical division name.

    See spec table — collapses sub-tiers (A1, A2 → A) and unifies LDY's/LDYS'
    variants with LAD.
    """
    if not raw:
        return ""
    s = raw.strip()
    upper = s.upper()
    # Strip trailing whitespace + collapse internal whitespace
    upper = re.sub(r"\s+", " ", upper)

    # Singles
    if upper.startswith("SINGLE"):
        return "Singles"

    # Mixed — preserve the gender-tier variant info (e.g. "MIXED B/A" → "Mixed B/A")
    if upper.startswith("MIXED"):
        rest = upper[len("MIXED"):].strip()
        return ("Mixed " + rest).strip() if rest else "Mixed"

    # Men tiers — collapse MEN A1/A2/A 1/A 2/A01/A02 → MEN A
    m = re.match(r"^MEN\s+([A-D])\s*\d*$", upper)
    if m:
        return f"Men {m.group(1)}"

    # Ladies tiers — accept LAD / LDY'S / LDYS' / LADIES variants;
    # collapse sub-tier suffixes (A1, A 1, A02, etc.).
    m = re.match(r"^(LAD|LADIES|LDY'?S?'?|LDS)\s+([A-D])\s*\d*$", upper)
    if m:
        return f"Lad {m.group(2)}"

    # Fallback — title-case the input so it has a stable form (rating engine
    # falls back to defaults for unrecognised divisions).
    return s.title()


def _gender_from_division(division: str) -> Optional[str]:
    """Map normalized division name to player gender."""
    if not division:
        return None
    if division.startswith("Men"):
        return "M"
    if division.startswith("Lad"):
        return "F"
    return None


def _is_singles(division: str) -> bool:
    return division.lower().startswith("single")


def _round_label_for_sheet(sheet_name: str) -> Optional[str]:
    """Map a sheet name to a round label (or None for MATCH RESULTS)."""
    s = sheet_name.strip().lower()
    if s == "final":
        return "final"
    if s in ("semi final", "semifinal", "semi-final"):
        return "semi-final"
    m = re.match(r"day\s*(\d+)", s)
    if m:
        return f"day {int(m.group(1))}"
    return None


def _round_label_for_encounter_label(label: str) -> Optional[str]:
    """Map an encounter anchor label (e.g. 'DAY 5', 'FINAL') to a round label."""
    if not isinstance(label, str):
        return None
    s = label.strip().upper()
    if s == "FINAL":
        return "final"
    if s in ("SEMI FINAL", "SEMI-FINAL", "SEMIFINAL"):
        return "semi-final"
    m = re.match(r"DAY\s*(\d+)", s)
    if m:
        return f"day {int(m.group(1))}"
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Sheet discovery
# ─────────────────────────────────────────────────────────────────────────────

def _is_match_sheet(sheet_name: str) -> bool:
    """Return True if the sheet name looks like a match-bearing sheet."""
    s = sheet_name.strip().lower()
    if re.match(r"day\s*\d+$", s):
        return True
    if s in ("semi final", "semifinal", "semi-final", "final"):
        return True
    if s in ("match results", "matches"):
        return True
    return False


def _find_encounter_anchors(ws) -> list[tuple[int, str]]:
    """Find all encounter anchor rows in a sheet.

    Returns a list of (anchor_row, label) where label is the original col-1
    text (e.g. 'DAY 1', 'FINAL', 'SEMI FINAL').

    An anchor is a cell in col 1 whose stripped value matches `DAY N` /
    `FINAL` / `SEMI FINAL`, AND the row +1 below contains a 'TEAM ' string
    in col 6 or col 8 (confirming this is a real encounter header rather
    than e.g. a section title elsewhere).
    """
    anchors: list[tuple[int, str]] = []
    max_row = min(ws.max_row, 2000)
    for r in range(1, max_row + 1):
        v = ws.cell(r, COL_DAYNAME).value
        if not isinstance(v, str):
            continue
        s = v.strip().upper()
        if not (
            re.match(r"^DAY\s*\d+$", s)
            or s == "FINAL"
            or s in ("SEMI FINAL", "SEMI-FINAL", "SEMIFINAL")
        ):
            continue
        # Confirm it's a real encounter header by looking for TEAM at row +1.
        team_a = ws.cell(r + ENC_TEAM_OFFSET, COL_SIDE_A_PLAYER).value
        team_b = ws.cell(r + ENC_TEAM_OFFSET, COL_SIDE_B_PLAYER).value
        is_team_row = (
            (isinstance(team_a, str) and "team" in team_a.lower())
            or (isinstance(team_b, str) and "team" in team_b.lower())
            # Some encounters have just captain names (no 'TEAM' prefix)
            # in San Michel 2024 bare file — accept if both cells have text
            # AND the row +2 has the DAY/DATE/CAT header.
            or (
                isinstance(team_a, str) and isinstance(team_b, str)
                and team_a.strip() and team_b.strip()
                and isinstance(ws.cell(r + ENC_HEADER_OFFSET, COL_DAYNAME).value, str)
                and ws.cell(r + ENC_HEADER_OFFSET, COL_DAYNAME).value.strip().upper() == "DAY"
            )
        )
        if not is_team_row:
            continue
        anchors.append((r, v.strip()))
    return anchors


# ─────────────────────────────────────────────────────────────────────────────
# Rubber + encounter extraction
# ─────────────────────────────────────────────────────────────────────────────

def _extract_rubber(ws, r: int, fallback_date: Optional[str]) -> Optional[dict]:
    """Extract a single rubber starting at row r (the SET 1 row).

    Returns dict with rubber data, or None if this row doesn't hold a rubber.
    """
    set1_marker = ws.cell(r, COL_SET_LABEL).value
    cat_raw = _strip_str(ws.cell(r, COL_CAT).value)
    if not _is_set1_marker(set1_marker) and not cat_raw:
        return None
    if not cat_raw:
        # Sometimes CAT cell is empty but SET 1 marker present — invalid.
        return None

    division = _normalize_division(cat_raw)
    is_singles_rubber = _is_singles(division)

    name_a1 = _strip_str(ws.cell(r, COL_SIDE_A_PLAYER).value)
    name_b1 = _strip_str(ws.cell(r, COL_SIDE_B_PLAYER).value)
    name_a2 = _strip_str(ws.cell(r + 1, COL_SIDE_A_PLAYER).value)
    name_b2 = _strip_str(ws.cell(r + 1, COL_SIDE_B_PLAYER).value)

    if not name_a1 or not name_b1:
        return None

    if is_singles_rubber:
        name_a2 = None
        name_b2 = None

    # Date can be at [r,2] or [r+1,2]; use fallback (previous rubber's date).
    date_iso = _coerce_date(ws.cell(r, COL_DATE).value)
    if date_iso is None:
        date_iso = _coerce_date(ws.cell(r + 1, COL_DATE).value)
    if date_iso is None:
        date_iso = fallback_date

    s1a = _coerce_score(ws.cell(r, COL_GAMES_A).value)
    s1b = _coerce_score(ws.cell(r, COL_GAMES_B).value)
    s2a = _coerce_score(ws.cell(r + 1, COL_GAMES_A).value)
    s2b = _coerce_score(ws.cell(r + 1, COL_GAMES_B).value)

    if s1a is None and s1b is None and s2a is None and s2b is None:
        # Empty / not-played rubber.
        return None

    return {
        "division": division,
        "side_a_p1": name_a1,
        "side_a_p2": name_a2,
        "side_b_p1": name_b1,
        "side_b_p2": name_b2,
        "date": date_iso,
        "set1_a": s1a,
        "set1_b": s1b,
        "set2_a": s2a,
        "set2_b": s2b,
        "is_singles": is_singles_rubber,
    }


def _detect_walkover_for_rubber(notes_text: Optional[str], division: str, scores: dict) -> bool:
    """Decide if a rubber is a walkover based on encounter NOTES text.

    Conservative: only flag walkover when NOTES mentions 'walkover' AND the
    note text contains the rubber's division name (case-insensitive prefix).
    """
    if not notes_text:
        return False
    if not _WALKOVER_RE.search(notes_text):
        return False
    # Look for mentions of the division gender+letter (e.g. "MEN B", "LDY B")
    # in the note. Crude but conservative.
    upper_note = notes_text.upper()
    div_upper = division.upper()  # e.g. "MEN B" or "LAD A" or "MIXED B/A"
    if div_upper in upper_note:
        return True
    # Also try LDY/LDS variants for Lad rubbers.
    if div_upper.startswith("LAD ") and ("LDY " + div_upper[4:]) in upper_note:
        return True
    if div_upper.startswith("LAD ") and ("LDY'S " + div_upper[4:]) in upper_note:
        return True
    return False


def _iter_encounter_rubbers(ws, anchor_row: int):
    """Yield (rubber_row, rubber_dict, walkover_bool) for each rubber in an encounter."""
    notes_text_cell = ws.cell(anchor_row + ENC_NOTES_OFFSET, COL_DATE).value
    notes_text = _strip_str(notes_text_cell)

    fallback_date: Optional[str] = None
    for i in range(ENC_RUBBERS_PER_BLOCK):
        r = anchor_row + ENC_FIRST_RUBBER_OFFSET + 2 * i
        rubber = _extract_rubber(ws, r, fallback_date=fallback_date)
        if rubber is None:
            continue
        if rubber.get("date"):
            fallback_date = rubber["date"]
        walkover = _detect_walkover_for_rubber(notes_text, rubber["division"], rubber)
        yield r, rubber, walkover


# ─────────────────────────────────────────────────────────────────────────────
# Match → DB row insertion
# ─────────────────────────────────────────────────────────────────────────────

def _insert_match(
    conn: sqlite3.Connection,
    tournament_id: int,
    ingestion_run_id: int,
    round_label: Optional[str],
    rubber: dict,
    walkover: bool,
    source_file_id: int,
    fallback_date: str,
) -> int:
    """Insert a match + its sides + per-set scores. Returns the match_id."""
    walkover_int = 1 if walkover else 0
    is_singles = rubber["is_singles"]
    match_type = "singles" if is_singles else "doubles"
    played_on = rubber["date"] or fallback_date

    cur = conn.execute(
        "INSERT INTO matches "
        "(tournament_id, played_on, match_type, division, round, ingestion_run_id, walkover) "
        "VALUES (?, ?, ?, ?, ?, ?, ?)",
        (tournament_id, played_on, match_type, rubber["division"], round_label, ingestion_run_id, walkover_int),
    )
    match_id = cur.lastrowid

    a1_id = players_mod.get_or_create_player(conn, rubber["side_a_p1"], source_file_id)
    a2_id = (
        players_mod.get_or_create_player(conn, rubber["side_a_p2"], source_file_id)
        if rubber["side_a_p2"] is not None
        else None
    )
    b1_id = players_mod.get_or_create_player(conn, rubber["side_b_p1"], source_file_id)
    b2_id = (
        players_mod.get_or_create_player(conn, rubber["side_b_p2"], source_file_id)
        if rubber["side_b_p2"] is not None
        else None
    )

    gender = _gender_from_division(rubber["division"])
    if gender is not None:
        for pid in (a1_id, a2_id, b1_id, b2_id):
            if pid is None:
                continue
            conn.execute(
                "UPDATE players SET gender = ? WHERE id = ? AND gender IS NULL",
                (gender, pid),
            )

    sets_won_a = 0
    sets_won_b = 0
    games_won_a = 0
    games_won_b = 0
    set_rows: list[tuple[int, int, int, int]] = []

    for set_no, ga_raw, gb_raw in (
        (1, rubber["set1_a"], rubber["set1_b"]),
        (2, rubber["set2_a"], rubber["set2_b"]),
    ):
        if ga_raw is None and gb_raw is None:
            continue
        ga = ga_raw if ga_raw is not None else 0
        gb = gb_raw if gb_raw is not None else 0
        was_tb = (ga == 7 or gb == 7)
        set_rows.append((set_no, ga, gb, int(was_tb)))
        games_won_a += ga
        games_won_b += gb
        if ga > gb:
            sets_won_a += 1
        elif gb > ga:
            sets_won_b += 1

    for set_no, ga, gb, tb in set_rows:
        conn.execute(
            "INSERT INTO match_set_scores (match_id, set_number, side_a_games, side_b_games, was_tiebreak) "
            "VALUES (?, ?, ?, ?, ?)",
            (match_id, set_no, ga, gb, tb),
        )

    if sets_won_a > sets_won_b:
        won_a, won_b = 1, 0
    elif sets_won_b > sets_won_a:
        won_a, won_b = 0, 1
    else:
        won_a, won_b = 0, 0

    conn.execute(
        "INSERT INTO match_sides (match_id, side, player1_id, player2_id, sets_won, games_won, won) "
        "VALUES (?, 'A', ?, ?, ?, ?, ?)",
        (match_id, a1_id, a2_id, sets_won_a, games_won_a, won_a),
    )
    conn.execute(
        "INSERT INTO match_sides (match_id, side, player1_id, player2_id, sets_won, games_won, won) "
        "VALUES (?, 'B', ?, ?, ?, ?, ?)",
        (match_id, b1_id, b2_id, sets_won_b, games_won_b, won_b),
    )

    return match_id


# ─────────────────────────────────────────────────────────────────────────────
# Top-level parse
# ─────────────────────────────────────────────────────────────────────────────

def _ensure_default_club(conn: sqlite3.Connection) -> int:
    row = conn.execute("SELECT id FROM clubs WHERE slug = ?", (DEFAULT_CLUB_SLUG,)).fetchone()
    if row is not None:
        return row[0]
    cur = conn.execute(
        "INSERT INTO clubs (name, slug) VALUES (?, ?)",
        (DEFAULT_CLUB_NAME, DEFAULT_CLUB_SLUG),
    )
    return cur.lastrowid


def _supersede_prior_runs_for_file(
    conn: sqlite3.Connection, source_file_id: int, new_run_id: int
) -> Optional[int]:
    row = conn.execute(
        "SELECT id FROM ingestion_runs "
        "WHERE source_file_id = ? AND id != ? "
        "ORDER BY id DESC LIMIT 1",
        (source_file_id, new_run_id),
    ).fetchone()
    if row is None:
        return None
    prior_latest_id = row[0]
    conn.execute(
        "UPDATE matches "
        "SET superseded_by_run_id = ? "
        "WHERE ingestion_run_id IN ("
        "    SELECT id FROM ingestion_runs WHERE source_file_id = ? AND id != ?"
        ") AND superseded_by_run_id IS NULL",
        (new_run_id, source_file_id, new_run_id),
    )
    conn.execute(
        "UPDATE ingestion_runs "
        "SET status = 'superseded' "
        "WHERE source_file_id = ? AND id != ? AND status != 'superseded'",
        (source_file_id, new_run_id),
    )
    return prior_latest_id


def _derive_tournament_name(filename: str, wb) -> str:
    """Derive tournament name from the file's title cells, falling back to filename."""
    for sn in wb.sheetnames:
        if not _is_match_sheet(sn):
            continue
        ws = wb[sn]
        for r, c in ((1, 3), (1, 4), (1, 5), (1, 6), (3, 3), (5, 1), (6, 1)):
            v = ws.cell(r, c).value
            if isinstance(v, str):
                txt = v.strip()
                if txt and not txt.startswith("#REF") and txt.upper() != "DAY 1":
                    return txt
        break
    base = os.path.splitext(filename)[0].strip()
    base = re.sub(r"\s+", " ", base)
    return base


def _detect_year(filename: str, tournament_name: str, fallback_date: Optional[str]) -> int:
    m = _YEAR_IN_FILENAME_RE.search(filename)
    if m:
        return int(m.group(1))
    m = re.search(r"\b(20\d{2})\b", tournament_name)
    if m:
        return int(m.group(1))
    if fallback_date:
        try:
            return int(fallback_date[:4])
        except ValueError:
            pass
    return 2024


def parse(xlsx_path: str, db_conn: sqlite3.Connection) -> int:
    """Parse a legacy team-tournament file into the DB. Returns ingestion_run_id."""
    sha256 = _sha256_of_file(xlsx_path)
    filename = os.path.basename(xlsx_path)

    with db_conn:
        club_id = _ensure_default_club(db_conn)

        # Reuse source_files row by (filename, sha256) when present; fall back to
        # sha256 alone so a scraper-renamed-but-identical file dedups correctly
        # (otherwise two source_files rows → two tournaments rows → duplicate
        # active matches; tournament-duplication bug, fixed Apr 2026).
        existing = db_conn.execute(
            "SELECT id FROM source_files WHERE original_filename = ? AND sha256 = ? "
            "ORDER BY id DESC LIMIT 1",
            (filename, sha256),
        ).fetchone()
        if existing is None:
            existing = db_conn.execute(
                "SELECT id FROM source_files WHERE sha256 = ? "
                "ORDER BY id DESC LIMIT 1",
                (sha256,),
            ).fetchone()
        if existing is not None:
            source_file_id = existing[0]
        else:
            cur = db_conn.execute(
                "INSERT INTO source_files (club_id, original_filename, sha256) VALUES (?, ?, ?)",
                (club_id, filename, sha256),
            )
            source_file_id = cur.lastrowid

        cur = db_conn.execute(
            "INSERT INTO ingestion_runs (source_file_id, status, agent_version) "
            "VALUES (?, 'running', ?)",
            (source_file_id, AGENT_VERSION),
        )
        ingestion_run_id = cur.lastrowid

        prior_run_id = _supersede_prior_runs_for_file(db_conn, source_file_id, ingestion_run_id)
        if prior_run_id is not None:
            db_conn.execute(
                "UPDATE ingestion_runs SET supersedes_run_id = ? WHERE id = ?",
                (prior_run_id, ingestion_run_id),
            )

        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=False)
        try:
            tournament_name = _derive_tournament_name(filename, wb)

            # Pre-pass: scan for any date in any rubber to use as year fallback.
            first_date: Optional[str] = None
            for sn in wb.sheetnames:
                if not _is_match_sheet(sn):
                    continue
                ws = wb[sn]
                for r in range(9, min(ws.max_row + 1, 50)):
                    d = _coerce_date(ws.cell(r, COL_DATE).value)
                    if d is not None:
                        first_date = d
                        break
                if first_date:
                    break

            tournament_year = _detect_year(filename, tournament_name, first_date)

            # Get-or-create on (club_id, name, year) — prevents two scraper
            # ingests of the same physical tournament under different filenames
            # producing two tournament rows with overlapping active matches.
            existing_t = db_conn.execute(
                "SELECT id FROM tournaments WHERE club_id = ? AND name = ? AND year = ? "
                "ORDER BY id LIMIT 1",
                (club_id, tournament_name, tournament_year),
            ).fetchone()
            if existing_t is not None:
                tournament_id = existing_t[0]
            else:
                cur = db_conn.execute(
                    "INSERT INTO tournaments (club_id, name, year, format, source_file_id) "
                    "VALUES (?, ?, ?, ?, ?)",
                    (club_id, tournament_name, tournament_year, TOURNAMENT_FORMAT, source_file_id),
                )
                tournament_id = cur.lastrowid

            quality_notes: list[str] = []
            n_matches_inserted = 0
            n_singles = 0
            n_walkovers = 0
            sheets_processed = []

            for sheet_name in wb.sheetnames:
                if not _is_match_sheet(sheet_name):
                    continue
                ws = wb[sheet_name]
                sheet_round_label = _round_label_for_sheet(sheet_name)

                anchors = _find_encounter_anchors(ws)
                if not anchors:
                    continue

                sheet_match_count = 0
                for anchor_row, anchor_label in anchors:
                    # Per-encounter round label: if the sheet itself has no
                    # mappable round (e.g. MATCH RESULTS), derive from the
                    # encounter's anchor label.
                    round_label = sheet_round_label or _round_label_for_encounter_label(anchor_label)

                    for r, rubber, walkover in _iter_encounter_rubbers(ws, anchor_row):
                        fallback_date_for_match = first_date or f"{tournament_year}-01-01"
                        _insert_match(
                            db_conn,
                            tournament_id,
                            ingestion_run_id,
                            round_label,
                            rubber,
                            walkover,
                            source_file_id,
                            fallback_date_for_match,
                        )
                        n_matches_inserted += 1
                        sheet_match_count += 1
                        if walkover:
                            n_walkovers += 1
                        if rubber["is_singles"]:
                            n_singles += 1

                sheets_processed.append({
                    "sheet": sheet_name,
                    "round": sheet_round_label,
                    "encounters": len(anchors),
                    "matches_inserted": sheet_match_count,
                })
        finally:
            wb.close()

        quality_report = {
            "n_matches_inserted": n_matches_inserted,
            "n_walkovers": n_walkovers,
            "n_singles": n_singles,
            "sheets_processed": sheets_processed,
            "notes": quality_notes,
        }

        db_conn.execute(
            "UPDATE ingestion_runs "
            "SET status = 'completed', completed_at = datetime('now'), quality_report_jsonb = ? "
            "WHERE id = ?",
            (json.dumps(quality_report), ingestion_run_id),
        )

    return ingestion_run_id
