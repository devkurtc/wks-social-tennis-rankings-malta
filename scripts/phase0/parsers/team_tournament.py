"""Parser for the modern VLTC Team-Tournament template family (T-P0-014).

Spec: `scripts/phase0/parser_spec_team_tournament.md`.

Files this parser handles (modern "Day N" template):
    - Antes Insurance Team Tournament IMO Joe results 2025.xlsx
    - San Michel Results 2025 Results.xlsx
    - San Michel Results 2026.xlsx
    - Tennis Trade Team Tournament - Results.xlsx (single-row variant)
    - Results Tennis Trade Team Tournament(1).xlsx (single-row variant)
    - (Empty templates: Antes "results sets" / "results with sets", Tennis Trade
      "Results" / "Results(1)" — these produce 0 matches; not an error.)

Files NOT handled (older single-sheet "DAY" template — separate parser needed):
    - TENNIS TRADE Team Tournament 2023.xlsx
    - SAN MICHEL TEAM TOURNAMENT 2023.xlsx
    - SAN MICHEL TEAM TOURNAMENT 2025.xlsx
    - PKF Team Tournament 2023.xlsx
    - PKF Team Tournament 2024.xlsx

Two row-layout variants are auto-detected per sheet:
    - Two-row: rubber occupies rows r and r+1; col 10/11 hold per-set games
      (set 1 on row r, set 2 on row r+1). Antes 2025, San Michel 2025/2026.
    - Single-row: rubber occupies only row r; col 10/11 hold TOTAL games for
      the whole match. Tennis Trade 2024/2025. Optional walkover annotation
      at col 12.

Public API:
    parse(xlsx_path: str, db_conn: sqlite3.Connection) -> int
        Returns the new ingestion_run_id.

Idempotency: re-loading the same file (matched by sha256+filename) creates a
new ingestion_runs row, supersedes prior matches via superseded_by_run_id.
Same convention as sports_experience_2025.py.
"""

from __future__ import annotations

import hashlib
import json
import os
import re
import sqlite3
import sys
from typing import Optional

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import players as players_mod  # noqa: E402

AGENT_VERSION = "phase0-team-tournament-parser-1.0"
TOURNAMENT_FORMAT = "doubles_team"
DEFAULT_CLUB_NAME = "VLTC"
DEFAULT_CLUB_SLUG = "vltc"

# Column anchors (1-indexed openpyxl)
COL_TIME = 3
COL_RUBBER = 4
COL_SIDE_A_P1 = 5
COL_SIDE_A_P2 = 6
COL_VS = 7
COL_SIDE_B_P1 = 8
COL_SIDE_B_P2 = 9
COL_GAMES_A = 10
COL_GAMES_B = 11
COL_SETS_OR_NOTE = 12
COL_SETS_B = 13

# Rubber stride (rows between consecutive rubber anchors within one panel)
RUBBER_STRIDE = 3

# Date regex covers DD/MM/YYYY, DD-MM-YYYY, DD.MM.YYYY, with 2- or 4-digit year.
_DATE_RE = re.compile(r"(\d{1,2})[/.\-](\d{1,2})[/.\-](\d{2,4})")
# Textual-month variant: "11 July 2025", "1st January 2026"
_MONTH_NAMES = {
    "january": 1, "jan": 1, "february": 2, "feb": 2, "march": 3, "mar": 3,
    "april": 4, "apr": 4, "may": 5, "june": 6, "jun": 6, "july": 7, "jul": 7,
    "august": 8, "aug": 8, "september": 9, "sep": 9, "sept": 9, "october": 10,
    "oct": 10, "november": 11, "nov": 11, "december": 12, "dec": 12,
}
_TEXT_DATE_RE = re.compile(
    r"(\d{1,2})(?:st|nd|rd|th)?\s+("
    + "|".join(sorted(_MONTH_NAMES.keys(), key=len, reverse=True))
    + r")\s+(\d{2,4})",
    re.IGNORECASE,
)
_YEAR_IN_FILENAME_RE = re.compile(r"\b(20\d{2})\b")
_WALKOVER_RE = re.compile(r"\bw\s*/?\s*o\b|\bwalk\s*over\b", re.IGNORECASE)
_PRO_SUB_RE = re.compile(r"\(\s*pro\s*\)", re.IGNORECASE)


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _sha256_of_file(path: str) -> str:
    """Compute SHA-256 hex digest of a file."""
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _coerce_score(value) -> Optional[int]:
    """Coerce an Excel score cell to int. Returns None for blank cells.

    Excel reads numbers as floats (e.g. `6.0`). Strings parsed if numeric.
    `0.0` is preserved as `0` (legitimate bagel score).
    """
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return int(value)
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        try:
            return int(float(s))
        except ValueError:
            return None
    return None


def _strip_str(value) -> Optional[str]:
    """Return stripped string or None for non-strings/empty values."""
    if value is None:
        return None
    if isinstance(value, str):
        s = value.strip()
        return s if s else None
    if isinstance(value, (int, float)):
        return str(value).strip()
    return None


def _split_pro_substitute(player_cell: str) -> str:
    """Handle a `'Original (pro) Substitute'` cell.

    Returns the FIRST half (the player who started the rubber). Rationale in
    spec edge case 3: the original player started the match, so credit them
    for the result.
    """
    parts = _PRO_SUB_RE.split(player_cell, maxsplit=1)
    return parts[0].strip()


def _parse_first_date(text: str) -> Optional[str]:
    """Find the FIRST date in a free-text cell and return it as ISO 8601.

    Accepts numeric `DD/MM/YYYY` (or `-`/`.` separators, 2- or 4-digit year)
    and textual `DD Month YYYY` (`11 July 2025`).
    """
    if not isinstance(text, str):
        return None

    # Try the numeric form first (most common across the family).
    numeric = _DATE_RE.search(text)
    text_match = _TEXT_DATE_RE.search(text)

    # Pick whichever match starts earlier (so 'FINAL - 11 July 2025' picks the textual form).
    candidates = []
    if numeric:
        candidates.append(("numeric", numeric.start(), numeric))
    if text_match:
        candidates.append(("text", text_match.start(), text_match))
    if not candidates:
        return None
    candidates.sort(key=lambda t: t[1])
    kind, _, m = candidates[0]

    if kind == "numeric":
        d, mo, y = m.group(1), m.group(2), m.group(3)
        if len(y) == 2:
            y = "20" + y
        try:
            di, moi, yi = int(d), int(mo), int(y)
        except ValueError:
            return None
    else:
        d = m.group(1)
        month_name = m.group(2).lower()
        y = m.group(3)
        if len(y) == 2:
            y = "20" + y
        try:
            di = int(d)
            moi = _MONTH_NAMES[month_name]
            yi = int(y)
        except (ValueError, KeyError):
            return None

    if not (1 <= di <= 31 and 1 <= moi <= 12 and 2000 <= yi <= 2099):
        return None
    return f"{yi:04d}-{moi:02d}-{di:02d}"


def _detect_year_from_filename(filename: str) -> Optional[int]:
    """Find a 20YY year in a filename. Returns None if absent."""
    m = _YEAR_IN_FILENAME_RE.search(filename)
    if not m:
        return None
    return int(m.group(1))


def _is_vs_divider(value) -> bool:
    """Return True if cell is the 'vs'/'VS' divider."""
    if not isinstance(value, str):
        return False
    return value.strip().rstrip(".").lower() == "vs"


def _gender_from_rubber(rubber: str) -> Optional[str]:
    """Map a rubber-type label to a player gender, or None if mixed/unknown."""
    if not rubber:
        return None
    r = rubber.strip().lower()
    if r.startswith("men") or r.startswith("man"):
        return "M"
    if (
        r.startswith("lad")
        or r.startswith("ldy")
        or r.startswith("lds")
        or r.startswith("ladies")
    ):
        return "F"
    return None


def _round_label_for_sheet(sheet_name: str) -> Optional[str]:
    """Map a sheet name to a `matches.round` label."""
    s = sheet_name.strip().lower()
    if s == "final":
        return "final"
    if s in ("semi final", "semifinal", "semi-final"):
        return "semi-final"
    m = re.match(r"day\s*(\d+)", s)
    if m:
        return f"day {int(m.group(1))}"
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Sheet structure detection
# ─────────────────────────────────────────────────────────────────────────────

def _find_court_panels(ws) -> list[tuple[int, int]]:
    """Find all court panel header rows in a sheet.

    Returns a list of (court_label_row, header_row) tuples — header_row is
    where 'Time'/'Rubber' headers live.

    Primary strategy: a 'court panel' is identified by a cell in col 3 (or
    nearby) whose value starts with 'Court' (case-insensitive); the header
    row is 1-4 rows below.

    Fallback: if no court labels found, scan for any 'Time' header row in
    col 3 — Final sheets often have no court label but still use the same
    Time/Rubber/Games header structure. Each header row found this way
    becomes its own pseudo-panel.
    """
    panels = []
    for r in range(1, min(ws.max_row + 1, 500)):
        for c in (COL_TIME, COL_TIME - 1, COL_TIME + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip().lower().startswith("court"):
                header_row = None
                for hr in range(r + 1, r + 5):
                    hv = ws.cell(hr, COL_TIME).value
                    if isinstance(hv, str) and hv.strip().lower() == "time":
                        header_row = hr
                        break
                if header_row is not None:
                    panels.append((r, header_row))
                break

    if panels:
        return panels

    # Fallback: any "Time" header at col 3 with a "Rubber" header at col 4.
    for r in range(1, min(ws.max_row + 1, 500)):
        v = ws.cell(r, COL_TIME).value
        if not (isinstance(v, str) and v.strip().lower() == "time"):
            continue
        v2 = ws.cell(r, COL_RUBBER).value
        if isinstance(v2, str) and v2.strip().lower() == "rubber":
            panels.append((r, r))
    return panels


def _detect_two_row_variant(ws, header_row: int) -> bool:
    """Decide whether a sheet uses two-row (set 1+2) or single-row variant.

    Heuristic: presence of 'Sets' header at [header_row, COL_SETS_OR_NOTE]
    indicates two-row variant. Fallback: scan rubbers below header for any
    that have a numeric in [r+1, COL_GAMES_A] or [r+1, COL_GAMES_B].
    """
    sets_header = ws.cell(header_row, COL_SETS_OR_NOTE).value
    if isinstance(sets_header, str) and sets_header.strip().lower() == "sets":
        return True
    # Fallback: probe the first 6 rubbers below the header for set-2 numerics.
    for r in range(header_row + 1, header_row + 1 + 6 * RUBBER_STRIDE, RUBBER_STRIDE):
        # Don't look at set-2 row if rubber row itself is empty.
        rubber = _strip_str(ws.cell(r, COL_RUBBER).value)
        if not rubber:
            continue
        for off in (1,):
            v_a = ws.cell(r + off, COL_GAMES_A).value
            v_b = ws.cell(r + off, COL_GAMES_B).value
            if isinstance(v_a, (int, float)) or isinstance(v_b, (int, float)):
                return True
    return False


def _find_sheet_date(ws) -> Optional[str]:
    """Find a date string at one of the typical sheet-header positions.

    Tries [6,3], [5,3], [6,2], [5,2], [6,4]. Returns ISO date or None.
    """
    candidates = [(6, 3), (5, 3), (6, 2), (5, 2), (6, 4), (3, 3), (2, 3)]
    for r, c in candidates:
        v = ws.cell(r, c).value
        iso = _parse_first_date(v) if isinstance(v, str) else None
        if iso is not None:
            return iso
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Match extraction (one rubber)
# ─────────────────────────────────────────────────────────────────────────────

def _extract_rubber(ws, r: int, two_row: bool) -> Optional[dict]:
    """Extract a single rubber anchored at row r. Returns None if not a rubber."""
    rubber_label = _strip_str(ws.cell(r, COL_RUBBER).value)
    name_a1 = _strip_str(ws.cell(r, COL_SIDE_A_P1).value)
    name_b1 = _strip_str(ws.cell(r, COL_SIDE_B_P1).value)
    if not rubber_label:
        return None
    if not name_a1 or not name_b1:
        return None
    name_a2 = _strip_str(ws.cell(r, COL_SIDE_A_P2).value)
    name_b2 = _strip_str(ws.cell(r, COL_SIDE_B_P2).value)

    # Pro-substitute handling: if a name cell contains '(pro)', take FIRST half.
    if name_a1 and _PRO_SUB_RE.search(name_a1):
        name_a1 = _split_pro_substitute(name_a1)
    if name_a2 and _PRO_SUB_RE.search(name_a2):
        name_a2 = _split_pro_substitute(name_a2)
    if name_b1 and _PRO_SUB_RE.search(name_b1):
        name_b1 = _split_pro_substitute(name_b1)
    if name_b2 and _PRO_SUB_RE.search(name_b2):
        name_b2 = _split_pro_substitute(name_b2)

    walkover_note = _strip_str(ws.cell(r, COL_SETS_OR_NOTE).value)
    is_walkover = bool(walkover_note and isinstance(walkover_note, str) and _WALKOVER_RE.search(walkover_note))

    if two_row:
        s1a = _coerce_score(ws.cell(r, COL_GAMES_A).value)
        s1b = _coerce_score(ws.cell(r, COL_GAMES_B).value)
        s2a = _coerce_score(ws.cell(r + 1, COL_GAMES_A).value)
        s2b = _coerce_score(ws.cell(r + 1, COL_GAMES_B).value)
        # Skip if no scores at all on either set.
        if s1a is None and s1b is None and s2a is None and s2b is None:
            return None
        return {
            "rubber": rubber_label,
            "side_a_p1": name_a1,
            "side_a_p2": name_a2,
            "side_b_p1": name_b1,
            "side_b_p2": name_b2,
            "two_row": True,
            "set1_a": s1a,
            "set1_b": s1b,
            "set2_a": s2a,
            "set2_b": s2b,
            "walkover": is_walkover,
        }
    else:
        ta = _coerce_score(ws.cell(r, COL_GAMES_A).value)
        tb = _coerce_score(ws.cell(r, COL_GAMES_B).value)
        if ta is None and tb is None:
            return None
        return {
            "rubber": rubber_label,
            "side_a_p1": name_a1,
            "side_a_p2": name_a2,
            "side_b_p1": name_b1,
            "side_b_p2": name_b2,
            "two_row": False,
            "total_a": ta,
            "total_b": tb,
            "walkover": is_walkover,
        }


def _iter_panel_rubbers(ws, header_row: int, two_row: bool, panels_below: list[int]):
    """Yield (row, rubber_data) for all rubbers in a court panel.

    Court panels can contain multiple sub-panels (e.g., Day 1 Court 2 has
    Team-A-vs-E rubbers, then a Total row, then Team-D-vs-F rubbers — all
    under the same court header). Sub-panels start at non-uniform offsets
    (typically separated by Total rows + a blank row), so a fixed-stride walk
    misses the second sub-panel. We instead scan every row in the panel range
    and try to extract a rubber wherever a row matches the rubber pattern.

    Stops at the next panel's header_row (exclusive). If no next panel,
    stops at ws.max_row (or 200, whichever is smaller — Day sheets in the
    family don't exceed ~115 rows of real content).
    """
    next_panel_row = min(panels_below) if panels_below else None
    end_row = next_panel_row if next_panel_row is not None else min(ws.max_row + 1, 250)

    r = header_row + 1
    while r < end_row:
        # Skip header-row helpers like blank spacers before trying.
        rubber_data = _extract_rubber(ws, r, two_row)
        if rubber_data is not None:
            yield r, rubber_data
            # Skip the next row in two-row variant since it's the set-2 row.
            r += 2 if two_row else 1
        else:
            r += 1


# ─────────────────────────────────────────────────────────────────────────────
# Match → DB row insertion
# ─────────────────────────────────────────────────────────────────────────────

def _insert_match(
    conn: sqlite3.Connection,
    tournament_id: int,
    ingestion_run_id: int,
    division: str,
    round_label: Optional[str],
    played_on: str,
    rubber_data: dict,
    source_file_id: int,
) -> int:
    """Insert a match + its sides + per-set scores. Returns the match_id."""
    walkover = 1 if rubber_data.get("walkover") else 0
    # Detect singles: side A or side B has only one named player (player2 cell empty).
    is_singles = (
        rubber_data["side_a_p2"] is None
        and rubber_data["side_b_p2"] is None
    )
    match_type = "singles" if is_singles else "doubles"

    cur = conn.execute(
        "INSERT INTO matches "
        "(tournament_id, played_on, match_type, division, round, ingestion_run_id, walkover) "
        "VALUES (?, ?, ?, ?, ?, ?, ?)",
        (tournament_id, played_on, match_type, division, round_label, ingestion_run_id, walkover),
    )
    match_id = cur.lastrowid

    # Player creation. Pass raw names verbatim.
    a1_id = players_mod.get_or_create_player(conn, rubber_data["side_a_p1"], source_file_id)
    a2_id = (
        players_mod.get_or_create_player(conn, rubber_data["side_a_p2"], source_file_id)
        if rubber_data["side_a_p2"] is not None
        else None
    )
    b1_id = players_mod.get_or_create_player(conn, rubber_data["side_b_p1"], source_file_id)
    b2_id = (
        players_mod.get_or_create_player(conn, rubber_data["side_b_p2"], source_file_id)
        if rubber_data["side_b_p2"] is not None
        else None
    )

    # Gender update on first sight (only Men / Lad rubbers — leave Mixed/etc. NULL).
    gender = _gender_from_rubber(rubber_data["rubber"])
    if gender is not None:
        for pid in (a1_id, a2_id, b1_id, b2_id):
            if pid is None:
                continue
            conn.execute(
                "UPDATE players SET gender = ? WHERE id = ? AND gender IS NULL",
                (gender, pid),
            )

    # Build set-score rows + per-side aggregates.
    sets_won_a = 0
    sets_won_b = 0
    games_won_a = 0
    games_won_b = 0
    set_rows: list[tuple[int, int, int, int]] = []  # (set_number, side_a, side_b, was_tiebreak)

    if rubber_data.get("two_row"):
        s1a, s1b = rubber_data["set1_a"], rubber_data["set1_b"]
        s2a, s2b = rubber_data["set2_a"], rubber_data["set2_b"]

        if s1a is not None or s1b is not None:
            ga = s1a if s1a is not None else 0
            gb = s1b if s1b is not None else 0
            was_tb = (ga == 7 or gb == 7)
            set_rows.append((1, ga, gb, int(was_tb)))
            games_won_a += ga
            games_won_b += gb
            if ga > gb:
                sets_won_a += 1
            elif gb > ga:
                sets_won_b += 1

        if s2a is not None or s2b is not None:
            ga = s2a if s2a is not None else 0
            gb = s2b if s2b is not None else 0
            was_tb = (ga == 7 or gb == 7)
            set_rows.append((2, ga, gb, int(was_tb)))
            games_won_a += ga
            games_won_b += gb
            if ga > gb:
                sets_won_a += 1
            elif gb > ga:
                sets_won_b += 1
    else:
        # Single-row variant: only total games available.
        ta = rubber_data["total_a"]
        tb = rubber_data["total_b"]
        ta = ta if ta is not None else 0
        tb = tb if tb is not None else 0
        set_rows.append((1, ta, tb, 0))
        games_won_a = ta
        games_won_b = tb
        # Treat as a single "set" for the win indicator.
        if ta > tb:
            sets_won_a = 1
        elif tb > ta:
            sets_won_b = 1
        # Equal → both 0 (undecided).

    for set_no, ga, gb, tb in set_rows:
        conn.execute(
            "INSERT INTO match_set_scores (match_id, set_number, side_a_games, side_b_games, was_tiebreak) "
            "VALUES (?, ?, ?, ?, ?)",
            (match_id, set_no, ga, gb, tb),
        )

    if sets_won_a > sets_won_b:
        won_a, won_b = 1, 0
    elif sets_won_b > sets_won_a:
        won_a, won_b = 0, 1
    else:
        won_a, won_b = 0, 0

    conn.execute(
        "INSERT INTO match_sides (match_id, side, player1_id, player2_id, sets_won, games_won, won) "
        "VALUES (?, 'A', ?, ?, ?, ?, ?)",
        (match_id, a1_id, a2_id, sets_won_a, games_won_a, won_a),
    )
    conn.execute(
        "INSERT INTO match_sides (match_id, side, player1_id, player2_id, sets_won, games_won, won) "
        "VALUES (?, 'B', ?, ?, ?, ?, ?)",
        (match_id, b1_id, b2_id, sets_won_b, games_won_b, won_b),
    )

    return match_id


# ─────────────────────────────────────────────────────────────────────────────
# Top-level parse
# ─────────────────────────────────────────────────────────────────────────────

def _ensure_default_club(conn: sqlite3.Connection) -> int:
    """Ensure a default VLTC club row exists; return its id."""
    row = conn.execute("SELECT id FROM clubs WHERE slug = ?", (DEFAULT_CLUB_SLUG,)).fetchone()
    if row is not None:
        return row[0]
    cur = conn.execute(
        "INSERT INTO clubs (name, slug) VALUES (?, ?)",
        (DEFAULT_CLUB_NAME, DEFAULT_CLUB_SLUG),
    )
    return cur.lastrowid


def _supersede_prior_runs_for_file(
    conn: sqlite3.Connection, source_file_id: int, new_run_id: int
) -> Optional[int]:
    """Mark all prior matches loaded from this source_file as superseded.

    Same semantics as sports_experience_2025.py.
    """
    row = conn.execute(
        "SELECT id FROM ingestion_runs "
        "WHERE source_file_id = ? AND id != ? "
        "ORDER BY id DESC LIMIT 1",
        (source_file_id, new_run_id),
    ).fetchone()
    if row is None:
        return None
    prior_latest_id = row[0]
    conn.execute(
        "UPDATE matches "
        "SET superseded_by_run_id = ? "
        "WHERE ingestion_run_id IN ("
        "    SELECT id FROM ingestion_runs WHERE source_file_id = ? AND id != ?"
        ") AND superseded_by_run_id IS NULL",
        (new_run_id, source_file_id, new_run_id),
    )
    conn.execute(
        "UPDATE ingestion_runs "
        "SET status = 'superseded' "
        "WHERE source_file_id = ? AND id != ? AND status != 'superseded'",
        (source_file_id, new_run_id),
    )
    return prior_latest_id


def _is_match_sheet(sheet_name: str) -> bool:
    """Return True if the sheet name looks like a match-bearing sheet."""
    s = sheet_name.strip().lower()
    if re.match(r"day\s*\d+", s):
        return True
    if s in ("semi final", "semifinal", "semi-final", "final"):
        return True
    return False


def _derive_tournament_name(filename: str, wb) -> str:
    """Derive a human-readable tournament name from the file."""
    # Try [3,3] / [2,3] / [3,2] / [2,2] of the first match sheet.
    for sn in wb.sheetnames:
        if not _is_match_sheet(sn):
            continue
        ws = wb[sn]
        for r, c in ((3, 3), (2, 3), (3, 2), (2, 2), (1, 3), (1, 4)):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip() and v.strip().upper() != "VITTORIOSA LAWN TENNIS CLUB":
                # Strip surrounding quotes/whitespace and trailing periods.
                txt = v.strip().strip('“”"\'').strip().rstrip(".")
                if txt:
                    return txt
        break
    # Fallback: derive from filename (drop extension, collapse whitespace).
    base = os.path.splitext(filename)[0].strip()
    base = re.sub(r"\s+", " ", base)
    return base


def _derive_tournament_year(filename: str, fallback_iso_date: Optional[str]) -> int:
    """Find the tournament year. Prefer filename; fall back to first match date."""
    y = _detect_year_from_filename(filename)
    if y is not None:
        return y
    if fallback_iso_date is not None:
        try:
            return int(fallback_iso_date[:4])
        except ValueError:
            pass
    # Last-resort: current calendar year. Phase 0 won't hit this path on real files.
    return 2025


def parse(xlsx_path: str, db_conn: sqlite3.Connection) -> int:
    """Parse a team-tournament file into the DB. Returns the new ingestion_run_id."""
    sha256 = _sha256_of_file(xlsx_path)
    filename = os.path.basename(xlsx_path)

    with db_conn:
        club_id = _ensure_default_club(db_conn)

        # Reuse source_files row if (filename, sha256) already exists.
        existing = db_conn.execute(
            "SELECT id FROM source_files WHERE original_filename = ? AND sha256 = ? "
            "ORDER BY id DESC LIMIT 1",
            (filename, sha256),
        ).fetchone()
        if existing is not None:
            source_file_id = existing[0]
        else:
            cur = db_conn.execute(
                "INSERT INTO source_files (club_id, original_filename, sha256) VALUES (?, ?, ?)",
                (club_id, filename, sha256),
            )
            source_file_id = cur.lastrowid

        cur = db_conn.execute(
            "INSERT INTO ingestion_runs (source_file_id, status, agent_version) "
            "VALUES (?, 'running', ?)",
            (source_file_id, AGENT_VERSION),
        )
        ingestion_run_id = cur.lastrowid

        prior_run_id = _supersede_prior_runs_for_file(db_conn, source_file_id, ingestion_run_id)
        if prior_run_id is not None:
            db_conn.execute(
                "UPDATE ingestion_runs SET supersedes_run_id = ? WHERE id = ?",
                (prior_run_id, ingestion_run_id),
            )

        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=False)
        try:
            tournament_name = _derive_tournament_name(filename, wb)

            # Pre-pass: gather every match-sheet's date so we can compute the year fallback.
            sheet_dates: dict[str, Optional[str]] = {}
            for sn in wb.sheetnames:
                if not _is_match_sheet(sn):
                    continue
                sheet_dates[sn] = _find_sheet_date(wb[sn])

            first_date = next(
                (d for d in sheet_dates.values() if d is not None), None
            )
            tournament_year = _derive_tournament_year(filename, first_date)

            cur = db_conn.execute(
                "INSERT INTO tournaments (club_id, name, year, format, source_file_id) "
                "VALUES (?, ?, ?, ?, ?)",
                (club_id, tournament_name, tournament_year, TOURNAMENT_FORMAT, source_file_id),
            )
            tournament_id = cur.lastrowid

            quality_notes: list[str] = []
            n_matches_inserted = 0
            n_singles = 0
            n_walkovers = 0
            sheets_processed = []

            for sheet_name in wb.sheetnames:
                if not _is_match_sheet(sheet_name):
                    continue
                ws = wb[sheet_name]
                round_label = _round_label_for_sheet(sheet_name)
                sheet_date = sheet_dates.get(sheet_name)
                if sheet_date is None:
                    # Fall back to first sheet date or to year-Jan-1.
                    sheet_date = first_date or f"{tournament_year}-01-01"
                    quality_notes.append(
                        f"sheet {sheet_name!r}: no date found at usual cells; using {sheet_date}"
                    )

                panels = _find_court_panels(ws)
                if not panels:
                    continue

                sheet_match_count = 0

                # Build a list of header rows so we know where each panel ends.
                header_rows = [hr for (_, hr) in panels]

                for idx, (court_row, header_row) in enumerate(panels):
                    panels_below = header_rows[idx + 1:]
                    two_row = _detect_two_row_variant(ws, header_row)
                    for r, rubber_data in _iter_panel_rubbers(ws, header_row, two_row, panels_below):
                        if rubber_data is None:
                            continue
                        _insert_match(
                            db_conn,
                            tournament_id,
                            ingestion_run_id,
                            rubber_data["rubber"],
                            round_label,
                            sheet_date,
                            rubber_data,
                            source_file_id,
                        )
                        n_matches_inserted += 1
                        sheet_match_count += 1
                        if rubber_data.get("walkover"):
                            n_walkovers += 1
                        if rubber_data["side_a_p2"] is None and rubber_data["side_b_p2"] is None:
                            n_singles += 1

                sheets_processed.append({
                    "sheet": sheet_name,
                    "round": round_label,
                    "date": sheet_date,
                    "panels": len(panels),
                    "matches_inserted": sheet_match_count,
                })
        finally:
            wb.close()

        quality_report = {
            "n_matches_inserted": n_matches_inserted,
            "n_walkovers": n_walkovers,
            "n_singles": n_singles,
            "sheets_processed": sheets_processed,
            "notes": quality_notes,
        }

        db_conn.execute(
            "UPDATE ingestion_runs "
            "SET status = 'completed', completed_at = datetime('now'), quality_report_jsonb = ? "
            "WHERE id = ?",
            (json.dumps(quality_report), ingestion_run_id),
        )

    return ingestion_run_id
