"""Parser for VLTC Mixed Doubles (and same-template) Excel files.

Spec: `scripts/phase0/parser_spec_mixed_doubles.md`.

Files this parser handles:
  - `ESS Mixed Tournament Div and Results 2025.xlsx`
  - `ESS Mixed Tournament Div and Results 2024.xlsx`
  - `Elektra Mixed Tournament Div and Results 2023.xlsx`

These files share the SAME column layout as `sports_experience_2025.py`
(two side-by-side blocks at cols A–N and P–AC, `vs.` divider, set scores
at cols 2/5/8 and 17/20/23). The differences:

  - Sheet names: `Division 1`..`Division N` (no Men/Lad split).
  - Sub-block positions vary per file/division — must be discovered
    dynamically (not hard-coded as in SE 2025).
  - Pair separator is `' / '` with spaces (handled by existing splitter).
  - Mixed gender — `players.gender` is left NULL.
  - Final block uses single-row pair-string layout only.

Public API:
    parse(xlsx_path: str, db_conn: sqlite3.Connection) -> int
        Returns the new ingestion_run_id.

This parser reuses helpers from `sports_experience_2025.py` (the lower-level
match-extraction functions) to avoid logic duplication.
"""

from __future__ import annotations

import hashlib
import json
import os
import re
import sqlite3
import sys
from typing import Iterable, Optional

import openpyxl

# Path-relative imports — same pattern as sports_experience_2025.py.
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import players as players_mod  # noqa: E402

# Reuse helpers / constants from the SE parser.
from parsers import sports_experience_2025 as se  # noqa: E402

AGENT_VERSION = "phase0-mixed-doubles-parser-1.0"
TOURNAMENT_FORMAT = "doubles_division"
DEFAULT_CLUB_NAME = "VLTC"
DEFAULT_CLUB_SLUG = "vltc"

# Reuse SE column block definitions verbatim.
LEFT_BLOCK = se.LEFT_BLOCK
RIGHT_BLOCK = se.RIGHT_BLOCK


# ─────────────────────────────────────────────────────────────────────────────
# Tolerant block extractor (allows missing `vs.` in this block as long as
# the OTHER block has it — observed in some Mixed sheets, e.g. ESS 2025
# Division 6 row 9, where col 1 row 10 is empty but col 16 row 10 is `vs.`)
# ─────────────────────────────────────────────────────────────────────────────

def _extract_block_match_tolerant(ws, row: int, block: dict) -> Optional[dict]:
    """Extract one match from a block, allowing the `vs.` divider to be absent
    from THIS block as long as the other block has it on the same row.
    """
    name_col = block["name_col"]
    set1_col = block["set1_col"]
    set2_col = block["set2_col"]
    tie_col = block["tie_col"]

    pair_a_cell = ws.cell(row, name_col).value
    pair_b_cell = ws.cell(row + 2, name_col).value
    divider_this = ws.cell(row + 1, name_col).value
    divider_other = ws.cell(row + 1, 1 if name_col != 1 else 16).value

    if not se._is_pair_string(pair_a_cell):
        return None
    if not se._is_pair_string(pair_b_cell):
        return None
    if not (se._is_vs_divider(divider_this) or se._is_vs_divider(divider_other)):
        return None

    s1a = se._coerce_score(ws.cell(row, set1_col).value)
    s1b = se._coerce_score(ws.cell(row + 2, set1_col).value)
    s2a = se._coerce_score(ws.cell(row, set2_col).value)
    s2b = se._coerce_score(ws.cell(row + 2, set2_col).value)
    tba = se._coerce_score(ws.cell(row, tie_col).value)
    tbb = se._coerce_score(ws.cell(row + 2, tie_col).value)

    if all(x is None for x in (s1a, s1b, s2a, s2b, tba, tbb)):
        # Unplayed
        return None

    return {
        "pair_a": pair_a_cell.strip(),
        "pair_b": pair_b_cell.strip(),
        "set1_a": s1a,
        "set1_b": s1b,
        "set2_a": s2a,
        "set2_b": s2b,
        "tb_a": tba,
        "tb_b": tbb,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

# Label patterns: "Division 1", "Division 1 - Group A", "Division 1 - Round 2"
# Tolerant: case-insensitive, allows trailing whitespace.
_LABEL_RE = re.compile(
    r"^\s*Division\s+\d+(\s*-\s*(Group|Round)\s+[A-Za-z0-9]+)?\s*$",
    re.IGNORECASE,
)


def _is_block_label(value) -> bool:
    """True if the cell value matches a 'Division N' / 'Division N - Group X'
    / 'Division N - Round X' label pattern.
    """
    if not isinstance(value, str):
        return False
    return bool(_LABEL_RE.match(value))


def _extract_year_from_filename(filename: str) -> int:
    """Pull the last 4-digit year-like substring from the filename.

    Falls back to 0 if no match — caller should validate.
    """
    matches = re.findall(r"\b(20\d{2}|19\d{2})\b", filename)
    if not matches:
        return 0
    return int(matches[-1])


def _read_tournament_name(wb: openpyxl.Workbook) -> Optional[str]:
    """Read the tournament title from cell [1,1] of the first match sheet
    (the one whose name starts with 'Division').
    """
    for sn in wb.sheetnames:
        if not sn.lower().startswith("division"):
            continue
        ws = wb[sn]
        v = ws.cell(1, 1).value
        if isinstance(v, str) and v.strip():
            cleaned = v.replace("\xa0", " ")
            lines = [ln.strip() for ln in cleaned.split("\n") if ln.strip()]
            return lines[0] if lines else None
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Sub-block discovery
# ─────────────────────────────────────────────────────────────────────────────

def _find_sub_blocks(ws) -> list[tuple[str, int, int]]:
    """Discover sub-blocks (label, label_row, first_match_anchor_row) on a Division sheet.

    Strategy:
      1. Find every row where col 1 holds a 'Division N' / '...- Group X' /
         '...- Round X' label.
      2. For each label row, scan forward 1..8 rows for the first row where
         col 1 is a pair-string AND col 1 (or col 16) of the next row is `vs.`
         — that's the first match anchor for this sub-block.
      3. Defensive fallback: if no labels are found, treat row 9 (the SE 2025
         convention) as the only sub-block start using the sheet name itself
         as the label.

    Returns: list of (division_label_string, label_row, first_match_anchor_row)
    tuples, sorted by row.
    """
    label_anchors: list[tuple[str, int, int]] = []  # (label, label_row, first_anchor_row)

    # 1. Find all label rows in col 1.
    max_r = ws.max_row or 300
    for r in range(1, max_r + 1):
        v = ws.cell(r, 1).value
        if _is_block_label(v):
            label = v.strip()
            # Find first match anchor below this label. A 'match anchor' is a
            # row where col 1 is a pair string AND col 1 of row+2 is also a
            # pair string. The `vs.` divider on row+1 may be in col 1 OR col 16
            # (some sheets omit col 1's vs. for the first row band — only the
            # right-block divider is present).
            first_anchor = None
            for off in range(1, 8):
                rr = r + off
                if rr > max_r:
                    break
                cell_v = ws.cell(rr, 1).value
                pair_b = ws.cell(rr + 2, 1).value
                divider_left = ws.cell(rr + 1, 1).value
                divider_right = ws.cell(rr + 1, 16).value
                if (
                    se._is_pair_string(cell_v)
                    and se._is_pair_string(pair_b)
                    and (se._is_vs_divider(divider_left) or se._is_vs_divider(divider_right))
                ):
                    first_anchor = rr
                    break
            if first_anchor is not None:
                label_anchors.append((label, r, first_anchor))

    # 2. Defensive: if no labels were found via col 1, fall back to sheet
    # name as a single label and start at row 9.
    if not label_anchors:
        label_anchors.append((ws.title, 0, 9))

    return label_anchors


def _walk_anchors_to_end(ws, start_row: int) -> list[int]:
    """Walk match anchors from `start_row` stepping +4 while pattern holds.

    Returns the list of anchor rows that have a valid pair/vs/pair pattern
    in the LEFT block. Stops at first break.
    """
    anchors = []
    r = start_row
    max_r = ws.max_row or 300
    while r <= max_r:
        pair_a = ws.cell(r, LEFT_BLOCK["name_col"]).value
        if not se._is_pair_string(pair_a):
            # Some sheets place a vs. or pair B without a pair A on the
            # very last orphan row — stop here.
            break
        anchors.append(r)
        r += 4
    return anchors


# ─────────────────────────────────────────────────────────────────────────────
# Final block (single-row pair-string layout only)
# ─────────────────────────────────────────────────────────────────────────────

def _find_final_anchor(ws) -> Optional[int]:
    """Locate the row of the 'Final' label in col 16. None if no Final block."""
    max_r = ws.max_row or 300
    for r in range(1, max_r + 1):
        v = ws.cell(r, 16).value
        if isinstance(v, str) and v.strip().lower() == "final":
            return r
    return None


def _extract_mixed_final(ws, anchor_row: int) -> Optional[dict]:
    """Extract a Mixed-format Final block (single-row pair strings).

    Layout:
      [r,16]='Final'
      [r+2,16]='Players' / [r+3,16]='Score'  (headers)
      [r+4,16] = Pair A string (with '/' separator)
      [r+5,16] = 'vs.'
      [r+6,16] = Pair B string

    Scores live on the same rows as the pair-name strings (cols 17/20/23).
    """
    name_col = RIGHT_BLOCK["name_col"]
    set1_col = RIGHT_BLOCK["set1_col"]
    set2_col = RIGHT_BLOCK["set2_col"]
    tie_col = RIGHT_BLOCK["tie_col"]

    # Defensive: scan a few rows below the 'Final' anchor for the pair
    # strings (some files have slightly different offsets).
    pair_a_row = None
    for off in range(2, 8):
        v = ws.cell(anchor_row + off, name_col).value
        if isinstance(v, str) and "/" in v and not se._is_vs_divider(v):
            pair_a_row = anchor_row + off
            break
    if pair_a_row is None:
        return None

    # Find vs. divider below pair A row.
    vs_row = None
    for off in range(1, 4):
        v = ws.cell(pair_a_row + off, name_col).value
        if se._is_vs_divider(v):
            vs_row = pair_a_row + off
            break
    if vs_row is None:
        return None

    pair_b_row = vs_row + 1
    pair_b_v = ws.cell(pair_b_row, name_col).value
    if not isinstance(pair_b_v, str) or "/" not in pair_b_v:
        return None

    pair_a_str = ws.cell(pair_a_row, name_col).value.strip()
    pair_b_str = pair_b_v.strip()

    s1a = se._coerce_score(ws.cell(pair_a_row, set1_col).value)
    s1b = se._coerce_score(ws.cell(pair_b_row, set1_col).value)
    s2a = se._coerce_score(ws.cell(pair_a_row, set2_col).value)
    s2b = se._coerce_score(ws.cell(pair_b_row, set2_col).value)
    tba = se._coerce_score(ws.cell(pair_a_row, tie_col).value)
    tbb = se._coerce_score(ws.cell(pair_b_row, tie_col).value)

    if all(x is None for x in (s1a, s1b, s2a, s2b, tba, tbb)):
        return None

    return {
        "pair_a": pair_a_str,
        "pair_b": pair_b_str,
        "set1_a": s1a,
        "set1_b": s1b,
        "set2_a": s2a,
        "set2_b": s2b,
        "tb_a": tba,
        "tb_b": tbb,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Match → DB row insertion (Mixed: gender = None)
# ─────────────────────────────────────────────────────────────────────────────

def _insert_match(
    conn: sqlite3.Connection,
    tournament_id: int,
    ingestion_run_id: int,
    division: str,
    round_label: Optional[str],
    match_data: dict,
    source_file_id: int,
    placeholder_date: str,
) -> int:
    """Insert a match + its sides + per-set scores. Returns the match_id.

    This is a near-copy of `sports_experience_2025._insert_match` but does
    NOT set `players.gender` (mixed doubles has no per-sheet gender signal).
    """
    cur = conn.execute(
        "INSERT INTO matches (tournament_id, played_on, match_type, division, round, ingestion_run_id) "
        "VALUES (?, ?, 'doubles', ?, ?, ?)",
        (tournament_id, placeholder_date, division, round_label, ingestion_run_id),
    )
    match_id = cur.lastrowid

    a1_raw, a2_raw = se._split_pair(match_data["pair_a"])
    b1_raw, b2_raw = se._split_pair(match_data["pair_b"])
    a1_id = players_mod.get_or_create_player(conn, a1_raw, source_file_id)
    a2_id = players_mod.get_or_create_player(conn, a2_raw, source_file_id)
    b1_id = players_mod.get_or_create_player(conn, b1_raw, source_file_id)
    b2_id = players_mod.get_or_create_player(conn, b2_raw, source_file_id)
    # No gender set — Mixed doubles. Phase 1 reconciliation handles it.

    set_rows: list[tuple[int, int, int, int]] = []
    s1a, s1b = match_data["set1_a"], match_data["set1_b"]
    s2a, s2b = match_data["set2_a"], match_data["set2_b"]
    tba, tbb = match_data["tb_a"], match_data["tb_b"]

    if s1a is not None or s1b is not None:
        ga = s1a if s1a is not None else 0
        gb = s1b if s1b is not None else 0
        was_tb = (ga == 7 or gb == 7)
        set_rows.append((1, ga, gb, int(was_tb)))
    if s2a is not None or s2b is not None:
        ga = s2a if s2a is not None else 0
        gb = s2b if s2b is not None else 0
        was_tb = (ga == 7 or gb == 7)
        set_rows.append((2, ga, gb, int(was_tb)))
    if tba is not None or tbb is not None:
        ga = tba if tba is not None else 0
        gb = tbb if tbb is not None else 0
        next_set_no = 3 if len(set_rows) == 2 else (max(r[0] for r in set_rows) + 1 if set_rows else 1)
        set_rows.append((next_set_no, ga, gb, 1))

    for set_no, ga, gb, tb in set_rows:
        conn.execute(
            "INSERT INTO match_set_scores (match_id, set_number, side_a_games, side_b_games, was_tiebreak) "
            "VALUES (?, ?, ?, ?, ?)",
            (match_id, set_no, ga, gb, tb),
        )

    sets_won_a = sets_won_b = games_won_a = games_won_b = 0
    has_super_tb = False
    super_tb_a = super_tb_b = 0
    for set_no, ga, gb, tb in set_rows:
        if set_no <= 2:
            games_won_a += ga
            games_won_b += gb
            if ga > gb:
                sets_won_a += 1
            elif gb > ga:
                sets_won_b += 1
        else:
            has_super_tb = True
            super_tb_a, super_tb_b = ga, gb

    if sets_won_a > sets_won_b:
        won_a, won_b = 1, 0
    elif sets_won_b > sets_won_a:
        won_a, won_b = 0, 1
    elif has_super_tb:
        if super_tb_a > super_tb_b:
            won_a, won_b = 1, 0
        elif super_tb_b > super_tb_a:
            won_a, won_b = 0, 1
        else:
            won_a, won_b = 0, 0
    else:
        won_a, won_b = 0, 0

    conn.execute(
        "INSERT INTO match_sides (match_id, side, player1_id, player2_id, sets_won, games_won, won) "
        "VALUES (?, 'A', ?, ?, ?, ?, ?)",
        (match_id, a1_id, a2_id, sets_won_a, games_won_a, won_a),
    )
    conn.execute(
        "INSERT INTO match_sides (match_id, side, player1_id, player2_id, sets_won, games_won, won) "
        "VALUES (?, 'B', ?, ?, ?, ?, ?)",
        (match_id, b1_id, b2_id, sets_won_b, games_won_b, won_b),
    )

    return match_id


# ─────────────────────────────────────────────────────────────────────────────
# Top-level parse
# ─────────────────────────────────────────────────────────────────────────────

def _ensure_default_club(conn: sqlite3.Connection) -> int:
    row = conn.execute("SELECT id FROM clubs WHERE slug = ?", (DEFAULT_CLUB_SLUG,)).fetchone()
    if row is not None:
        return row[0]
    cur = conn.execute(
        "INSERT INTO clubs (name, slug) VALUES (?, ?)",
        (DEFAULT_CLUB_NAME, DEFAULT_CLUB_SLUG),
    )
    return cur.lastrowid


def parse(xlsx_path: str, db_conn: sqlite3.Connection) -> int:
    """Parse a Mixed Doubles (or same-template division) file into the DB.

    Returns the new ingestion_run_id.
    """
    sha256 = se._sha256_of_file(xlsx_path)
    filename = os.path.basename(xlsx_path)
    year = _extract_year_from_filename(filename)
    if year == 0:
        # Fall back to current year if filename has no year — shouldn't happen
        # for the known files but be defensive.
        year = 2024
    placeholder_date = f"{year}-01-01"

    with db_conn:
        club_id = _ensure_default_club(db_conn)

        # source_files row
        existing = db_conn.execute(
            "SELECT id FROM source_files WHERE original_filename = ? AND sha256 = ? "
            "ORDER BY id DESC LIMIT 1",
            (filename, sha256),
        ).fetchone()
        if existing is not None:
            source_file_id = existing[0]
        else:
            cur = db_conn.execute(
                "INSERT INTO source_files (club_id, original_filename, sha256) VALUES (?, ?, ?)",
                (club_id, filename, sha256),
            )
            source_file_id = cur.lastrowid

        # ingestion_runs row
        cur = db_conn.execute(
            "INSERT INTO ingestion_runs (source_file_id, status, agent_version) "
            "VALUES (?, 'running', ?)",
            (source_file_id, AGENT_VERSION),
        )
        ingestion_run_id = cur.lastrowid

        # Supersede prior runs for this file
        prior_run_id = se._supersede_prior_runs_for_file(db_conn, source_file_id, ingestion_run_id)
        if prior_run_id is not None:
            db_conn.execute(
                "UPDATE ingestion_runs SET supersedes_run_id = ? WHERE id = ?",
                (prior_run_id, ingestion_run_id),
            )

        # Open workbook (need full mode for random cell access)
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        try:
            tournament_name = _read_tournament_name(wb) or filename.rsplit(".", 1)[0]

            cur = db_conn.execute(
                "INSERT INTO tournaments (club_id, name, year, format, source_file_id) "
                "VALUES (?, ?, ?, ?, ?)",
                (club_id, tournament_name, year, TOURNAMENT_FORMAT, source_file_id),
            )
            tournament_id = cur.lastrowid

            skipped_unplayed: list[dict] = []
            n_matches_inserted = 0

            for sheet_name in wb.sheetnames:
                if not sheet_name.lower().startswith("division"):
                    continue
                ws = wb[sheet_name]

                sub_blocks = _find_sub_blocks(ws)
                # Track which anchor rows we've already processed so a later
                # defensive sweep doesn't double-walk them.
                walked_anchor_rows: set[int] = set()

                for label, _label_row, first_anchor in sub_blocks:
                    anchors = _walk_anchors_to_end(ws, first_anchor)
                    for r in anchors:
                        walked_anchor_rows.add(r)
                        # Extract LEFT block match
                        left = _extract_block_match_tolerant(ws, r, LEFT_BLOCK)
                        if left is not None:
                            _insert_match(
                                db_conn,
                                tournament_id,
                                ingestion_run_id,
                                label,
                                None,
                                left,
                                source_file_id,
                                placeholder_date,
                            )
                            n_matches_inserted += 1
                        else:
                            # Could be unplayed; check for the pair/vs/pair shape
                            if (
                                se._is_pair_string(ws.cell(r, LEFT_BLOCK["name_col"]).value)
                                and se._is_vs_divider(ws.cell(r + 1, LEFT_BLOCK["name_col"]).value)
                                and se._is_pair_string(ws.cell(r + 2, LEFT_BLOCK["name_col"]).value)
                            ):
                                skipped_unplayed.append({
                                    "sheet": sheet_name,
                                    "row": r,
                                    "block": "left",
                                    "division": label,
                                    "pair_A": str(ws.cell(r, LEFT_BLOCK["name_col"]).value).strip(),
                                    "pair_B": str(ws.cell(r + 2, LEFT_BLOCK["name_col"]).value).strip(),
                                    "reason": "no scores recorded",
                                })

                        # Extract RIGHT block match
                        right = _extract_block_match_tolerant(ws, r, RIGHT_BLOCK)
                        if right is not None:
                            _insert_match(
                                db_conn,
                                tournament_id,
                                ingestion_run_id,
                                label,
                                None,
                                right,
                                source_file_id,
                                placeholder_date,
                            )
                            n_matches_inserted += 1
                        else:
                            if (
                                se._is_pair_string(ws.cell(r, RIGHT_BLOCK["name_col"]).value)
                                and se._is_vs_divider(ws.cell(r + 1, RIGHT_BLOCK["name_col"]).value)
                                and se._is_pair_string(ws.cell(r + 2, RIGHT_BLOCK["name_col"]).value)
                            ):
                                skipped_unplayed.append({
                                    "sheet": sheet_name,
                                    "row": r,
                                    "block": "right",
                                    "division": label,
                                    "pair_A": str(ws.cell(r, RIGHT_BLOCK["name_col"]).value).strip(),
                                    "pair_B": str(ws.cell(r + 2, RIGHT_BLOCK["name_col"]).value).strip(),
                                    "reason": "no scores recorded",
                                })

                # Defensive sweep for un-labeled sub-blocks (e.g. ESS 2024 Div 6
                # Round 2 has no label row; just a sub-header at row 33 and the
                # first match at row 34). Find any pair-string + vs. + pair-string
                # row band whose anchor wasn't walked above. Attribute orphans
                # to the most recent labelled sub-block whose label_row is
                # above the orphan.
                max_r = ws.max_row or 300
                for r in range(1, max_r - 1):
                    if r in walked_anchor_rows:
                        continue
                    pair_a = ws.cell(r, LEFT_BLOCK["name_col"]).value
                    pair_b = ws.cell(r + 2, LEFT_BLOCK["name_col"]).value
                    divider_left = ws.cell(r + 1, LEFT_BLOCK["name_col"]).value
                    divider_right = ws.cell(r + 1, RIGHT_BLOCK["name_col"]).value
                    if not (
                        se._is_pair_string(pair_a)
                        and se._is_pair_string(pair_b)
                        and (se._is_vs_divider(divider_left) or se._is_vs_divider(divider_right))
                    ):
                        continue

                    # Most recent label-row above this row — attribute orphan
                    # to that division/group/round.
                    applicable_label = sheet_name
                    for lab, lrow, _anchor in sub_blocks:
                        if lrow < r:
                            applicable_label = lab

                    # Walk the orphan run.
                    orphan_anchors = _walk_anchors_to_end(ws, r)
                    for rr in orphan_anchors:
                        walked_anchor_rows.add(rr)
                        left = _extract_block_match_tolerant(ws, rr, LEFT_BLOCK)
                        if left is not None:
                            _insert_match(
                                db_conn,
                                tournament_id,
                                ingestion_run_id,
                                applicable_label,
                                None,
                                left,
                                source_file_id,
                                placeholder_date,
                            )
                            n_matches_inserted += 1
                        right = _extract_block_match_tolerant(ws, rr, RIGHT_BLOCK)
                        if right is not None:
                            _insert_match(
                                db_conn,
                                tournament_id,
                                ingestion_run_id,
                                applicable_label,
                                None,
                                right,
                                source_file_id,
                                placeholder_date,
                            )
                            n_matches_inserted += 1

                # Final block (single-row pair-string layout)
                final_anchor = _find_final_anchor(ws)
                if final_anchor is not None:
                    final_match = _extract_mixed_final(ws, final_anchor)
                    if final_match is not None:
                        # Strip group/round suffix from the division for the final.
                        # Use the leading 'Division N' part of the most-relevant
                        # label, if present in sub_blocks; otherwise use sheet name.
                        if sub_blocks:
                            base_label = sub_blocks[0][0]
                            base_label = re.sub(
                                r"\s*-\s*(Group|Round)\s+.*$",
                                "",
                                base_label,
                                flags=re.IGNORECASE,
                            ).strip()
                        else:
                            base_label = sheet_name
                        _insert_match(
                            db_conn,
                            tournament_id,
                            ingestion_run_id,
                            base_label,
                            "final",
                            final_match,
                            source_file_id,
                            placeholder_date,
                        )
                        n_matches_inserted += 1

            quality_report = {
                "n_matches_inserted": n_matches_inserted,
                "skipped_unplayed_matches": skipped_unplayed,
                "placeholder_date_used": placeholder_date,
                "notes": [
                    "File contains no per-match dates; played_on is a placeholder.",
                    "Mixed doubles — players.gender left NULL.",
                ],
            }

            db_conn.execute(
                "UPDATE ingestion_runs "
                "SET status = 'completed', completed_at = datetime('now'), quality_report_jsonb = ? "
                "WHERE id = ?",
                (json.dumps(quality_report), ingestion_run_id),
            )
        finally:
            wb.close()

    return ingestion_run_id
