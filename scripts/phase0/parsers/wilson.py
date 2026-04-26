"""Parser for Wilson Autumn / Spring Team Tournaments (2017–2021) — T-P0-014.

Spec: `scripts/phase0/parser_spec_wilson.md`.

This parser handles the older VLTC team-tournament template family used for
Wilson Autumn (2017–2021) and Wilson Spring (2018, 2019). Format:
  - 6 teams (A–F) play each other across multiple "Day N" sheets.
  - Each Day-N sheet contains 3–4 court blocks; each court hosts 8 rubbers
    (4 vs 4 between two teams).
  - Each rubber is a doubles match across 2 score-rows (set 1 on row r, set 2
    on row r+1) plus optional super-tiebreak.
  - Semi Final and Final sheets use the same rubber layout.

Pre-resolved decisions (per spec):
  1. The `'Time'` header anchor column varies (col 2 vs col 3) — even within
     the same workbook. We auto-detect by scanning rows 1–15, cols 1–5.
  2. Tied rubbers (1-1 sets, no super-TB string) are recorded with both sides
     `won=0`, `sets_won=1` — flagged in quality report.
  3. Retirements (`'ret'` in set-2 cells) are recorded as 1-set wins by the
     set-1 winner with `match.walkover=1`.
  4. Date placeholders: Day-N → year-09-01 + 7*(N-1) (Autumn) or year-04-01
     (Spring); Semi Final → year-10-15 / year-05-15; Final → year-10-30 /
     year-05-30. Real dates parsed from day-headers when present.
  5. Templates and the in-cell year typo in Wilson 2017 (which says "2015") are
     IGNORED — year always comes from the filename.

`.xls` (legacy) and `.xlsx` are handled via a uniform Workbook adapter so the
match-extraction logic is format-agnostic. xlrd is required for `.xls`.

Public API:
    parse(xlsx_path: str, db_conn: sqlite3.Connection) -> int
        Returns the new ingestion_run_id.
"""

from __future__ import annotations

import hashlib
import json
import os
import re
import sqlite3
import sys
from typing import Optional

# Path-relative import — works for both `python -m scripts.phase0.parsers.<x>`
# and module-level `import` from cli.py / tests.
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import players as players_mod  # noqa: E402

AGENT_VERSION = "phase0-wilson-parser-1.0"
TOURNAMENT_FORMAT = "doubles_team"
DEFAULT_CLUB_NAME = "VLTC"
DEFAULT_CLUB_SLUG = "vltc"

# Rubber-category prefix → gender hint used to set players.gender on first
# sight. Mixed doubles ('Mxd', 'MXD') is conservatively skipped.
_GENDER_PREFIXES_M = ("men ", "men")
_GENDER_PREFIXES_F = ("lad ", "lad", "ladies")

# Regex for parsing optional 'T.B. 7-9' or 'TB 9-7' style super-tiebreak strings.
_TB_RE = re.compile(r"T\.?\s*B\.?\s*(\d+)\s*[-/]\s*(\d+)", re.IGNORECASE)

# Regex to extract a date from day-header strings like 'DAY 1 - 3/11/2021'.
_DATE_RE = re.compile(r"(\d{1,2})/(\d{1,2})/(\d{4})")

# Year regex for filename → tournament year.
_FILE_YEAR_RE = re.compile(r"(20\d{2})")


# ─────────────────────────────────────────────────────────────────────────────
# Workbook adapter — uniform .xls / .xlsx access
# ─────────────────────────────────────────────────────────────────────────────

class _Workbook:
    """Minimal workbook adapter so the rest of the parser is .xls/.xlsx agnostic.

    Provides `.sheet_names()` and `.cell(sheet_name, row, col)` (1-indexed).
    Returns Python None for blank cells; ints/floats/strings otherwise.
    """

    def __init__(self, path: str) -> None:
        self.path = path
        self._is_xls = path.lower().endswith(".xls")
        if self._is_xls:
            import xlrd  # type: ignore
            self._wb = xlrd.open_workbook(path)
            self._sheets_by_name = {s.name: s for s in self._wb.sheets()}
        else:
            import openpyxl
            # data_only=True so formula cells return cached values.
            self._wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
            self._sheets_by_name = {sn: self._wb[sn] for sn in self._wb.sheetnames}

    def sheet_names(self) -> list[str]:
        return list(self._sheets_by_name.keys())

    def cell(self, sheet_name: str, row: int, col: int):
        """Return the cell value at (row, col) — 1-indexed in both axes."""
        ws = self._sheets_by_name[sheet_name]
        if self._is_xls:
            try:
                v = ws.cell_value(row - 1, col - 1)
            except IndexError:
                return None
            # xlrd represents blank as empty string. Coerce to None for symmetry.
            if v == "":
                return None
            return v
        # openpyxl: 1-indexed natively.
        v = ws.cell(row, col).value
        if v == "":
            return None
        return v

    def max_row(self, sheet_name: str) -> int:
        ws = self._sheets_by_name[sheet_name]
        return ws.nrows if self._is_xls else (ws.max_row or 0)

    def max_col(self, sheet_name: str) -> int:
        ws = self._sheets_by_name[sheet_name]
        return ws.ncols if self._is_xls else (ws.max_column or 0)

    def close(self) -> None:
        if not self._is_xls:
            self._wb.close()


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _sha256_of_file(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _coerce_score(value) -> Optional[int]:
    """Coerce an Excel score cell to int. Returns None for blank cells."""
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return int(value)
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        try:
            return int(float(s))
        except ValueError:
            return None
    return None


def _is_retirement_marker(value) -> bool:
    if not isinstance(value, str):
        return False
    return value.strip().lower() == "ret"


def _is_player_name(value) -> bool:
    """A player-name cell is a non-empty string that isn't a header / 'vs' /
    a totals row label / a time string.
    """
    if not isinstance(value, str):
        return False
    s = value.strip()
    if not s:
        return False
    sl = s.lower()
    if sl in ("vs", "vs.", "v"):
        return False
    if sl.startswith("total "):
        return False
    if sl == "time" or sl == "rubber" or sl == "games" or sl == "sets":
        return False
    # Time strings like '6.30 pm', '7.45pm' — these belong in the Time col,
    # not the player col. Reject anything matching a time pattern.
    if re.match(r"^\d{1,2}[\.:]\d{2}\s*(am|pm)?$", sl):
        return False
    if re.match(r"^\d{1,2}\s*(am|pm)$", sl):
        return False
    # Court / day labels.
    if sl.startswith("court ") or sl.startswith("day "):
        return False
    # Team labels.
    if "team " in sl and ":" in sl:
        return False
    return True


def _extract_year_from_filename(filename: str) -> int:
    m = _FILE_YEAR_RE.search(filename)
    if m is None:
        raise ValueError(f"could not extract year from filename: {filename!r}")
    return int(m.group(1))


def _classify_season(filename: str) -> str:
    """Return 'autumn' or 'spring' from filename."""
    fn = filename.lower()
    if "spring" in fn:
        return "spring"
    return "autumn"  # default — Autumn is the more common variant


def _tournament_name(filename: str) -> str:
    season = _classify_season(filename)
    year = _extract_year_from_filename(filename)
    season_pretty = "Spring" if season == "spring" else "Autumn"
    return f"Wilson {season_pretty} Team Tournament {year}"


def _placeholder_date(season: str, year: int, sheet_name: str) -> str:
    """Return an ISO date placeholder for a given Day-N / Semi Final / Final sheet."""
    sn = sheet_name.strip().lower()
    if sn.startswith("day"):
        # Try to extract day number.
        m = re.search(r"day\s*(\d+)", sn)
        n = int(m.group(1)) if m else 1
        if season == "spring":
            base_month, base_day = 4, 1
        else:
            base_month, base_day = 9, 1
        # day1 = base, day2 = base+7, day5 = base+28
        from datetime import date, timedelta
        d = date(year, base_month, base_day) + timedelta(days=7 * (n - 1))
        return d.isoformat()
    if "semi" in sn:
        return f"{year}-{'05' if season == 'spring' else '10'}-15"
    if sn == "final" or sn.startswith("final"):
        return f"{year}-{'05' if season == 'spring' else '10'}-30"
    return f"{year}-09-15"


def _extract_real_date(day_header_value, year: int) -> Optional[str]:
    """Extract a real ISO date from a day-header string like 'DAY 1 - 3/11/2021'.

    Returns None if no date is parseable. Ignores ranges (uses the first date).
    """
    if not isinstance(day_header_value, str):
        return None
    m = _DATE_RE.search(day_header_value)
    if m is None:
        return None
    d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
    # Sanity check year matches filename year (or close to it — Wilson 2021's
    # Semi Final has '21/1/22' which is 2022).
    if not (year - 1 <= y <= year + 1):
        return None
    try:
        from datetime import date
        return date(y, mo, d).isoformat()
    except ValueError:
        return None


# ─────────────────────────────────────────────────────────────────────────────
# Per-sheet column anchor detection
# ─────────────────────────────────────────────────────────────────────────────

def _find_time_anchor(wb: _Workbook, sheet_name: str) -> Optional[tuple[int, int]]:
    """Scan rows 1–15, cols 1–5 for a literal `'Time'` header cell.

    Returns (header_row, time_col) or None if not found. The first occurrence
    wins; subsequent court blocks share the same column anchor.
    """
    for r in range(1, 16):
        for c in range(1, 6):
            v = wb.cell(sheet_name, r, c)
            if isinstance(v, str) and v.strip().lower() == "time":
                return (r, c)
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Match (rubber) extraction
# ─────────────────────────────────────────────────────────────────────────────

def _extract_rubber(
    wb: _Workbook,
    sheet_name: str,
    row: int,
    time_col: int,
) -> Optional[dict]:
    """Extract one rubber starting at `row`. Returns a dict or None.

    Layout (offsets from time_col):
      +1=Rubber, +2=A1, +3=A2, +4=vs, +5=B1, +6=B2,
      +7=A games (set 1 on row r, set 2 on row r+1),
      +8=B games (set 1 on row r, set 2 on row r+1),
      +9=A sets-won, +10=B sets-won
      (optional) +11=tiebreak string ('T.B. 7-9') on row r+1
    """
    rubber_col = time_col + 1
    a1_col = time_col + 2
    a2_col = time_col + 3
    b1_col = time_col + 5
    b2_col = time_col + 6
    games_a_col = time_col + 7
    games_b_col = time_col + 8
    tb_col = time_col + 11  # optional

    a1 = wb.cell(sheet_name, row, a1_col)
    a2 = wb.cell(sheet_name, row, a2_col)
    b1 = wb.cell(sheet_name, row, b1_col)
    b2 = wb.cell(sheet_name, row, b2_col)

    # All four name cells must be valid player-name strings.
    if not (_is_player_name(a1) and _is_player_name(a2) and _is_player_name(b1) and _is_player_name(b2)):
        return None

    rubber_cat_raw = wb.cell(sheet_name, row, rubber_col)
    rubber_cat = ""
    if isinstance(rubber_cat_raw, str):
        # Normalize whitespace; preserve case for downstream display.
        rubber_cat = re.sub(r"\s+", " ", rubber_cat_raw).strip()

    # Set scores
    s1a_raw = wb.cell(sheet_name, row, games_a_col)
    s1b_raw = wb.cell(sheet_name, row, games_b_col)
    s2a_raw = wb.cell(sheet_name, row + 1, games_a_col)
    s2b_raw = wb.cell(sheet_name, row + 1, games_b_col)

    # Retirement detection — any 'ret' in set 2 cells.
    retired = _is_retirement_marker(s2a_raw) or _is_retirement_marker(s2b_raw)

    s1a = _coerce_score(s1a_raw)
    s1b = _coerce_score(s1b_raw)
    s2a = _coerce_score(s2a_raw) if not retired else None
    s2b = _coerce_score(s2b_raw) if not retired else None

    # Optional tiebreak cell (Wilson 2020 Semi Final has 'T.B. 7-9').
    tb_a = tb_b = None
    tb_raw = wb.cell(sheet_name, row + 1, tb_col)
    if isinstance(tb_raw, str):
        m = _TB_RE.search(tb_raw)
        if m is not None:
            tb_a, tb_b = int(m.group(1)), int(m.group(2))

    # Unplayed detection: all set cells blank/0 → skip.
    if (
        s1a in (None, 0)
        and s1b in (None, 0)
        and s2a in (None, 0)
        and s2b in (None, 0)
        and tb_a is None
        and tb_b is None
    ):
        # If all four set cells are explicitly None (truly empty) — skip.
        if s1a is None and s1b is None and s2a is None and s2b is None:
            return {
                "_skip_unplayed": True,
                "rubber": rubber_cat,
                "pair_a": (str(a1).strip(), str(a2).strip()),
                "pair_b": (str(b1).strip(), str(b2).strip()),
            }
        # 0-0 0-0 is technically "played but no games won by either side" —
        # treat as unplayed too (defensive; appears in Wilson 2020 Semi Final).

    return {
        "rubber": rubber_cat,
        "pair_a": (str(a1).strip(), str(a2).strip()),
        "pair_b": (str(b1).strip(), str(b2).strip()),
        "set1_a": s1a,
        "set1_b": s1b,
        "set2_a": s2a,
        "set2_b": s2b,
        "tb_a": tb_a,
        "tb_b": tb_b,
        "retired": retired,
    }


def _iter_rubbers(wb: _Workbook, sheet_name: str, time_col: int, header_row: int):
    """Yield (row, rubber_dict) for every rubber in the sheet.

    Walks every row from header_row+1 to max_row, attempting `_extract_rubber`
    at each. Skips rows that don't contain rubber data (totals, blanks, court
    headers, etc.).
    """
    max_r = wb.max_row(sheet_name)
    r = header_row + 1
    while r <= max_r:
        # Quick check: is there a player name at offset +2?
        peek = wb.cell(sheet_name, r, time_col + 2)
        if _is_player_name(peek):
            rubber = _extract_rubber(wb, sheet_name, r, time_col)
            if rubber is not None:
                yield r, rubber
                r += 2  # Skip past set-2 row
                continue
        r += 1


# ─────────────────────────────────────────────────────────────────────────────
# Match → DB row insertion
# ─────────────────────────────────────────────────────────────────────────────

def _gender_for_rubber(rubber: str) -> Optional[str]:
    """Return 'M'/'F'/None inferred from rubber category."""
    if not rubber:
        return None
    rl = rubber.lower()
    if rl.startswith("men"):
        return "M"
    if rl.startswith("lad"):
        return "F"
    return None  # Mixed (Mxd / MXD) — conservatively skip


def _round_for_sheet(sheet_name: str) -> Optional[str]:
    sn = sheet_name.strip().lower()
    if "semi" in sn:
        return "semi-final"
    if sn == "final" or sn.startswith("final"):
        return "final"
    return None


def _insert_match(
    conn: sqlite3.Connection,
    tournament_id: int,
    ingestion_run_id: int,
    division: str,
    round_label: Optional[str],
    played_on: str,
    rubber_data: dict,
    source_file_id: int,
    quality_report: dict,
    sheet_name: str,
    row: int,
) -> Optional[int]:
    """Insert one rubber as a doubles match. Returns match_id or None if skipped."""
    a1_raw, a2_raw = rubber_data["pair_a"]
    b1_raw, b2_raw = rubber_data["pair_b"]
    s1a, s1b = rubber_data["set1_a"], rubber_data["set1_b"]
    s2a, s2b = rubber_data["set2_a"], rubber_data["set2_b"]
    tb_a, tb_b = rubber_data["tb_a"], rubber_data["tb_b"]
    retired = rubber_data["retired"]

    # Resolve gender from rubber category (if Men/Lad).
    gender = _gender_for_rubber(division)
    walkover = 1 if retired else 0

    cur = conn.execute(
        "INSERT INTO matches (tournament_id, played_on, match_type, division, round, "
        "ingestion_run_id, walkover) "
        "VALUES (?, ?, 'doubles', ?, ?, ?, ?)",
        (tournament_id, played_on, division, round_label, ingestion_run_id, walkover),
    )
    match_id = cur.lastrowid

    a1_id = players_mod.get_or_create_player(conn, a1_raw, source_file_id)
    a2_id = players_mod.get_or_create_player(conn, a2_raw, source_file_id)
    b1_id = players_mod.get_or_create_player(conn, b1_raw, source_file_id)
    b2_id = players_mod.get_or_create_player(conn, b2_raw, source_file_id)

    if gender is not None:
        for pid in (a1_id, a2_id, b1_id, b2_id):
            conn.execute(
                "UPDATE players SET gender = ? WHERE id = ? AND gender IS NULL",
                (gender, pid),
            )

    # Build set rows. Set 1 is always present (we wouldn't be here if
    # all-blank). Set 2 only if non-blank. Super-tb only if parsed.
    set_rows: list[tuple[int, int, int, int]] = []  # (set_no, ga, gb, was_tb)
    if s1a is not None or s1b is not None:
        ga = s1a if s1a is not None else 0
        gb = s1b if s1b is not None else 0
        set_rows.append((1, ga, gb, 1 if (ga == 7 or gb == 7) else 0))
    if s2a is not None or s2b is not None:
        ga = s2a if s2a is not None else 0
        gb = s2b if s2b is not None else 0
        set_rows.append((2, ga, gb, 1 if (ga == 7 or gb == 7) else 0))
    if tb_a is not None or tb_b is not None:
        ga = tb_a if tb_a is not None else 0
        gb = tb_b if tb_b is not None else 0
        next_set_no = (max(r[0] for r in set_rows) + 1) if set_rows else 1
        set_rows.append((next_set_no, ga, gb, 1))

    for set_no, ga, gb, tb in set_rows:
        conn.execute(
            "INSERT INTO match_set_scores (match_id, set_number, side_a_games, side_b_games, was_tiebreak) "
            "VALUES (?, ?, ?, ?, ?)",
            (match_id, set_no, ga, gb, tb),
        )

    # Per-side aggregates: regular sets only.
    sets_won_a = sets_won_b = 0
    games_won_a = games_won_b = 0
    has_super_tb = False
    super_tb_a = super_tb_b = 0
    for set_no, ga, gb, tb in set_rows:
        if set_no <= 2:
            games_won_a += ga
            games_won_b += gb
            if ga > gb:
                sets_won_a += 1
            elif gb > ga:
                sets_won_b += 1
        else:
            has_super_tb = True
            super_tb_a, super_tb_b = ga, gb

    # Determine winner.
    if retired:
        # Side that won set 1 wins by retirement.
        if (s1a or 0) > (s1b or 0):
            won_a, won_b = 1, 0
        elif (s1b or 0) > (s1a or 0):
            won_a, won_b = 0, 1
        else:
            won_a, won_b = 0, 0
    elif sets_won_a > sets_won_b:
        won_a, won_b = 1, 0
    elif sets_won_b > sets_won_a:
        won_a, won_b = 0, 1
    elif has_super_tb:
        if super_tb_a > super_tb_b:
            won_a, won_b = 1, 0
        elif super_tb_b > super_tb_a:
            won_a, won_b = 0, 1
        else:
            won_a, won_b = 0, 0
    else:
        # Tied 1-1, no super-tb cell — undecided. Log to quality report.
        won_a, won_b = 0, 0
        quality_report["tied_rubbers_undecided"].append({
            "sheet": sheet_name,
            "row": row,
            "division": division,
            "pair_A": f"{a1_raw}/{a2_raw}",
            "pair_B": f"{b1_raw}/{b2_raw}",
        })

    if retired:
        quality_report["retired_rubbers"].append({
            "sheet": sheet_name,
            "row": row,
            "division": division,
            "pair_A": f"{a1_raw}/{a2_raw}",
            "pair_B": f"{b1_raw}/{b2_raw}",
        })

    conn.execute(
        "INSERT INTO match_sides (match_id, side, player1_id, player2_id, sets_won, games_won, won) "
        "VALUES (?, 'A', ?, ?, ?, ?, ?)",
        (match_id, a1_id, a2_id, sets_won_a, games_won_a, won_a),
    )
    conn.execute(
        "INSERT INTO match_sides (match_id, side, player1_id, player2_id, sets_won, games_won, won) "
        "VALUES (?, 'B', ?, ?, ?, ?, ?)",
        (match_id, b1_id, b2_id, sets_won_b, games_won_b, won_b),
    )
    return match_id


# ─────────────────────────────────────────────────────────────────────────────
# Top-level parse
# ─────────────────────────────────────────────────────────────────────────────

def _ensure_default_club(conn: sqlite3.Connection) -> int:
    row = conn.execute("SELECT id FROM clubs WHERE slug = ?", (DEFAULT_CLUB_SLUG,)).fetchone()
    if row is not None:
        return row[0]
    cur = conn.execute(
        "INSERT INTO clubs (name, slug) VALUES (?, ?)",
        (DEFAULT_CLUB_NAME, DEFAULT_CLUB_SLUG),
    )
    return cur.lastrowid


def _supersede_prior_runs_for_file(
    conn: sqlite3.Connection, source_file_id: int, new_run_id: int
) -> Optional[int]:
    row = conn.execute(
        "SELECT id FROM ingestion_runs "
        "WHERE source_file_id = ? AND id != ? "
        "ORDER BY id DESC LIMIT 1",
        (source_file_id, new_run_id),
    ).fetchone()
    if row is None:
        return None
    prior_latest_id = row[0]
    conn.execute(
        "UPDATE matches "
        "SET superseded_by_run_id = ? "
        "WHERE ingestion_run_id IN ("
        "    SELECT id FROM ingestion_runs WHERE source_file_id = ? AND id != ?"
        ") AND superseded_by_run_id IS NULL",
        (new_run_id, source_file_id, new_run_id),
    )
    conn.execute(
        "UPDATE ingestion_runs "
        "SET status = 'superseded' "
        "WHERE source_file_id = ? AND id != ? AND status != 'superseded'",
        (source_file_id, new_run_id),
    )
    return prior_latest_id


def parse(xlsx_path: str, db_conn: sqlite3.Connection) -> int:
    """Parse a Wilson team-tournament file into the DB.

    Accepts both `.xlsx` (2019, 2020, 2021) and `.xls` (2017, 2018; 2018/2019
    Spring) files. Skips template files (names containing 'Template').

    Returns the new ingestion_run_id.
    """
    filename = os.path.basename(xlsx_path)
    if "template" in filename.lower():
        raise ValueError(f"refusing to parse template file: {filename!r}")

    sha256 = _sha256_of_file(xlsx_path)
    year = _extract_year_from_filename(filename)
    season = _classify_season(filename)
    tournament_name = _tournament_name(filename)

    with db_conn:
        club_id = _ensure_default_club(db_conn)

        # source_files row — reuse if exact (filename, sha) match.
        existing = db_conn.execute(
            "SELECT id FROM source_files WHERE original_filename = ? AND sha256 = ? "
            "ORDER BY id DESC LIMIT 1",
            (filename, sha256),
        ).fetchone()
        if existing is not None:
            source_file_id = existing[0]
        else:
            cur = db_conn.execute(
                "INSERT INTO source_files (club_id, original_filename, sha256) VALUES (?, ?, ?)",
                (club_id, filename, sha256),
            )
            source_file_id = cur.lastrowid

        cur = db_conn.execute(
            "INSERT INTO ingestion_runs (source_file_id, status, agent_version) "
            "VALUES (?, 'running', ?)",
            (source_file_id, AGENT_VERSION),
        )
        ingestion_run_id = cur.lastrowid

        prior_run_id = _supersede_prior_runs_for_file(db_conn, source_file_id, ingestion_run_id)
        if prior_run_id is not None:
            db_conn.execute(
                "UPDATE ingestion_runs SET supersedes_run_id = ? WHERE id = ?",
                (prior_run_id, ingestion_run_id),
            )

        cur = db_conn.execute(
            "INSERT INTO tournaments (club_id, name, year, format, source_file_id) "
            "VALUES (?, ?, ?, ?, ?)",
            (club_id, tournament_name, year, TOURNAMENT_FORMAT, source_file_id),
        )
        tournament_id = cur.lastrowid

        wb = _Workbook(xlsx_path)
        try:
            quality_report = {
                "tournament_year_from_filename": year,
                "n_matches_inserted": 0,
                "n_sheets_processed": 0,
                "sheets_skipped_no_time_anchor": [],
                "unplayed_rubbers": [],
                "retired_rubbers": [],
                "tied_rubbers_undecided": [],
                "placeholder_dates_used": True,
                "notes": [
                    "Wilson team tournaments use a 10-pt match super-tiebreak; per-rubber "
                    "tiebreak winners are not always recoverable from the cells. Tied 1-1 "
                    "rubbers are flagged in tied_rubbers_undecided for the rating engine to skip.",
                    "Year is taken from the filename (2017 file's in-cell title says 2015).",
                    "Per-day dates use a placeholder schedule when the file does not record them.",
                ],
            }

            n_matches_inserted = 0
            for sheet_name in wb.sheet_names():
                # Only process Day-N, Semi Final, Final sheets.
                sn_lower = sheet_name.strip().lower()
                if not (sn_lower.startswith("day") or "semi" in sn_lower or sn_lower == "final" or sn_lower.startswith("final")):
                    continue

                anchor = _find_time_anchor(wb, sheet_name)
                if anchor is None:
                    quality_report["sheets_skipped_no_time_anchor"].append(sheet_name)
                    continue
                header_row, time_col = anchor

                # Try to extract a real date from any day-header cell in the
                # sheet (rows 1–8, cols time_col-1..time_col+1).
                real_date = None
                for r in range(1, header_row):
                    for c in range(max(1, time_col - 1), time_col + 3):
                        v = wb.cell(sheet_name, r, c)
                        d = _extract_real_date(v, year)
                        if d is not None:
                            real_date = d
                            break
                    if real_date is not None:
                        break
                played_on = real_date or _placeholder_date(season, year, sheet_name)
                round_label = _round_for_sheet(sheet_name)

                quality_report["n_sheets_processed"] += 1

                for row, rubber in _iter_rubbers(wb, sheet_name, time_col, header_row):
                    if rubber.get("_skip_unplayed"):
                        quality_report["unplayed_rubbers"].append({
                            "sheet": sheet_name,
                            "row": row,
                            "division": rubber["rubber"],
                            "pair_A": "/".join(rubber["pair_a"]),
                            "pair_B": "/".join(rubber["pair_b"]),
                        })
                        continue
                    division = rubber["rubber"] or "(unknown)"
                    mid = _insert_match(
                        db_conn,
                        tournament_id,
                        ingestion_run_id,
                        division,
                        round_label,
                        played_on,
                        rubber,
                        source_file_id,
                        quality_report,
                        sheet_name,
                        row,
                    )
                    if mid is not None:
                        n_matches_inserted += 1

            quality_report["n_matches_inserted"] = n_matches_inserted

            db_conn.execute(
                "UPDATE ingestion_runs "
                "SET status = 'completed', completed_at = datetime('now'), quality_report_jsonb = ? "
                "WHERE id = ?",
                (json.dumps(quality_report), ingestion_run_id),
            )
        finally:
            wb.close()

    return ingestion_run_id
