"""Parser for `TCK CHOSEN TOUNAMENT DIVISIONS 2024.xlsx` (T-P0-014, flat-list family).

Spec: `scripts/phase0/parser_spec_tck_chosen_2024.md`.

Distinguishing features vs the existing SE / Mixed parsers:
  - Layout is a FLAT LIST (not side-by-side blocks).
  - Each row is a single match with columns:
        DATE | TIME | COURT | DIV | TEAM | VS | TEAM | RESULTS [| SCRATCHED-OVERRIDE]
  - The score is a single string like `'7-5   6-1'` or `'6-4   5-7   10-4'`.
  - Walkovers: `W/O`, `W/0` (typo), `wo`, `WO`, `SCRATCHED` — flag
    `match.walkover = 1`.
  - SCRATCHED rows MAY carry a real score in col 9; if so, use the col-9
    score for set rows (winner determined from that score) but still flag
    walkover.
  - Date is per-row (Excel datetime in col 1).
  - Pair separator inside col-5/col-7 is `'/'` with optional whitespace
    (handled by `sports_experience_2025._split_pair`).

Tournament format on `tournaments.format`: `'doubles_division'` (same as
SE / Mixed — a divisional round-robin with fixed pairs). The flat-list
row layout doesn't change the schema mapping.

Public API:
    parse(xlsx_path: str, db_conn: sqlite3.Connection) -> int
        Returns the new ingestion_run_id.
"""

from __future__ import annotations

import datetime as _dt
import json
import os
import re
import sqlite3
import sys
from typing import Optional

import openpyxl

# Path-relative imports — same pattern as the other parsers.
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import players as players_mod  # noqa: E402

# Reuse helpers from the SE parser (sha256, split_pair, supersede).
from parsers import sports_experience_2025 as se  # noqa: E402

AGENT_VERSION = "phase0-tck-chosen-2024-parser-1.0"
TOURNAMENT_FORMAT = "doubles_division"
DEFAULT_CLUB_NAME = "VLTC"
DEFAULT_CLUB_SLUG = "vltc"
PLACEHOLDER_DATE_FALLBACK = "2024-01-01"

# Sheet → (division-string, gender) map. If a sheet not in this map shows up,
# we derive a best-effort division string from the sheet-name digit.
SHEET_DEFS: dict[str, dict] = {
    "MEN 1ST DIV": {"division": "Men Division 1", "gender": "M"},
    "MEN 2ND DIV": {"division": "Men Division 2", "gender": "M"},
    "MEN 3RD DIV": {"division": "Men Division 3", "gender": "M"},
    "LDYS 1ST DIV": {"division": "Ladies Division 1", "gender": "F"},
    "LDYS 2ND DIV": {"division": "Ladies Division 2", "gender": "F"},
}

# Walkover token recognition (case-insensitive, after .strip())
_WALKOVER_TOKENS = {"w/o", "w/0", "wo", "scratched"}

# Column anchors (1-indexed)
COL_DATE = 1
COL_TEAM_A = 5
COL_TEAM_B = 7
COL_RESULTS = 8
COL_RESULTS_OVERRIDE = 9  # for SCRATCHED rows that ALSO have a recorded score

# Cell value used to find the header row (col 8)
HEADER_TOKEN = "RESULTS"


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _is_walkover_token(value) -> bool:
    if not isinstance(value, str):
        return False
    return value.strip().lower() in _WALKOVER_TOKENS


def _coerce_date(value) -> Optional[str]:
    """Coerce a cell value to an ISO 8601 YYYY-MM-DD date string.

    Accepts datetime/date objects directly. Returns None for blanks or
    unparseable inputs.
    """
    if value is None:
        return None
    if isinstance(value, _dt.datetime):
        return value.date().isoformat()
    if isinstance(value, _dt.date):
        return value.isoformat()
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        # Best-effort: try YYYY-MM-DD parse.
        try:
            return _dt.date.fromisoformat(s).isoformat()
        except ValueError:
            return None
    return None


def _looks_like_pair(value) -> bool:
    """A pair string is a non-empty string containing at least one '/'."""
    if not isinstance(value, str):
        return False
    s = value.strip()
    return bool(s) and "/" in s


def _parse_score_string(score_str: str) -> Optional[list[tuple[int, int]]]:
    """Parse a score string like `'7-5   6-1'` into list of (a_games, b_games).

    Tolerant to:
      - any amount of whitespace between sets
      - leading/trailing whitespace overall
      - tab/newline as separator (just calls .split())

    Returns None if the string contains nothing usable. Tokens that don't
    match `<int>-<int>` are skipped (with stderr log).
    """
    if not isinstance(score_str, str):
        return None
    s = score_str.strip()
    if not s:
        return None
    sets: list[tuple[int, int]] = []
    for tok in s.split():
        m = re.fullmatch(r"(\d+)-(\d+)", tok)
        if not m:
            print(
                f"[tck_chosen_2024] skipping unparseable score token "
                f"{tok!r} in score string {score_str!r}",
                file=sys.stderr,
            )
            continue
        sets.append((int(m.group(1)), int(m.group(2))))
    return sets if sets else None


def _is_set_tiebreak(set_number: int, ga: int, gb: int) -> int:
    """Return 1 if this set should be flagged `was_tiebreak`, else 0.

    Rules:
      - 7-X or X-7 normal set with a tiebreak game (e.g. 7-6, 6-7) → flag.
        (Mirrors the SE-parser convention that any 7 in a 1st/2nd-set
        score implies a tiebreak.)
      - 3rd set whose max value >= 9 — assume 10-point match super-TB.
    """
    if set_number >= 3 and max(ga, gb) >= 9:
        return 1
    if set_number <= 2 and (ga == 7 or gb == 7):
        return 1
    return 0


def _is_super_tb(set_number: int, ga: int, gb: int) -> bool:
    """Was this set a 10-point match super-tiebreak (i.e. NOT a normal set)?"""
    return set_number >= 3 and max(ga, gb) >= 9


def _find_header_row(ws) -> Optional[int]:
    """Locate the row whose col 8 holds the literal `RESULTS` header.

    Searches rows 1..30 (defensive — observed positions are 13 or 15).
    """
    for r in range(1, 31):
        v = ws.cell(r, COL_RESULTS).value
        if isinstance(v, str) and v.strip().upper() == HEADER_TOKEN:
            return r
    return None


def _derive_division_from_sheet(sheet_name: str) -> tuple[str, Optional[str]]:
    """Best-effort division/gender derivation for unknown sheet names.

    Returns (division_string, gender) where gender is 'M' / 'F' / None.
    """
    sn = sheet_name.upper()
    digit_match = re.search(r"(\d+)", sn)
    div_n = digit_match.group(1) if digit_match else "?"
    if sn.startswith("MEN"):
        return f"Men Division {div_n}", "M"
    if sn.startswith("LDY") or sn.startswith("LAD"):
        return f"Ladies Division {div_n}", "F"
    return f"Division {div_n}", None


def _read_tournament_name(wb) -> Optional[str]:
    """Read the tournament title from cell [2,1] of the first sheet that has it."""
    for sn in wb.sheetnames:
        ws = wb[sn]
        v = ws.cell(2, 1).value
        if isinstance(v, str) and v.strip():
            return v.strip().replace("\xa0", " ")
    return None


def _extract_year_from_filename(filename: str) -> int:
    matches = re.findall(r"\b(20\d{2}|19\d{2})\b", filename)
    return int(matches[-1]) if matches else 2024


# ─────────────────────────────────────────────────────────────────────────────
# Match → DB row insertion
# ─────────────────────────────────────────────────────────────────────────────

def _insert_match(
    conn: sqlite3.Connection,
    tournament_id: int,
    ingestion_run_id: int,
    division: str,
    played_on: str,
    pair_a: str,
    pair_b: str,
    set_scores: list[tuple[int, int]],
    walkover: bool,
    source_file_id: int,
    gender: Optional[str],
) -> int:
    """Insert a match + its sides + per-set scores. Returns the match_id.

    `set_scores` is a list of (a_games, b_games) per set, in order.
    For walkovers with no recorded score, pass [(6, 0)].
    """
    cur = conn.execute(
        "INSERT INTO matches (tournament_id, played_on, match_type, division, "
        "round, ingestion_run_id, walkover) "
        "VALUES (?, ?, 'doubles', ?, NULL, ?, ?)",
        (tournament_id, played_on, division, ingestion_run_id, 1 if walkover else 0),
    )
    match_id = cur.lastrowid

    a1_raw, a2_raw = se._split_pair(pair_a)
    b1_raw, b2_raw = se._split_pair(pair_b)
    a1_id = players_mod.get_or_create_player(conn, a1_raw, source_file_id)
    a2_id = players_mod.get_or_create_player(conn, a2_raw, source_file_id)
    b1_id = players_mod.get_or_create_player(conn, b1_raw, source_file_id)
    b2_id = players_mod.get_or_create_player(conn, b2_raw, source_file_id)

    if gender is not None:
        for pid in (a1_id, a2_id, b1_id, b2_id):
            conn.execute(
                "UPDATE players SET gender = ? WHERE id = ? AND gender IS NULL",
                (gender, pid),
            )

    # Insert per-set scores, computing aggregates as we go.
    sets_won_a = sets_won_b = games_won_a = games_won_b = 0
    has_super_tb = False
    super_tb_a = super_tb_b = 0
    for idx, (ga, gb) in enumerate(set_scores, start=1):
        was_tb = _is_set_tiebreak(idx, ga, gb)
        conn.execute(
            "INSERT INTO match_set_scores (match_id, set_number, side_a_games, "
            "side_b_games, was_tiebreak) VALUES (?, ?, ?, ?, ?)",
            (match_id, idx, ga, gb, was_tb),
        )
        if _is_super_tb(idx, ga, gb):
            has_super_tb = True
            super_tb_a, super_tb_b = ga, gb
        else:
            games_won_a += ga
            games_won_b += gb
            if ga > gb:
                sets_won_a += 1
            elif gb > ga:
                sets_won_b += 1

    # Decide winner. Regular sets resolve it normally; if tied at 1-1 in
    # regular sets and a super-TB is present, the super-TB decides.
    if sets_won_a > sets_won_b:
        won_a, won_b = 1, 0
    elif sets_won_b > sets_won_a:
        won_a, won_b = 0, 1
    elif has_super_tb:
        if super_tb_a > super_tb_b:
            won_a, won_b = 1, 0
        elif super_tb_b > super_tb_a:
            won_a, won_b = 0, 1
        else:
            won_a, won_b = 0, 0
    else:
        # Walkover or pathological — default to side A win for walkovers
        # (the listed first team gets the W/O).
        if walkover:
            won_a, won_b = 1, 0
        else:
            won_a, won_b = 0, 0

    conn.execute(
        "INSERT INTO match_sides (match_id, side, player1_id, player2_id, "
        "sets_won, games_won, won) VALUES (?, 'A', ?, ?, ?, ?, ?)",
        (match_id, a1_id, a2_id, sets_won_a, games_won_a, won_a),
    )
    conn.execute(
        "INSERT INTO match_sides (match_id, side, player1_id, player2_id, "
        "sets_won, games_won, won) VALUES (?, 'B', ?, ?, ?, ?, ?)",
        (match_id, b1_id, b2_id, sets_won_b, games_won_b, won_b),
    )

    return match_id


# ─────────────────────────────────────────────────────────────────────────────
# Per-row extraction
# ─────────────────────────────────────────────────────────────────────────────

def _process_match_row(
    ws,
    row: int,
    sheet_name: str,
    division: str,
    gender: Optional[str],
    conn: sqlite3.Connection,
    tournament_id: int,
    ingestion_run_id: int,
    source_file_id: int,
    placeholder_date: str,
) -> tuple[bool, Optional[str]]:
    """Try to parse + insert one match from `ws.row(row)`.

    Returns (inserted_bool, skip_reason_or_None).
    """
    pair_a_v = ws.cell(row, COL_TEAM_A).value
    pair_b_v = ws.cell(row, COL_TEAM_B).value
    date_v = ws.cell(row, COL_DATE).value
    result_v = ws.cell(row, COL_RESULTS).value
    result_v9 = ws.cell(row, COL_RESULTS_OVERRIDE).value

    # If both team cells are blank, this is a spacer row — skip silently.
    if not _looks_like_pair(pair_a_v) and not _looks_like_pair(pair_b_v):
        return False, None

    # Need both sides to have a pair-like string.
    if not (_looks_like_pair(pair_a_v) and _looks_like_pair(pair_b_v)):
        print(
            f"[tck_chosen_2024] skipping {sheet_name!r} row {row}: missing pair "
            f"(team_a={pair_a_v!r} team_b={pair_b_v!r})",
            file=sys.stderr,
        )
        return False, "missing pair"

    pair_a = pair_a_v.strip()
    pair_b = pair_b_v.strip()

    played_on = _coerce_date(date_v) or placeholder_date

    # Determine score + walkover flag.
    is_walkover = False
    set_scores: Optional[list[tuple[int, int]]] = None

    if _is_walkover_token(result_v):
        is_walkover = True
        # SCRATCHED with col-9 score override → use that score for sets.
        if result_v9 is not None:
            set_scores = _parse_score_string(result_v9)
        if set_scores is None:
            # Walkover with no recorded score: use 6-0 placeholder.
            set_scores = [(6, 0)]
    elif isinstance(result_v, str) and result_v.strip():
        set_scores = _parse_score_string(result_v)
        if set_scores is None:
            # Result string present but unparseable.
            print(
                f"[tck_chosen_2024] skipping {sheet_name!r} row {row}: "
                f"unparseable score {result_v!r}",
                file=sys.stderr,
            )
            return False, "unparseable score"
    else:
        # No result recorded — incomplete match. Log + skip.
        print(
            f"[tck_chosen_2024] skipping {sheet_name!r} row {row}: no result "
            f"(team_a={pair_a!r} team_b={pair_b!r})",
            file=sys.stderr,
        )
        return False, "no result"

    try:
        _insert_match(
            conn,
            tournament_id,
            ingestion_run_id,
            division,
            played_on,
            pair_a,
            pair_b,
            set_scores,
            is_walkover,
            source_file_id,
            gender,
        )
    except ValueError as exc:
        # _split_pair fail (more than one '/' or zero) — log + skip.
        print(
            f"[tck_chosen_2024] skipping {sheet_name!r} row {row}: {exc}",
            file=sys.stderr,
        )
        return False, str(exc)
    return True, None


# ─────────────────────────────────────────────────────────────────────────────
# Top-level parse
# ─────────────────────────────────────────────────────────────────────────────

def _ensure_default_club(conn: sqlite3.Connection) -> int:
    row = conn.execute("SELECT id FROM clubs WHERE slug = ?", (DEFAULT_CLUB_SLUG,)).fetchone()
    if row is not None:
        return row[0]
    cur = conn.execute(
        "INSERT INTO clubs (name, slug) VALUES (?, ?)",
        (DEFAULT_CLUB_NAME, DEFAULT_CLUB_SLUG),
    )
    return cur.lastrowid


def parse(xlsx_path: str, db_conn: sqlite3.Connection) -> int:
    """Parse the TCK Chosen Tournament Divisions 2024 file into the DB.

    Returns the new ingestion_run_id.
    """
    sha256 = se._sha256_of_file(xlsx_path)
    filename = os.path.basename(xlsx_path)
    year = _extract_year_from_filename(filename)
    placeholder_date = f"{year}-01-01"

    with db_conn:
        club_id = _ensure_default_club(db_conn)

        # source_files row — reuse by (filename, sha256) when present; fall back
        # to sha256 alone so a scraper-renamed-but-identical file dedups
        # correctly (otherwise two source_files rows → two tournaments rows →
        # duplicate active matches; tournament-duplication bug, fixed Apr 2026).
        existing = db_conn.execute(
            "SELECT id FROM source_files WHERE original_filename = ? AND sha256 = ? "
            "ORDER BY id DESC LIMIT 1",
            (filename, sha256),
        ).fetchone()
        if existing is None:
            existing = db_conn.execute(
                "SELECT id FROM source_files WHERE sha256 = ? "
                "ORDER BY id DESC LIMIT 1",
                (sha256,),
            ).fetchone()
        if existing is not None:
            source_file_id = existing[0]
        else:
            cur = db_conn.execute(
                "INSERT INTO source_files (club_id, original_filename, sha256) VALUES (?, ?, ?)",
                (club_id, filename, sha256),
            )
            source_file_id = cur.lastrowid

        # ingestion_runs row
        cur = db_conn.execute(
            "INSERT INTO ingestion_runs (source_file_id, status, agent_version) "
            "VALUES (?, 'running', ?)",
            (source_file_id, AGENT_VERSION),
        )
        ingestion_run_id = cur.lastrowid

        # Supersede prior runs for this file
        prior_run_id = se._supersede_prior_runs_for_file(db_conn, source_file_id, ingestion_run_id)
        if prior_run_id is not None:
            db_conn.execute(
                "UPDATE ingestion_runs SET supersedes_run_id = ? WHERE id = ?",
                (prior_run_id, ingestion_run_id),
            )

        # Open workbook (full mode for random cell access).
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        try:
            tournament_name = _read_tournament_name(wb) or filename.rsplit(".", 1)[0]

            # Get-or-create on (club_id, name, year) — prevents two scraper
            # ingests of the same physical tournament under different filenames
            # producing two tournament rows with overlapping active matches.
            existing_t = db_conn.execute(
                "SELECT id FROM tournaments WHERE club_id = ? AND name = ? AND year = ? "
                "ORDER BY id LIMIT 1",
                (club_id, tournament_name, year),
            ).fetchone()
            if existing_t is not None:
                tournament_id = existing_t[0]
            else:
                cur = db_conn.execute(
                    "INSERT INTO tournaments (club_id, name, year, format, source_file_id) "
                    "VALUES (?, ?, ?, ?, ?)",
                    (club_id, tournament_name, year, TOURNAMENT_FORMAT, source_file_id),
                )
                tournament_id = cur.lastrowid

            n_matches_inserted = 0
            n_walkovers = 0
            skipped: list[dict] = []

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # Look up division/gender; fall back to derivation.
                if sheet_name in SHEET_DEFS:
                    division = SHEET_DEFS[sheet_name]["division"]
                    gender = SHEET_DEFS[sheet_name]["gender"]
                else:
                    division, gender = _derive_division_from_sheet(sheet_name)
                    print(
                        f"[tck_chosen_2024] unknown sheet {sheet_name!r} — "
                        f"derived division={division!r} gender={gender!r}",
                        file=sys.stderr,
                    )

                header_row = _find_header_row(ws)
                if header_row is None:
                    print(
                        f"[tck_chosen_2024] sheet {sheet_name!r}: no RESULTS "
                        f"header row found in first 30 rows — skipping sheet",
                        file=sys.stderr,
                    )
                    continue

                # Iterate rows from header_row + 2 (skip the spacer line below header)
                # downward. Stop once we've seen too many consecutive blank rows.
                first_data_row = header_row + 2
                max_r = ws.max_row or first_data_row
                blank_streak = 0
                BLANK_TOLERANCE = 30  # rows of blank space before we conclude end

                for r in range(first_data_row, max_r + 1):
                    pair_a_v = ws.cell(r, COL_TEAM_A).value
                    pair_b_v = ws.cell(r, COL_TEAM_B).value
                    date_v = ws.cell(r, COL_DATE).value
                    result_v = ws.cell(r, COL_RESULTS).value

                    # If row is fully empty in the data columns, count it as blank.
                    if (
                        not _looks_like_pair(pair_a_v)
                        and not _looks_like_pair(pair_b_v)
                        and date_v is None
                        and (result_v is None or (isinstance(result_v, str) and not result_v.strip()))
                    ):
                        blank_streak += 1
                        if blank_streak >= BLANK_TOLERANCE:
                            break
                        continue
                    blank_streak = 0

                    # Track walkovers BEFORE inserting (the inserter consumes the row).
                    is_walkover_row = _is_walkover_token(result_v)
                    inserted, reason = _process_match_row(
                        ws,
                        r,
                        sheet_name,
                        division,
                        gender,
                        db_conn,
                        tournament_id,
                        ingestion_run_id,
                        source_file_id,
                        placeholder_date,
                    )
                    if inserted:
                        n_matches_inserted += 1
                        if is_walkover_row:
                            n_walkovers += 1
                    elif reason is not None:
                        skipped.append({
                            "sheet": sheet_name,
                            "row": r,
                            "reason": reason,
                            "team_a": str(pair_a_v).strip() if pair_a_v else None,
                            "team_b": str(pair_b_v).strip() if pair_b_v else None,
                            "result": str(result_v).strip() if isinstance(result_v, str) else None,
                        })

            quality_report = {
                "n_matches_inserted": n_matches_inserted,
                "n_walkovers": n_walkovers,
                "skipped_rows": skipped,
                "notes": [
                    "Flat-list format — each row = one match.",
                    "Walkover tokens (W/O, W/0, wo, SCRATCHED) → walkover=1 + 6-0 placeholder.",
                    "SCRATCHED with col-9 score uses the col-9 score for set rows.",
                    "Default walkover winner = side A (the listed first team) — "
                    "no annotation in source file says which side took the W/O.",
                ],
            }

            db_conn.execute(
                "UPDATE ingestion_runs "
                "SET status = 'completed', completed_at = datetime('now'), "
                "quality_report_jsonb = ? WHERE id = ?",
                (json.dumps(quality_report), ingestion_run_id),
            )
        finally:
            wb.close()

    return ingestion_run_id
