"""Extract captain-assigned class labels (A1, A2, ..., D3) from team-tournament
"Team Selection" sheets.

Used by team-tournament parsers (modern + legacy) to populate
`player_team_assignments`. The class label is a captain-assigned slot
designation, not derived from results — this module just reads what's already
in the source spreadsheet.

Layout (modern Antes / Tennis Trade / San Michel files):
    Row 2:  | (col 0) | (col 1) | A | B | C | D | E | F        ← team letters
    Row 3:  | (col 0) | CAPTAIN | <name1> | <name2> | ...      ← captains per team
    Row 5:  | (col 0) | MEN     | (blank)                       ← gender header
    Row 6:  | (col 0) | A1      | <player A1, team A> | <A1 team B> | ...
    Row 7:  | (col 0) | A2      | <player A2, team A> | ...
    Row 8:  | (col 0) | A3      | ...
    Row 10+: B1, B2, B3, C1..C3, D1..D3
    Row 22+: LADIES section: A1..A3, B1..B3, C1..C3 (no D for ladies usually)
"""

from __future__ import annotations

import re
import sqlite3
from typing import Iterator

import openpyxl

# Class label pattern: letter (A-D) + digit
_CLASS_RE = re.compile(r"^([A-D])(\d+)$", re.IGNORECASE)


def _norm(cell) -> str:
    """Normalize a cell value to a stripped string (or empty)."""
    if cell is None:
        return ""
    return str(cell).strip()


def _is_class_label(value: str) -> tuple[str, int] | None:
    """Return (tier_letter, slot_number) if value is a class label like 'A1'."""
    if not value:
        return None
    m = _CLASS_RE.match(value.strip().upper())
    if not m:
        return None
    return m.group(1).upper(), int(m.group(2))


def extract_team_selection(
    xlsx_path: str,
    sheet_name: str = "Team Selection",
) -> list[dict]:
    """Parse the Team Selection sheet and return a list of assignment records.

    Each record:
        {
            "team_letter": "A",            # or "B", "C", ...
            "captain_name": "Kelsey",      # raw from source
            "class_label": "A1",
            "tier_letter": "A",
            "slot_number": 1,
            "gender": "M",                 # 'M' for men section, 'F' for ladies
            "player_name": "Clayton Zammit Cesare",
        }

    If the sheet doesn't exist (some legacy files omit it), returns [].
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]

    # Read all rows once into a list of lists (small sheet; ~30 rows)
    rows = [
        [_norm(c) for c in row]
        for row in ws.iter_rows(values_only=True)
    ]

    # Find the team-letter header row. Two layout variants observed:
    #   Antes 2025:        'A', 'B', 'C', 'D', 'E', 'F'           (bare letters)
    #   San Michel 2026:   'TEAM A', 'TEAM B', 'TEAM C', ...      (with prefix)
    team_header_idx = None
    team_letter_cols: dict[str, int] = {}
    for i, row in enumerate(rows[:10]):
        positions: dict[str, int] = {}
        for col, cell in enumerate(row):
            if not cell:
                continue
            up = cell.upper().strip()
            # Bare single letter
            if len(up) == 1 and up in "ABCDEF":
                positions[up] = col
            # "TEAM X" form
            elif up.startswith("TEAM ") and len(up) == 6 and up[-1] in "ABCDEF":
                positions[up[-1]] = col
        if len(positions) >= 3:  # at least 3 teams to be plausible
            team_header_idx = i
            team_letter_cols = positions
            break

    if not team_letter_cols:
        return []

    # Find captain row (next non-empty row with 'CAPTAIN' in some col)
    captain_names: dict[str, str] = {}
    for i in range(team_header_idx + 1, min(team_header_idx + 4, len(rows))):
        row = rows[i]
        if any(c.upper() == "CAPTAIN" for c in row):
            for letter, col in team_letter_cols.items():
                if col < len(row):
                    captain_names[letter] = row[col]
            break

    # Walk subsequent rows looking for class labels in any column;
    # players are in the team-letter columns of the same row.
    assignments: list[dict] = []
    current_gender: str | None = None  # 'M' / 'F' tracked by section headers

    for row in rows[team_header_idx + 1:]:
        # Detect gender section headers: a cell containing 'MEN' or 'LADIES'
        for cell in row:
            up = cell.upper()
            if up == "MEN":
                current_gender = "M"
            elif up in ("LADIES", "WOMEN"):
                current_gender = "F"

        # Find the class label in this row (typically col 1 or 2)
        class_info = None
        for col_idx, cell in enumerate(row):
            ci = _is_class_label(cell)
            if ci:
                class_info = (col_idx, ci[0], ci[1])
                break
        if class_info is None:
            continue
        _, tier_letter, slot_number = class_info
        class_label = f"{tier_letter}{slot_number}"

        # For each team letter, the player's name is in that team's column
        for letter, col in team_letter_cols.items():
            if col >= len(row):
                continue
            player_name = row[col]
            if not player_name:
                continue
            # Skip if the cell IS the class label (happens when team-letter
            # column accidentally aligns with class-label column on small sheets)
            if _is_class_label(player_name):
                continue
            assignments.append({
                "team_letter": letter,
                "captain_name": captain_names.get(letter),
                "class_label": class_label,
                "tier_letter": tier_letter,
                "slot_number": slot_number,
                "gender": current_gender,
                "player_name": player_name,
            })

    return assignments


def store_team_selection(
    db_conn: sqlite3.Connection,
    tournament_id: int,
    source_file_id: int,
    assignments: list[dict],
    *,
    get_or_create_player_fn,
) -> int:
    """Insert assignments into player_team_assignments. Returns count inserted.

    `get_or_create_player_fn` is `players.get_or_create_player` (passed in to
    avoid a circular import). It resolves each raw player name to a player_id.
    """
    n = 0
    for a in assignments:
        player_id = get_or_create_player_fn(
            db_conn, a["player_name"], source_file_id
        )
        # Update gender on the player record if we now know it (and didn't before)
        if a.get("gender"):
            db_conn.execute(
                "UPDATE players SET gender = ? WHERE id = ? AND gender IS NULL",
                (a["gender"], player_id),
            )
        # INSERT OR REPLACE since (tournament_id, player_id) is unique;
        # if a player appears twice in the same tournament's selection (shouldn't
        # but defensive), the last one wins.
        db_conn.execute(
            """
            INSERT OR REPLACE INTO player_team_assignments
                (tournament_id, player_id, team_letter, captain_name,
                 class_label, tier_letter, slot_number, gender)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                tournament_id, player_id, a.get("team_letter"),
                a.get("captain_name"), a["class_label"],
                a["tier_letter"], a["slot_number"], a.get("gender"),
            ),
        )
        n += 1
    return n


# ---- Helpers for downstream use (CLI rank command) ----


def player_current_class(
    db_conn: sqlite3.Connection,
    player_id: int,
) -> tuple[str, str, int] | None:
    """Return the player's most-recent (class_label, tier_letter, slot_number).

    Most-recent = highest tournament_id (proxy for chronological order; works
    because tournaments are inserted in load order which is roughly chronological).
    Returns None if no team assignment exists.
    """
    row = db_conn.execute(
        """
        SELECT pta.class_label, pta.tier_letter, pta.slot_number
        FROM player_team_assignments pta
        JOIN tournaments t ON t.id = pta.tournament_id
        WHERE pta.player_id = ?
        ORDER BY t.year DESC, t.id DESC
        LIMIT 1
        """,
        (player_id,),
    ).fetchone()
    return row if row else None


def player_class_history(
    db_conn: sqlite3.Connection,
    player_id: int,
) -> list[dict]:
    """Return the player's full class-assignment history, oldest first."""
    rows = db_conn.execute(
        """
        SELECT t.year, t.name, pta.team_letter, pta.captain_name, pta.class_label
        FROM player_team_assignments pta
        JOIN tournaments t ON t.id = pta.tournament_id
        WHERE pta.player_id = ?
        ORDER BY t.year, t.id
        """,
        (player_id,),
    ).fetchall()
    return [
        {"year": r[0], "tournament": r[1], "team": r[2],
         "captain": r[3], "class": r[4]}
        for r in rows
    ]
